#!/usr/bin/env python3
"""Download and normalize budgets of regional governments.

Regional governments are a separate fact grain from municipalities.  The
shared ``public_entities`` dimension supplies stable identifiers, while
``regional_governments`` describes the jurisdiction tier and
``regional_budget_line_facts`` stores the money.  National classifications are
preserved; cross-country mappings remain a reviewed, versioned layer.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import io
import json
import os
import shutil
import tempfile
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Iterator

import requests
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from prepare_international_municipal_data import (
    bigquery_numeric_compatible,
    iter_dbf,
    jsonstat_rows,
    numeric_json,
    reported_amount,
    sha256,
    stable_code,
)


ROOT = Path(os.environ.get("CZBUDGET_WORKSPACE_ROOT", Path(__file__).resolve().parents[3]))
DEFAULT_CONFIG = ROOT / "website/pipeline/config/international_regional_sources.v1.json"
DEFAULT_CACHE = ROOT / "data/source_cache/international_regional"
DEFAULT_OUTPUT = ROOT / "outputs/international-regional"
USER_AGENT = "czbudget-regional-pipeline/1.0 (+https://publicspendingdata.org)"

TABLE_FILES = (
    "public_entities",
    "public_entity_sources",
    "classification_versions",
    "budget_nodes",
    "raw_budget_lines",
    "regional_governments",
    "regional_budget_line_facts",
    "regional_budget_coverage",
    "ingestion_runs",
)


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


class Bundle:
    def __init__(self, output: Path, raw_mode: str, raw_limit: int, gzip_output: bool):
        output.mkdir(parents=True, exist_ok=True)
        self.output = output
        self.raw_mode = raw_mode
        self.raw_limit = raw_limit
        self.counts: Counter[str] = Counter()
        self.handles: dict[str, Any] = {}
        self.entities: set[str] = set()
        self.regions: set[str] = set()
        self.classifications: set[str] = set()
        self.nodes: set[str] = set()
        self.sources: set[tuple[str, str | None]] = set()
        for table in TABLE_FILES:
            path = output / f"{table}.jsonl{'.gz' if gzip_output else ''}"
            self.handles[table] = gzip.open(path, "wt", encoding="utf-8") if gzip_output else path.open("w", encoding="utf-8")

    def write(self, table: str, row: dict[str, Any]) -> None:
        self.handles[table].write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")
        self.counts[table] += 1

    def raw(self, row: dict[str, Any]) -> None:
        if self.raw_mode == "none":
            return
        if self.raw_mode == "sample" and self.counts["raw_budget_lines"] >= self.raw_limit:
            return
        self.write("raw_budget_lines", row)

    def entity(self, row: dict[str, Any]) -> bool:
        key = row["public_entity_id"]
        if key in self.entities:
            return False
        self.entities.add(key)
        self.write("public_entities", row)
        return True

    def region(self, row: dict[str, Any]) -> bool:
        key = row["regional_government_id"]
        if key in self.regions:
            return False
        self.regions.add(key)
        self.write("regional_governments", row)
        return True

    def classification(self, row: dict[str, Any]) -> None:
        key = row["classification_id"]
        if key not in self.classifications:
            self.classifications.add(key)
            self.write("classification_versions", row)

    def node(self, row: dict[str, Any]) -> None:
        key = row["budget_node_id"]
        if key not in self.nodes:
            self.nodes.add(key)
            self.write("budget_nodes", row)

    def source(self, row: dict[str, Any]) -> None:
        key = (row["source_id"], row.get("public_entity_id"))
        if key not in self.sources:
            self.sources.add(key)
            self.write("public_entity_sources", row)

    def close(self) -> None:
        for handle in self.handles.values():
            handle.close()


class Context:
    def __init__(self, args: argparse.Namespace, config: dict[str, Any], bundle: Bundle):
        self.args = args
        self.config = config
        self.bundle = bundle
        self.loaded_at = utc_now()
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT, "Accept": "*/*"})
        retry = Retry(total=4, connect=4, read=4, backoff_factor=1.0, status_forcelist=(429, 500, 502, 503, 504), allowed_methods=("GET", "POST"))
        self.session.mount("https://", HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=10))

    def cache_path(self, country: str, source: dict[str, Any], year: int) -> Path:
        reused = source.get("reuse_path")
        if reused:
            candidate = ROOT / reused
            if candidate.exists() and not self.args.refresh:
                return candidate
        return self.args.cache_dir / country / source["filename"].format(year=year)

    def download(self, country: str, source: dict[str, Any], year: int) -> Path:
        path = self.cache_path(country, source, year)
        if path.exists() and not self.args.refresh:
            return path
        if self.args.offline:
            raise FileNotFoundError(f"Offline source missing: {path}")
        path.parent.mkdir(parents=True, exist_ok=True)
        response = self.session.get(source["url"], timeout=180, stream=True)
        response.raise_for_status()
        temporary = path.with_suffix(path.suffix + ".part")
        with temporary.open("wb") as handle:
            for chunk in response.iter_content(1024 * 1024):
                if chunk:
                    handle.write(chunk)
        temporary.replace(path)
        return path

    def source_row(self, country: str, source: dict[str, Any], year: int, path: Path, notes: str) -> None:
        archive_file = str(path.relative_to(ROOT)) if path.is_relative_to(ROOT) else str(path)
        self.bundle.source({
            "source_id": source["id"],
            "public_entity_id": None,
            "source_type": "regional_budget_detail",
            "source_name": source.get("table") or source["id"],
            "source_url": source["url"],
            "dataset_code": source.get("table"),
            "archive_file": archive_file,
            "archive_sha256": sha256(path),
            "retrieved_at": self.loaded_at,
            "notes": notes,
            "loaded_at": self.loaded_at,
        })


def source_for_year(cfg: dict[str, Any], year: int) -> list[dict[str, Any]]:
    return [row for row in cfg["sources"] if row.get("year") in (None, year)]


def entity_row(country: str, cfg: dict[str, Any], code: str, name: str, code_type: str, loaded_at: str) -> dict[str, Any]:
    alpha2 = cfg["alpha2"]
    return {
        "public_entity_id": f"{alpha2}:{code}",
        "entity_name": name,
        "entity_type": "regional_government",
        "country_code_alpha2": alpha2,
        "country_code_alpha3": country,
        "national_entity_code": code,
        "national_entity_code_type": code_type,
        "is_eu_capital": False,
        "is_extra_city": False,
        "default_currency_code": cfg["currency"],
        "eurostat_city_code": None,
        "eurostat_geography_name": None,
        "administrative_region_code": code,
        "administrative_region_name": name,
        "administrative_district_code": None,
        "administrative_district_name": None,
        "national_geography_code": code,
        "national_geography_code_type": code_type,
        "valid_from": None,
        "valid_to": None,
        "loaded_at": loaded_at,
    }


def regional_row(country: str, cfg: dict[str, Any], code: str, name: str, code_type: str, tier: dict[str, Any], source_id: str, loaded_at: str) -> dict[str, Any]:
    public_entity_id = f"{cfg['alpha2']}:{code}"
    return {
        "regional_government_id": public_entity_id,
        "public_entity_id": public_entity_id,
        "country_code": country,
        "national_region_code": code,
        "national_region_code_type": code_type,
        "government_type_code": tier["tier_code"],
        "tier_level": tier["level"],
        "name_native": name,
        "name_en": name,
        "name_cs": name,
        "nuts_code": None,
        "parent_regional_government_id": None,
        "is_capital_region": False,
        "valid_from": None,
        "valid_to": None,
        "source_id": source_id,
        "loaded_at": loaded_at,
    }


def classification_row(country: str, classification_id: str, side: str, name: str, source_url: str, year: int, loaded_at: str) -> dict[str, Any]:
    return {
        "classification_id": classification_id,
        "country_code": country,
        "budget_side": side,
        "government_scope": "regional_government",
        "valid_from_year": year,
        "valid_to_year": year,
        "classification_name": name,
        "legal_basis": None,
        "source_url": source_url,
        "notes": "Source-native regional classification; no cross-country mapping is implied.",
        "loaded_at": loaded_at,
    }


def node_row(country: str, classification_id: str, side: str, code: str, name: str, year: int, loaded_at: str, summary: bool = False) -> dict[str, Any]:
    return {
        "budget_node_id": f"{country}:{classification_id}:{code}",
        "classification_id": classification_id,
        "country_code": country,
        "budget_side": side,
        "government_scope": "regional_government",
        "node_code": code,
        "node_name_native": name or code,
        "node_name_en": name or None,
        "node_name_cs": None,
        "parent_budget_node_id": None,
        "hierarchy_level": 1,
        "hierarchy_path": [],
        "is_chapter": summary,
        "effective_from_year": year,
        "effective_to_year": year,
        "loaded_at": loaded_at,
    }


def fact_row(
    entity_id: str,
    tier_code: str,
    year: int,
    stage: str,
    side: str,
    function_code: str | None,
    economic_code: str,
    amount: Decimal,
    currency: str,
    source: dict[str, Any],
    run_id: str,
    loaded_at: str,
    function_class: str | None,
    economic_class: str,
    **extra: Any,
) -> dict[str, Any]:
    country_code = {"FR": "FRA", "PL": "POL", "SE": "SWE", "DK": "DNK"}[entity_id.split(":", 1)[0]]
    return {
        "public_entity_id": entity_id,
        "country_code": country_code,
        "regional_tier_code": tier_code,
        "fiscal_year": year,
        "fiscal_period": extra.pop("period", "FY"),
        "reporting_scope": extra.pop("scope", "standalone_regional_government"),
        "budget_stage": stage,
        "budget_side": side,
        "source_budget_item_type_code": extra.pop("item_type", None),
        "functional_code": function_code,
        "economic_code": economic_code,
        "functional_classification_id": function_class,
        "economic_classification_id": economic_class,
        "amount_local": numeric_json(amount),
        "currency_code": currency,
        "amount_eur": None,
        "fx_date": None,
        "is_consolidation_item": extra.pop("consolidation", False),
        "is_financing": side == "financing",
        "is_summary_row": extra.pop("summary", False),
        "source_row_number": extra.pop("row_number", None),
        "source_sheet": extra.pop("sheet", None),
        "source_id": source["id"],
        "ingestion_run_id": run_id,
        "coverage_type": extra.pop("coverage_type", "census"),
        "is_imputed": False,
        "quality_flags": extra.pop("quality_flags", []),
        "loaded_at": loaded_at,
    }


def raw_row(country: str, year: int, source: dict[str, Any], run_id: str, row_number: int, sheet: str, payload: dict[str, Any], loaded_at: str) -> dict[str, Any]:
    return {
        "country_code": country,
        "fiscal_year": year,
        "source_id": source["id"],
        "ingestion_run_id": run_id,
        "source_row_number": row_number,
        "source_sheet": sheet,
        "source_payload": payload,
        "source_url": source["url"],
        "loaded_at": loaded_at,
    }


def ingestion_row(source: dict[str, Any], run_id: str, path: Path, started_at: str, rows_read: int, rows_loaded: int) -> dict[str, Any]:
    return {
        "ingestion_run_id": run_id,
        "source_id": source["id"],
        "started_at": started_at,
        "completed_at": utc_now(),
        "status": "completed",
        "source_vintage": source.get("year") and str(source["year"]),
        "source_sha256": sha256(path),
        "rows_read": rows_read,
        "rows_loaded": rows_loaded,
        "warning_count": 0,
        "error_message": None,
    }


def coverage_row(country: str, year: int, tier: dict[str, Any], source_ids: list[str], entity_count: int, fact_count: int, stages: set[str], sides: set[str], loaded_at: str, limitations: list[str] | None = None) -> dict[str, Any]:
    expected = tier.get("expected_count")
    status = "complete" if expected is not None and entity_count == expected else ("partial" if expected is not None else "source_count_unreviewed")
    return {
        "coverage_id": f"{country}:{tier['tier_code']}:{year}",
        "country_code": country,
        "fiscal_year": year,
        "regional_tier_code": tier["tier_code"],
        "source_ids": source_ids,
        "entity_expected_count": expected,
        "entity_source_count": entity_count,
        "entity_loaded_count": entity_count,
        "fact_count": fact_count,
        "budget_stages": sorted(stages),
        "budget_sides": sorted(sides),
        "coverage_type": "census",
        "validation_status": status,
        "limitations": limitations or [],
        "assessed_at": loaded_at,
    }


def french_csv_rows(path: Path) -> Iterator[tuple[int, str, dict[str, Any]]]:
    with zipfile.ZipFile(path) as archive:
        member = next(name for name in archive.namelist() if name.lower().endswith(".csv"))
        with archive.open(member) as raw:
            reader = csv.DictReader(io.TextIOWrapper(raw, encoding="cp1252", newline=""), delimiter=";")
            for row_number, row in enumerate(reader, 2):
                yield row_number, member, row


def run_france(ctx: Context, country: str, cfg: dict[str, Any], year: int) -> dict[str, Any]:
    source = source_for_year(cfg, year)[0]
    path = ctx.download(country, source, year)
    ctx.source_row(country, source, year, path, "DGFiP regional and departmental nature-function accounting balances.")
    tiers = {row["source_code"]: row for row in cfg["expected_tiers"]}
    function_class = f"FR_DGFIP_REGIONAL_FUNCTION_{year}"
    economic_class = f"FR_DGFIP_REGIONAL_ACCOUNT_{year}"
    ctx.bundle.classification(classification_row(country, function_class, "mixed", "DGFiP regional functional classification", source["url"], year, ctx.loaded_at))
    ctx.bundle.classification(classification_row(country, economic_class, "mixed", "DGFiP regional chart of accounts", source["url"], year, ctx.loaded_at))
    run_id = f"{source['id']}-v1"
    started_at = utc_now()
    entities_by_tier: dict[str, set[str]] = defaultdict(set)
    facts_by_tier: Counter[str] = Counter()
    rows_read = rows_loaded = 0
    for row_number, member, row in french_csv_rows(path):
        tier = tiers.get(stable_code(row.get("CATEG")))
        if not tier:
            continue
        code = stable_code(row.get("siren"))
        if not code:
            continue
        if ctx.args.max_entities and code not in entities_by_tier[tier["tier_code"]] and len(entities_by_tier[tier["tier_code"]]) >= ctx.args.max_entities:
            continue
        rows_read += 1
        name = stable_code(row.get("LBUDG")) or code
        entity_id = f"FR:{code}"
        entities_by_tier[tier["tier_code"]].add(code)
        ctx.bundle.entity(entity_row(country, cfg, code, name, "FR_SIREN", ctx.loaded_at))
        ctx.bundle.region(regional_row(country, cfg, code, name, "FR_SIREN", tier, source["id"], ctx.loaded_at))
        ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, member, row, ctx.loaded_at))
        function = stable_code(row.get("FONCTION")) or "UNSPECIFIED"
        account = stable_code(row.get("COMPTE"))
        if not account:
            continue
        ctx.bundle.node(node_row(country, function_class, "mixed", function, function, year, ctx.loaded_at))
        ctx.bundle.node(node_row(country, economic_class, "mixed", account, account, year, ctx.loaded_at))
        scope = "main_budget" if stable_code(row.get("CBUDG")) == "1" else "supplementary_budget"
        flags = [f"nomenclature:{stable_code(row.get('NOMEN'))}", "accounting_balance_execution"]
        for side, column in (("expenditure", "OBNETDEB"), ("revenue", "OBNETCRE")):
            amount = reported_amount(row.get(column))
            if amount is None:
                continue
            ctx.bundle.write("regional_budget_line_facts", fact_row(
                entity_id, tier["tier_code"], year, "actual", side, function, account, amount,
                cfg["currency"], source, run_id, ctx.loaded_at, function_class, economic_class,
                scope=scope, row_number=row_number, sheet=member, quality_flags=flags,
            ))
            rows_loaded += 1
            facts_by_tier[tier["tier_code"]] += 1
    ctx.bundle.write("ingestion_runs", ingestion_row(source, run_id, path, started_at, rows_read, rows_loaded))
    for tier in cfg["expected_tiers"]:
        count = len(entities_by_tier[tier["tier_code"]])
        ctx.bundle.write("regional_budget_coverage", coverage_row(
            country, year, tier, [source["id"]], count, facts_by_tier[tier["tier_code"]], {"actual"}, {"revenue", "expenditure"}, ctx.loaded_at,
            ["Functional-balance publication covers only entities that report this presentation; supplementary budgets remain separate."],
        ))
    return {"entities": sum(map(len, entities_by_tier.values())), "facts": rows_loaded, "tiers": {key: len(value) for key, value in entities_by_tier.items()}}


def polish_dictionary(path: Path) -> dict[tuple[str, str, str, str, str], dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [stable_code(value).replace("\u200b", "") for value in next(rows)]
    result: dict[tuple[str, str, str, str, str], dict[str, str]] = {}
    for values in rows:
        row = dict(zip(headers, values))
        if stable_code(row.get("Typ Jst Nazwa")) != "Województwo":
            continue
        key = tuple(stable_code(row.get(code)).zfill(2 if code not in ("GT", "PT") else 1) for code in ("WK", "PK", "GK", "GT", "PT"))
        result[key] = {
            "name": stable_code(row.get("Nazwa")),
            "code": stable_code(row.get("Kod GUS")).zfill(7),
        }
    return result


def run_poland(ctx: Context, country: str, cfg: dict[str, Any], year: int) -> dict[str, Any]:
    sources = {row["kind"]: row for row in source_for_year(cfg, year)}
    paths = {kind: ctx.download(country, source, year) for kind, source in sources.items()}
    dictionary = polish_dictionary(paths["entities"])
    if ctx.args.max_entities:
        selected_keys = set(sorted(dictionary)[:ctx.args.max_entities])
        dictionary = {key: value for key, value in dictionary.items() if key in selected_keys}
    tier = cfg["expected_tiers"][0]
    for item in dictionary.values():
        ctx.bundle.entity(entity_row(country, cfg, item["code"], item["name"], "PL_TERYT", ctx.loaded_at))
        ctx.bundle.region(regional_row(country, cfg, item["code"], item["name"], "PL_TERYT", tier, sources["entities"]["id"], ctx.loaded_at))
    total_facts = 0
    stages: set[str] = set()
    for side in ("revenue", "expenditure"):
        source = sources[side]
        path = paths[side]
        ctx.source_row(country, source, year, path, "Polish Q4 JST return; voivodeship rows selected from the all-JST publication.")
        function_class = f"PL_REGIONAL_FUNCTION_{year}"
        economic_class = f"PL_REGIONAL_{side.upper()}_{year}"
        ctx.bundle.classification(classification_row(country, function_class, "mixed", "Polish regional division/chapter classification", source["url"], year, ctx.loaded_at))
        ctx.bundle.classification(classification_row(country, economic_class, side, f"Polish regional {side} paragraph classification", source["url"], year, ctx.loaded_at))
        run_id = f"{source['id']}-v1"
        started_at = utc_now()
        rows_read = rows_loaded = 0
        with zipfile.ZipFile(path) as archive:
            member = next(name for name in archive.namelist() if name.lower().endswith(".dbf"))
            with archive.open(member) as handle:
                for row_number, row in iter_dbf(handle):
                    key = tuple(stable_code(row.get(code)).zfill(2 if code not in ("GT", "PT") else 1) for code in ("WK", "PK", "GK", "GT", "PT"))
                    item = dictionary.get(key)
                    if not item:
                        continue
                    rows_read += 1
                    payload = {key: numeric_json(value) if isinstance(value, Decimal) else value for key, value in row.items()}
                    ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, member, payload, ctx.loaded_at))
                    function = f"{stable_code(row.get('DZIAL')).zfill(3)}.{stable_code(row.get('ROZDZIAL')).zfill(5)}"
                    economic = f"{stable_code(row.get('PAR')).zfill(3)}{stable_code(row.get('PAR4'))}"
                    ctx.bundle.node(node_row(country, function_class, "mixed", function, function, year, ctx.loaded_at))
                    ctx.bundle.node(node_row(country, economic_class, side, economic, economic, year, ctx.loaded_at))
                    for stage, column in (("revised", "R1"), ("actual", "R4")):
                        amount = reported_amount(row.get(column))
                        if amount is None:
                            continue
                        stages.add(stage)
                        ctx.bundle.write("regional_budget_line_facts", fact_row(
                            f"PL:{item['code']}", tier["tier_code"], year, stage, side, function, economic,
                            amount, cfg["currency"], source, run_id, ctx.loaded_at, function_class, economic_class,
                            row_number=row_number, sheet=member, quality_flags=["quarter_four_return"],
                        ))
                        rows_loaded += 1
        ctx.bundle.write("ingestion_runs", ingestion_row(source, run_id, path, started_at, rows_read, rows_loaded))
        total_facts += rows_loaded
    ctx.source_row(country, sources["entities"], year, paths["entities"], "Official JST dictionary; 16 voivodeships selected by source type.")
    ctx.bundle.write("regional_budget_coverage", coverage_row(
        country, year, tier, [row["id"] for row in sources.values()], len(dictionary), total_facts, stages, {"revenue", "expenditure"}, ctx.loaded_at,
        ["The Q4 return exposes revised and actual values, not the original enacted plan."],
    ))
    return {"entities": len(dictionary), "facts": total_facts, "tiers": {tier["tier_code"]: len(dictionary)}}


def fetch_scb(ctx: Context, country: str, source: dict[str, Any], year: int) -> tuple[Path, dict[str, Any], dict[str, Any]]:
    path = ctx.cache_path(country, source, year)
    metadata_url = source["url"]
    metadata = ctx.session.get(metadata_url, timeout=120).json()
    if path.exists() and not ctx.args.refresh:
        return path, metadata, json.loads(path.read_text(encoding="utf-8"))
    if ctx.args.offline:
        raise FileNotFoundError(f"Offline source missing: {path}")
    query = []
    for variable in metadata["variables"]:
        values = [str(year)] if variable.get("time") else variable["values"]
        query.append({"code": variable["code"], "selection": {"filter": "item", "values": values}})
    response = ctx.session.post(metadata_url, json={"query": query, "response": {"format": "json-stat2"}}, timeout=180)
    response.raise_for_status()
    data = response.json()
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".part")
    temporary.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    temporary.replace(path)
    return path, metadata, data


def run_sweden(ctx: Context, country: str, cfg: dict[str, Any], year: int) -> dict[str, Any]:
    tier = cfg["expected_tiers"][0]
    entities: set[str] = set()
    total_facts = 0
    source_ids: list[str] = []
    for source in source_for_year(cfg, year):
        path, metadata, data = fetch_scb(ctx, country, source, year)
        source_ids.append(source["id"])
        ctx.source_row(country, source, year, path, "SCB official region accounts; source values in SEK millions are converted to SEK.")
        region_var = next(item for item in metadata["variables"] if item["code"] == "Region")
        region_names = dict(zip(region_var["values"], region_var["valueTexts"]))
        selected_regions = [code for code in region_var["values"] if code != "00"]
        if ctx.args.max_entities:
            selected_regions = selected_regions[:ctx.args.max_entities]
        for code in selected_regions:
            entities.add(code)
            name = region_names[code]
            ctx.bundle.entity(entity_row(country, cfg, code, name, "SE_REGION_CODE", ctx.loaded_at))
            ctx.bundle.region(regional_row(country, cfg, code, name, "SE_REGION_CODE", tier, source["id"], ctx.loaded_at))
        run_id = f"{source['id']}-{year}-v1"
        started_at = utc_now()
        rows_read = rows_loaded = 0
        if source["kind"] == "actual_activity":
            line_var = next(item for item in metadata["variables"] if item["code"] == "Verksomrkom")
            line_names = dict(zip(line_var["values"], line_var["valueTexts"]))
            # SCB publishes overlapping totals and sub-activities in one cube.
            # The single-digit main activities plus service/comparability rows
            # form the non-overlapping top-level presentation used here.
            comparable_activity_codes = {
                code for code in line_var["values"]
                if (len(code) == 1 and code.isdigit()) or code in {"940-980", "Jamf"}
            }
            function_class = f"SE_REGIONAL_ACTIVITY_{year}"
            economic_class = f"SE_REGIONAL_ACTIVITY_MEASURE_{year}"
            ctx.bundle.classification(classification_row(country, function_class, "mixed", "SCB regional activity", source["url"], year, ctx.loaded_at))
            ctx.bundle.classification(classification_row(country, economic_class, "mixed", "SCB regional activity measure", source["url"], year, ctx.loaded_at))
            measures = {"000000A6": "expenditure", "000000A5": "revenue"}
            for row_number, row in enumerate(jsonstat_rows(data), 1):
                region = stable_code(row.get("Region"))
                if region not in entities:
                    continue
                rows_read += 1
                function = stable_code(row.get("Verksomrkom"))
                measure = stable_code(row.get("ContentsCode"))
                side = measures.get(measure)
                amount = reported_amount(row.get("value"))
                if side is None or amount is None or function not in comparable_activity_codes:
                    continue
                amount *= Decimal(1_000_000)
                ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, source["table"], row, ctx.loaded_at))
                ctx.bundle.node(node_row(country, function_class, "mixed", function, line_names.get(function, function), year, ctx.loaded_at))
                ctx.bundle.node(node_row(country, economic_class, side, measure, row.get("ContentsCode_label") or measure, year, ctx.loaded_at))
                ctx.bundle.write("regional_budget_line_facts", fact_row(
                    f"SE:{region}", tier["tier_code"], year, "actual", side, function, measure, amount,
                    cfg["currency"], source, run_id, ctx.loaded_at, function_class, economic_class,
                    row_number=row_number, sheet=source["table"], quality_flags=["source_unit_sek_millions"],
                ))
                rows_loaded += 1
        else:
            line_var = next(item for item in metadata["variables"] if item["code"] == "Resultatraknposter")
            line_names = dict(zip(line_var["values"], line_var["valueTexts"]))
            economic_class = f"SE_REGIONAL_INCOME_STATEMENT_{year}"
            ctx.bundle.classification(classification_row(country, economic_class, "mixed", "SCB regional income statement", source["url"], year, ctx.loaded_at))
            revenue_codes = {"010", "040", "054", "055", "060", "100"}
            expenditure_codes = {"022", "023", "025", "072", "110"}
            for row_number, row in enumerate(jsonstat_rows(data), 1):
                region = stable_code(row.get("Region"))
                if region not in entities:
                    continue
                rows_read += 1
                code = stable_code(row.get("Resultatraknposter"))
                side = "revenue" if code in revenue_codes else ("expenditure" if code in expenditure_codes else None)
                amount = reported_amount(row.get("value"))
                if side is None or amount is None:
                    continue
                amount = abs(amount) * Decimal(1_000_000)
                ctx.bundle.node(node_row(country, economic_class, side, code, line_names.get(code, code), year, ctx.loaded_at, summary=True))
                ctx.bundle.write("regional_budget_line_facts", fact_row(
                    f"SE:{region}", tier["tier_code"], year, "actual", side, None, code, amount,
                    cfg["currency"], source, run_id, ctx.loaded_at, None, economic_class,
                    summary=True, row_number=row_number, sheet=source["table"], quality_flags=["source_unit_sek_millions", "statement_summary_row"],
                ))
                rows_loaded += 1
        ctx.bundle.write("ingestion_runs", ingestion_row(source, run_id, path, started_at, rows_read, rows_loaded))
        total_facts += rows_loaded
    ctx.bundle.write("regional_budget_coverage", coverage_row(
        country, year, tier, source_ids, len(entities), total_facts, {"actual"}, {"revenue", "expenditure"}, ctx.loaded_at,
        ["Annual accounts are loaded; enacted regional budgets are not yet part of the central SCB source."],
    ))
    return {"entities": len(entities), "facts": total_facts, "tiers": {tier["tier_code"]: len(entities)}}


def fetch_denmark(ctx: Context, country: str, source: dict[str, Any], year: int) -> tuple[Path, dict[str, Any]]:
    path = ctx.cache_path(country, source, year)
    metadata_url = f"https://api.statbank.dk/v1/tableinfo/{source['table']}?lang=en"
    metadata_response = ctx.session.get(metadata_url, timeout=120)
    metadata_response.raise_for_status()
    metadata = metadata_response.json()
    if path.exists() and not ctx.args.refresh:
        return path, metadata
    if ctx.args.offline:
        raise FileNotFoundError(f"Offline source missing: {path}")
    region_var = metadata["variables"][0]
    regions = [item["id"] for item in region_var["values"] if item["id"] != "000"]
    if ctx.args.max_entities:
        regions = regions[:ctx.args.max_entities]
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".part")
    wrote_header = False
    with temporary.open("w", encoding="utf-8", newline="") as output:
        for region in regions:
            query = {"table": source["table"], "format": "BULK", "lang": "en", "valuePresentation": "Code", "variables": []}
            for variable in metadata["variables"]:
                code = variable["id"]
                values = [region] if code == region_var["id"] else ([str(year)] if variable.get("time") else [item["id"] for item in variable["values"]])
                query["variables"].append({"code": code, "values": values})
            response = ctx.session.post(source["url"], json=query, timeout=240)
            response.raise_for_status()
            lines = response.text.splitlines()
            if not lines:
                continue
            if not wrote_header:
                output.write(lines[0] + "\n")
                wrote_header = True
            for line in lines[1:]:
                output.write(line + "\n")
    temporary.replace(path)
    return path, metadata


def run_denmark(ctx: Context, country: str, cfg: dict[str, Any], year: int) -> dict[str, Any]:
    source = source_for_year(cfg, year)[0]
    path, metadata = fetch_denmark(ctx, country, source, year)
    ctx.source_row(country, source, year, path, "Statistics Denmark REGR55 annual region accounts; DKK thousands converted to DKK.")
    tier = cfg["expected_tiers"][0]
    variables = {item["id"]: item for item in metadata["variables"]}
    names = {key: {row["id"]: row["text"] for row in value["values"]} for key, value in variables.items()}
    region_codes = [row["id"] for row in variables["REGI07"]["values"] if row["id"] != "000"]
    if ctx.args.max_entities:
        region_codes = region_codes[:ctx.args.max_entities]
    for code in region_codes:
        name = names["REGI07"][code]
        ctx.bundle.entity(entity_row(country, cfg, code, name, "DK_REGION_CODE", ctx.loaded_at))
        ctx.bundle.region(regional_row(country, cfg, code, name, "DK_REGION_CODE", tier, source["id"], ctx.loaded_at))
    function_class = f"DK_REGR55_FUNCTION_{year}"
    economic_class = f"DK_REGR55_ACCOUNT_{year}"
    ctx.bundle.classification(classification_row(country, function_class, "mixed", metadata["text"] + " functions", source["url"], year, ctx.loaded_at))
    ctx.bundle.classification(classification_row(country, economic_class, "mixed", metadata["text"] + " authorized group and kind", source["url"], year, ctx.loaded_at))
    art_codes = set(names["ART"])
    non_leaf = {"UE", "UI", "TOT", "I", "S0", "S2", "S4", "S5", "S6", "S7", "S8", "S9"}
    leaf_art = art_codes - non_leaf
    run_id = f"{source['id']}-{year}-v1"
    started_at = utc_now()
    rows_read = rows_loaded = 0
    entities_with_facts: set[str] = set()
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        for row_number, row in enumerate(reader, 2):
            region = stable_code(row.get("REGI07"))
            if region not in region_codes:
                continue
            rows_read += 1
            function = stable_code(row.get("FUNK1"))
            grouping = stable_code(row.get("TVGRP"))
            art = stable_code(row.get("ART"))
            amount = reported_amount(row.get("INDHOLD"))
            if art not in leaf_art or amount is None:
                continue
            amount *= Decimal(1000)
            economic = f"{grouping}.{art}"
            side = "revenue" if art.startswith(("7", "8")) or art == "97" else "expenditure"
            consolidation = art.startswith("9") or grouping == "730"
            ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, source["table"], row, ctx.loaded_at))
            ctx.bundle.node(node_row(country, function_class, "mixed", function, names["FUNK1"].get(function, function), year, ctx.loaded_at))
            economic_name = " · ".join(filter(None, (names["TVGRP"].get(grouping), names["ART"].get(art))))
            ctx.bundle.node(node_row(country, economic_class, side, economic, economic_name, year, ctx.loaded_at))
            ctx.bundle.write("regional_budget_line_facts", fact_row(
                f"DK:{region}", tier["tier_code"], year, "actual", side, function, economic, amount,
                cfg["currency"], source, run_id, ctx.loaded_at, function_class, economic_class,
                consolidation=consolidation, row_number=row_number, sheet=source["table"], quality_flags=["source_unit_dkk_thousands"],
            ))
            rows_loaded += 1
            entities_with_facts.add(region)
    ctx.bundle.write("ingestion_runs", ingestion_row(source, run_id, path, started_at, rows_read, rows_loaded))
    ctx.bundle.write("regional_budget_coverage", coverage_row(
        country, year, tier, [source["id"]], len(entities_with_facts), rows_loaded, {"actual"}, {"revenue", "expenditure"}, ctx.loaded_at,
        [
            "REGR55 provides annual accounts; enacted regional budget plans require a separate source.",
            *([] if len(entities_with_facts) == len(region_codes) else [f"The {year} table currently contains facts for {len(entities_with_facts)} of {len(region_codes)} listed regions."]),
        ],
    ))
    return {"entities": len(region_codes), "entities_with_facts": len(entities_with_facts), "facts": rows_loaded, "tiers": {tier["tier_code"]: len(entities_with_facts)}}


ADAPTERS = {
    "france_dgfip": run_france,
    "poland_jst": run_poland,
    "sweden_scb": run_sweden,
    "denmark_statbank": run_denmark,
}


def iter_jsonl(path: Path) -> Iterator[dict[str, Any]]:
    candidates = (path, path.with_suffix(path.suffix + ".gz"))
    selected = next((item for item in candidates if item.exists()), None)
    if selected is None:
        return
    opener = gzip.open if selected.suffix == ".gz" else open
    with opener(selected, "rt", encoding="utf-8") as handle:
        for line in handle:
            if line.strip():
                yield json.loads(line)


def validate_bundle(output: Path) -> dict[str, Any]:
    errors: list[str] = []
    entities = {row["public_entity_id"] for row in iter_jsonl(output / "regional_governments.jsonl")}
    classifications = {row["classification_id"] for row in iter_jsonl(output / "classification_versions.jsonl")}
    fact_count = 0
    for row in iter_jsonl(output / "regional_budget_line_facts.jsonl"):
        fact_count += 1
        if row["public_entity_id"] not in entities:
            errors.append(f"fact references missing regional government {row['public_entity_id']}")
        for key in ("functional_classification_id", "economic_classification_id"):
            if row.get(key) and row[key] not in classifications:
                errors.append(f"fact references missing classification {row[key]}")
        if not bigquery_numeric_compatible(row["amount_local"]):
            errors.append(f"amount is not BigQuery NUMERIC compatible: {row['amount_local']}")
        if row["budget_side"] not in {"revenue", "expenditure", "financing"}:
            errors.append(f"invalid budget side {row['budget_side']}")
        if len(errors) >= 100:
            break
    if not entities:
        errors.append("bundle contains no regional governments")
    if not fact_count:
        errors.append("bundle contains no regional budget facts")
    return {"status": "passed" if not errors else "failed", "regional_government_count": len(entities), "fact_count": fact_count, "errors": errors}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--countries", default="FRA,POL,SWE,DNK")
    parser.add_argument("--years", default="2024,2025")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--cache-dir", type=Path, default=DEFAULT_CACHE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--max-entities", type=int)
    parser.add_argument("--raw-mode", choices=("all", "sample", "none"), default="sample")
    parser.add_argument("--raw-limit", type=int, default=10000)
    parser.add_argument("--gzip", action="store_true")
    parser.add_argument("--offline", action="store_true")
    parser.add_argument("--refresh", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    config = json.loads(args.config.read_text(encoding="utf-8"))
    requested = [value.strip().upper() for value in args.countries.split(",") if value.strip()]
    years = sorted({int(value.strip()) for value in args.years.split(",") if value.strip()})
    unknown = sorted(set(requested) - set(config["countries"]))
    if unknown:
        raise SystemExit(f"Unknown countries: {', '.join(unknown)}")
    bundle = Bundle(args.output_dir, args.raw_mode, args.raw_limit, args.gzip)
    context = Context(args, config, bundle)
    results: dict[str, Any] = {}
    failures: dict[str, str] = {}
    try:
        for country in requested:
            cfg = config["countries"][country]
            results[country] = {}
            for year in years:
                if year not in cfg["years"]:
                    results[country][str(year)] = {"unavailable": "year_not_configured"}
                    continue
                try:
                    results[country][str(year)] = ADAPTERS[cfg["adapter"]](context, country, cfg, year)
                except Exception as exc:
                    failures[f"{country}:{year}"] = f"{type(exc).__name__}: {exc}"
                    if not args.max_entities:
                        raise
    finally:
        bundle.close()
    validation = validate_bundle(args.output_dir)
    if validation["status"] != "passed":
        failures["bundle_validation"] = "; ".join(validation["errors"][:5])
    manifest = {
        "schema_version": "1.0.0",
        "generated_at": context.loaded_at,
        "countries_requested": requested,
        "years_requested": years,
        "country_results": results,
        "queued_sources": config.get("queued_sources", []),
        "output_rows": dict(bundle.counts),
        "raw_mode": args.raw_mode,
        "validation": validation,
        "failures": failures,
        "grain": "regional government × fiscal year/period × stage × source functional code × source economic code",
        "non_additivity_warning": "Regional and municipal budgets must not be added without matched intergovernmental-transfer elimination.",
    }
    (args.output_dir / "international_regional_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    if failures:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
