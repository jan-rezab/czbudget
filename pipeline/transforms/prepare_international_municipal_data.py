#!/usr/bin/env python3
"""Download and normalize official international municipal finance data.

The adapters preserve national codes and classifications.  They do not invent
missing budget stages, consolidate supplementary budgets, or claim census
coverage where the source does not provide it.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import io
import itertools
import json
import os
import re
import struct
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable, Iterator

import requests
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from raw_cache_archives import restore as restore_raw_cache


ROOT = Path(os.environ.get("CZBUDGET_WORKSPACE_ROOT", Path(__file__).resolve().parents[3]))
DEFAULT_CONFIG = ROOT / "website/pipeline/config/international_municipal_sources.json"
DEFAULT_CACHE = ROOT / "data/source_cache/international_municipal"
DEFAULT_OUTPUT = ROOT / "outputs/international-municipal"
USER_AGENT = "czbudget-municipal-pipeline/1.0 (+https://www.czbudget.cz)"

TABLE_FILES = (
    "public_entities",
    "public_entity_sources",
    "classification_versions",
    "budget_nodes",
    "raw_budget_lines",
    "municipal_budget_line_facts",
    "public_entity_balance_sheet_facts",
    "public_entity_cash_facts",
    "ingestion_runs",
)


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


# Placeholders official publishers use to mean "no value was reported".
# Eurostat and the Nordic statistics offices use ".." and ":"; DGFiP, StatsWales
# and the German kameral exports use a dash. None of these is a zero.
MISSING_VALUE_TOKENS = frozenset({"..", ":", "-", "—", "–", "nan", "None", "NaN", "N/A", "n/a"})


def decimal_value(value: Any) -> Decimal | None:
    """Parse an official source amount.

    Returns ``None`` when the source reported **no value** — an empty cell, or
    one of the documented missing-value placeholders in
    :data:`MISSING_VALUE_TOKENS`. It never returns ``Decimal(0)`` for those.

    This is deliberate and load-bearing. The published methodology promises
    that a missing value is never treated as zero, so "the publisher reported
    0" and "the publisher reported nothing" have to stay distinguishable all
    the way from the source cell to the fact tables. Callers must branch on
    ``None`` explicitly and must not coerce it to zero; a genuine reported
    zero comes back as ``Decimal(0)``, which is falsy but *not* ``None``.

    Raises ``ValueError`` for text that is present but not a number, so an
    unexpected source schema fails loudly instead of silently reading as zero.
    """
    if value is None:
        return None
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if not text or text in MISSING_VALUE_TOKENS:
        return None
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Not a decimal value: {value!r}") from exc


def reported_amount(value: Any) -> Decimal | None:
    """Return a *loadable* amount, or ``None`` if there is nothing to load.

    Collapses the two distinct no-fact cases every adapter shares: a missing
    value (:func:`decimal_value` returned ``None``) and a reported zero, which
    carries no money and so is not emitted as a budget line. Neither is ever
    written as a zero fact. Callers that must tell the two apart keep calling
    :func:`decimal_value` directly; this helper only replaces the common
    ``if not amount: continue`` guard, which could not distinguish them.
    """
    amount = decimal_value(value)
    if amount is None or not amount:
        return None
    return amount


def numeric_json(value: Decimal) -> str:
    # BigQuery NUMERIC accepts at most nine fractional digits. Some official
    # spreadsheets expose binary-derived ratios with a longer decimal tail;
    # round only those values so source-native integer/decimal formatting stays
    # intact while every emitted amount remains warehouse-loadable.
    if value.as_tuple().exponent < -9:
        value = value.quantize(Decimal("0.000000001"), rounding=ROUND_HALF_UP)
    return format(value, "f")


def bigquery_numeric_compatible(value: Any) -> bool:
    """Return whether a serialized decimal fits BigQuery NUMERIC(38, 9)."""
    number = Decimal(str(value))
    scale = max(0, -number.as_tuple().exponent)
    integer_digits = max(1, number.adjusted() + 1)
    return scale <= 9 and integer_digits <= 29


def source_api_key(source: dict[str, Any]) -> str:
    """Return the API key for a source, read from the environment.

    Credentials are never stored in the source configuration. A source that
    needs one declares the *name* of the environment variable that carries it
    via ``api_key_env``; the value is read at run time. Missing configuration
    or a missing/blank variable is a hard failure, so a refresh can never
    silently fall through to an unauthenticated request. The key value itself
    is never echoed, logged, or included in an error message.
    """
    variable = source.get("api_key_env")
    if not variable:
        raise SystemExit(
            f"Source {source.get('id', '<unknown>')!r} requires an API key but declares no "
            "'api_key_env'. Add the environment variable name to the source configuration; "
            "never store the key itself there."
        )
    key = os.environ.get(variable, "").strip()
    if not key:
        raise SystemExit(
            f"Environment variable {variable} is not set (or is empty), and source "
            f"{source.get('id', '<unknown>')!r} cannot be refreshed without it. Export the "
            "credential in the shell that runs this transform, or use --offline / a cached "
            "snapshot."
        )
    return key


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def stable_code(value: Any) -> str:
    return str(value or "").strip()


def configured_years(args: argparse.Namespace, cfg: dict[str, Any]) -> list[int]:
    """Resolve requested years while keeping the original --year interface."""
    if args.year:
        return [args.year]
    if args.years:
        return sorted({int(value.strip()) for value in args.years.split(",") if value.strip()})
    return [int(value) for value in cfg.get("years", [cfg["year"]])]


def config_for_year(cfg: dict[str, Any], year: int) -> dict[str, Any]:
    result = dict(cfg)
    result["year"] = year
    result["sources"] = [
        source for source in cfg["sources"]
        if source.get("year") in (None, year)
        and (not source.get("years") or year in source["years"])
    ]
    return result


class JsonlBundle:
    def __init__(self, output: Path, raw_mode: str, raw_limit: int, gzip_output: bool = False):
        output.mkdir(parents=True, exist_ok=True)
        self.output = output
        self.gzip_output = gzip_output
        self.handles = {}
        for name in TABLE_FILES:
            path = output / f"{name}.jsonl{'.gz' if gzip_output else ''}"
            self.handles[name] = gzip.open(path, "wt", encoding="utf-8") if gzip_output else path.open("w", encoding="utf-8")
        self.counts: Counter[str] = Counter()
        self.raw_mode = raw_mode
        self.raw_limit = raw_limit
        self._seen_entities: set[str] = set()
        self._seen_nodes: set[str] = set()
        self._seen_versions: set[str] = set()

    def write(self, table: str, row: dict[str, Any]) -> None:
        self.handles[table].write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")
        self.counts[table] += 1

    def raw(self, row: dict[str, Any]) -> None:
        if self.raw_mode == "none":
            return
        if self.raw_mode == "sample" and self.counts["raw_budget_lines"] >= self.raw_limit:
            return
        self.write("raw_budget_lines", row)

    def entity(self, row: dict[str, Any]) -> bool:
        key = row["public_entity_id"]
        if key in self._seen_entities:
            return False
        self._seen_entities.add(key)
        self.write("public_entities", row)
        return True

    def classification(self, row: dict[str, Any]) -> None:
        key = row["classification_id"]
        if key not in self._seen_versions:
            self._seen_versions.add(key)
            self.write("classification_versions", row)

    def node(self, row: dict[str, Any]) -> None:
        key = row["budget_node_id"]
        if key not in self._seen_nodes:
            self._seen_nodes.add(key)
            self.write("budget_nodes", row)

    def close(self) -> None:
        for handle in self.handles.values():
            handle.close()


class Context:
    def __init__(self, config: dict[str, Any], args: argparse.Namespace, bundle: JsonlBundle):
        self.config = config
        self.args = args
        self.bundle = bundle
        self.loaded_at = utc_now()
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT, "Accept": "*/*"})
        retry = Retry(total=4, connect=4, read=4, backoff_factor=1.0, status_forcelist=(429, 500, 502, 503, 504), allowed_methods=("GET", "POST"))
        self.session.mount("https://", HTTPAdapter(max_retries=retry, pool_connections=max(10, args.api_workers), pool_maxsize=max(10, args.api_workers)))
        self.overrides = {}
        for item in args.source_file:
            if "=" not in item:
                raise SystemExit(f"--source-file must be SOURCE_ID=PATH, got {item!r}")
            source_id, path = item.split("=", 1)
            self.overrides[source_id] = Path(path).expanduser().resolve()

    def cache_path(self, country: str, source: dict[str, Any]) -> Path:
        if source["id"] in self.overrides:
            return self.overrides[source["id"]]
        filename = source.get("filename") or f"{source['id']}.json"
        return self.args.cache_dir / country / filename

    def download(self, country: str, source: dict[str, Any], optional: bool = False) -> Path | None:
        path = self.cache_path(country, source)
        if path.exists() and not self.args.refresh:
            return path
        if self.args.offline:
            if optional:
                return None
            raise FileNotFoundError(f"Offline source missing: {path}")
        path.parent.mkdir(parents=True, exist_ok=True)
        headers = {}
        token = self.args.openbudget_token or os.environ.get("OPENBUDGET_API_TOKEN")
        if token and country == "UKR":
            headers["Authorization"] = f"Bearer {token}"
        try:
            method = source.get("method", "GET").upper()
            with self.session.request(method, source["url"], headers=headers, data=source.get("form"), timeout=120, stream=True) as response:
                response.raise_for_status()
                temporary = path.with_suffix(path.suffix + ".part")
                with temporary.open("wb") as handle:
                    for chunk in response.iter_content(1024 * 1024):
                        if chunk:
                            handle.write(chunk)
                temporary.replace(path)
        except requests.RequestException:
            if optional:
                return None
            raise
        return path

    def api_json(self, method: str, url: str, **kwargs: Any) -> Any:
        response = self.session.request(method, url, timeout=120, **kwargs)
        response.raise_for_status()
        return response.json()

    def source_row(self, entity_id: str | None, source: dict[str, Any], country: str, notes: str) -> None:
        path = self.cache_path(country, source)
        self.bundle.write("public_entity_sources", {
            "source_id": source["id"], "public_entity_id": entity_id,
            "source_type": source["kind"], "source_name": source["id"],
            "source_url": source["url"], "dataset_code": source.get("table"),
            "archive_file": str(path.relative_to(ROOT)) if path.exists() and path.is_relative_to(ROOT) else (str(path) if path.exists() else None),
            "archive_sha256": sha256(path) if path.exists() else None,
            "retrieved_at": self.loaded_at, "notes": notes, "loaded_at": self.loaded_at,
        })


def entity_row(country: str, code: str, name: str, currency: str, loaded_at: str, **extra: Any) -> dict[str, Any]:
    alpha2 = {"POL": "PL", "DNK": "DK", "UKR": "UA", "FRA": "FR", "SWE": "SE", "GBR": "GB", "USA": "US", "CHE": "CH", "DEU": "DE", "PRY": "PY"}[country]
    return {
        "public_entity_id": f"{alpha2}:{code}", "entity_name": name,
        "entity_type": extra.pop("entity_type", "municipality"),
        "country_code_alpha2": alpha2, "country_code_alpha3": country,
        "national_entity_code": code, "national_entity_code_type": extra.pop("code_type", f"{alpha2}_MUNICIPALITY_CODE"),
        "is_eu_capital": False, "is_extra_city": False, "default_currency_code": currency,
        "eurostat_city_code": None, "eurostat_geography_name": None,
        "administrative_region_code": extra.pop("region_code", None),
        "administrative_region_name": extra.pop("region_name", None),
        "administrative_district_code": extra.pop("district_code", None),
        "administrative_district_name": extra.pop("district_name", None),
        "national_geography_code": extra.pop("geography_code", code),
        "national_geography_code_type": extra.pop("geography_type", f"{alpha2}_MUNICIPALITY_CODE"),
        "valid_from": None, "valid_to": None, "loaded_at": loaded_at,
    }


def classification_row(country: str, classification_id: str, side: str, name: str, source_url: str, year: int, loaded_at: str) -> dict[str, Any]:
    return {
        "classification_id": classification_id, "country_code": country, "budget_side": side,
        "government_scope": "municipal", "valid_from_year": year, "valid_to_year": None,
        "classification_name": name, "legal_basis": None, "source_url": source_url,
        "notes": "National source classification preserved without cross-country remapping.", "loaded_at": loaded_at,
    }


def node_row(country: str, classification_id: str, side: str, code: str, name: str, year: int, loaded_at: str) -> dict[str, Any]:
    return {
        "budget_node_id": f"{country}:{classification_id}:{code}", "classification_id": classification_id,
        "country_code": country, "budget_side": side, "government_scope": "municipal",
        "node_code": code, "node_name_native": name or code, "node_name_en": None, "node_name_cs": None,
        "parent_budget_node_id": None, "hierarchy_level": 1, "hierarchy_path": [], "is_chapter": False,
        "effective_from_year": year, "effective_to_year": None, "loaded_at": loaded_at,
    }


def fact_row(
    country: str, entity_id: str, year: int, stage: str, side: str, function_code: str | None,
    economic_code: str, amount: Decimal, currency: str, source: dict[str, Any], run_id: str,
    loaded_at: str, function_class: str | None, economic_class: str, **extra: Any,
) -> dict[str, Any]:
    return {
        "public_entity_id": entity_id, "fiscal_year": year, "fiscal_period": extra.pop("period", "FY"),
        "reporting_scope": extra.pop("scope", "standalone_municipality"), "budget_stage": stage,
        "budget_side": side, "source_budget_item_type_code": extra.pop("item_type", None),
        "functional_paragraph_code": function_code, "economic_item_code": economic_code,
        "functional_classification_id": function_class, "economic_classification_id": economic_class,
        "amount_local": numeric_json(amount), "currency_code": currency, "amount_eur": None, "fx_date": None,
        "is_consolidation_item": False, "is_financing": side == "financing", "is_summary_row": extra.pop("summary", False),
        "source_row_number": extra.pop("row_number", None), "source_sheet": extra.pop("sheet", None),
        "source_id": source["id"], "ingestion_run_id": run_id,
        "coverage_type": extra.pop("coverage_type", "census"), "is_imputed": extra.pop("is_imputed", False),
        "quality_flags": extra.pop("quality_flags", []), "loaded_at": loaded_at,
    }


def raw_row(country: str, year: int, source: dict[str, Any], run_id: str, row_number: int, sheet: str, payload: dict[str, Any], loaded_at: str) -> dict[str, Any]:
    return {
        "country_code": country, "fiscal_year": year, "source_id": source["id"],
        "ingestion_run_id": run_id, "source_row_number": row_number, "source_sheet": sheet,
        "source_payload": payload, "source_url": source["url"], "loaded_at": loaded_at,
    }


def iter_dbf(handle: io.BufferedReader, encoding: str = "cp1250") -> Iterator[tuple[int, dict[str, Any]]]:
    header = handle.read(32)
    if len(header) != 32:
        raise ValueError("Invalid DBF header")
    record_count = struct.unpack("<I", header[4:8])[0]
    header_length = struct.unpack("<H", header[8:10])[0]
    record_length = struct.unpack("<H", header[10:12])[0]
    fields = []
    while handle.tell() < header_length - 1:
        descriptor = handle.read(32)
        if not descriptor or descriptor[0] == 0x0D:
            break
        name = descriptor[:11].split(b"\0", 1)[0].decode("ascii")
        fields.append((name, chr(descriptor[11]), descriptor[16], descriptor[17]))
    handle.seek(header_length)
    for row_number in range(1, record_count + 1):
        record = handle.read(record_length)
        if len(record) < record_length:
            break
        if record[:1] == b"*":
            continue
        position = 1
        row: dict[str, Any] = {}
        for name, kind, length, decimals in fields:
            raw = record[position:position + length]
            position += length
            text = raw.decode(encoding, errors="replace").strip()
            if kind in "NF" and text:
                # ``None`` here means the DBF numeric column held a missing-value
                # placeholder; it is preserved as absent rather than as zero.
                row[name] = decimal_value(text)
            else:
                row[name] = text or None
        yield row_number, row


ODS_TABLE_NS = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
ODS_OFFICE_NS = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
ODS_TEXT_NS = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"


def iter_ods_rows(path: Path, sheet_name: str, max_columns: int = 1000) -> Iterator[tuple[int, list[Any]]]:
    """Read one ODS sheet using only the Python standard library.

    Government ODS releases often encode long blank tails as repeated cells or
    rows. Blank repetitions are collapsed, while populated repetitions are
    retained. This keeps the parser deterministic without requiring LibreOffice
    or a third-party ODF package in production.
    """
    q = lambda namespace, name: f"{{{namespace}}}{name}"
    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("content.xml"))
    table = next((item for item in root.iter(q(ODS_TABLE_NS, "table")) if item.get(q(ODS_TABLE_NS, "name")) == sheet_name), None)
    if table is None:
        raise ValueError(f"ODS workbook missing sheet: {sheet_name}")
    output_row = 0
    for row_element in table.iter(q(ODS_TABLE_NS, "table-row")):
        row: list[Any] = []
        for cell in row_element:
            if cell.tag not in {q(ODS_TABLE_NS, "table-cell"), q(ODS_TABLE_NS, "covered-table-cell")}:
                continue
            repeat = min(int(cell.get(q(ODS_TABLE_NS, "number-columns-repeated"), "1")), max_columns - len(row))
            value_type = cell.get(q(ODS_OFFICE_NS, "value-type"))
            value: Any = None
            if value_type in {"float", "currency", "percentage"}:
                raw_value = cell.get(q(ODS_OFFICE_NS, "value"))
                value = Decimal(raw_value) if raw_value not in (None, "") else None
            elif value_type == "boolean":
                value = cell.get(q(ODS_OFFICE_NS, "boolean-value")) == "true"
            elif value_type == "date":
                value = cell.get(q(ODS_OFFICE_NS, "date-value"))
            else:
                pieces = ["".join(paragraph.itertext()) for paragraph in cell.iter(q(ODS_TEXT_NS, "p"))]
                value = "\n".join(piece for piece in pieces if piece) or None
            row.extend([value] * max(0, repeat))
            if len(row) >= max_columns:
                break
        while row and row[-1] is None:
            row.pop()
        row_repeat = int(row_element.get(q(ODS_TABLE_NS, "number-rows-repeated"), "1"))
        repetitions = row_repeat if row else 1
        for _ in range(repetitions):
            output_row += 1
            yield output_row, list(row)


def _pl_dictionary(path: Path) -> dict[tuple[str, str, str, str, str], dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [stable_code(value).replace("\u200b", "") for value in next(rows)]
    result = {}
    for values in rows:
        row = dict(zip(headers, values))
        key = tuple(stable_code(row.get(code)).zfill(2 if code != "GT" and code != "PT" else 1) for code in ("WK", "PK", "GK", "GT", "PT"))
        kind = stable_code(row.get("Typ Jst Nazwa"))
        if "gmina" in kind.lower() or "miasto na prawach" in kind.lower():
            result[key] = {"name": stable_code(row.get("Nazwa")), "code": stable_code(row.get("Kod GUS")).zfill(7), "type": kind}
    return result


def run_poland(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    year, currency = cfg["year"], cfg["currency"]
    sources = {source["kind"]: source for source in cfg["sources"]}
    paths = {kind: ctx.download(country, source) for kind, source in sources.items()}
    dictionary = _pl_dictionary(paths["entities"])
    selected = list(dictionary.values())[:ctx.args.max_entities or None]
    selected_codes = {item["code"] for item in selected}
    for key, item in dictionary.items():
        if item["code"] not in selected_codes:
            continue
        ctx.bundle.entity(entity_row(country, item["code"], item["name"], currency, ctx.loaded_at, code_type="PL_TERYT"))
    for side, kind in (("revenue", "revenue"), ("expenditure", "expenditure")):
        source = sources[kind]
        run_id = f"{source['id']}-v1"
        function_class = f"PL_JST_FUNCTION_{year}"
        economic_class = f"PL_JST_{side.upper()}_{year}"
        ctx.bundle.classification(classification_row(country, function_class, "mixed", "Polish JST division/chapter classification", source["url"], year, ctx.loaded_at))
        ctx.bundle.classification(classification_row(country, economic_class, side, f"Polish JST {side} paragraph classification", source["url"], year, ctx.loaded_at))
        rows_read = rows_loaded = 0
        with zipfile.ZipFile(paths[kind]) as archive:
            member = next(name for name in archive.namelist() if name.lower().endswith(".dbf"))
            with archive.open(member) as handle:
                for row_number, row in iter_dbf(handle):
                    key = tuple(stable_code(row.get(code)).zfill(2 if code not in ("GT", "PT") else 1) for code in ("WK", "PK", "GK", "GT", "PT"))
                    item = dictionary.get(key)
                    if not item or item["code"] not in selected_codes:
                        continue
                    rows_read += 1
                    payload = {key: numeric_json(value) if isinstance(value, Decimal) else value for key, value in row.items()}
                    ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, member, payload, ctx.loaded_at))
                    function = f"{stable_code(row.get('DZIAL')).zfill(3)}.{stable_code(row.get('ROZDZIAL')).zfill(5)}"
                    economic = f"{stable_code(row.get('PAR')).zfill(3)}{stable_code(row.get('PAR4'))}"
                    ctx.bundle.node(node_row(country, function_class, side, function, function, year, ctx.loaded_at))
                    ctx.bundle.node(node_row(country, economic_class, side, economic, economic, year, ctx.loaded_at))
                    for stage, column in (("revised", "R1"), ("actual", "R4")):
                        amount = reported_amount(row.get(column))
                        if amount is None:
                            continue
                        ctx.bundle.write("municipal_budget_line_facts", fact_row(
                            country, f"PL:{item['code']}", year, stage, side, function, economic, amount,
                            currency, source, run_id, ctx.loaded_at, function_class, economic_class,
                            row_number=row_number, sheet=member, quality_flags=["quarter_four_return"],
                        ))
                        rows_loaded += 1
        write_run(ctx, run_id, source, year, paths[kind], rows_read, rows_loaded)
        ctx.source_row(None, source, country, cfg["coverage"])
    ctx.source_row(None, sources["entities"], country, cfg["coverage"])
    return {"entities": len(selected_codes)}


def statbank_metadata(ctx: Context, table: str) -> dict[str, Any]:
    return ctx.api_json("GET", f"https://api.statbank.dk/v1/tableinfo/{table}?lang=en")


def run_denmark(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    year, currency = cfg["year"], cfg["currency"]
    budget_table = next(source["table"] for source in cfg["sources"] if source["kind"] == "enacted")
    budget_meta = statbank_metadata(ctx, budget_table)
    region_var = budget_meta["variables"][0]
    aggregate_codes = {"000", "081", "082", "083", "084", "085", "910", "920", "930", "940", "960"}
    municipalities = [value for value in region_var["values"] if value["id"] not in aggregate_codes]
    municipalities = municipalities[:ctx.args.max_entities or None]
    for municipality in municipalities:
        ctx.bundle.entity(entity_row(country, municipality["id"], municipality["text"], currency, ctx.loaded_at, code_type="DK_KOMKODE"))
    for source in cfg["sources"]:
        meta = statbank_metadata(ctx, source["table"])
        variables = {item["id"]: item for item in meta["variables"]}
        region_code = meta["variables"][0]["id"]
        function_class = f"DK_{source['table']}_FUNCTION"
        # The 100-series tables expose the authorized account below the older
        # function x kind cube. Keep ownership and grouping in the normalized
        # economic key so equal ART codes are not silently collapsed.
        economic_class = f"DK_{source['table']}_ACCOUNT"
        ctx.bundle.classification(classification_row(country, function_class, "mixed", meta["text"] + " functions", source["url"], year, ctx.loaded_at))
        ctx.bundle.classification(classification_row(country, economic_class, "mixed", meta["text"] + " kinds", source["url"], year, ctx.loaded_at))
        function_names = {item["id"]: item["text"] for item in variables["FUNKTION"]["values"]}
        art_names = {item["id"]: item["text"] for item in variables["ART"]["values"]}
        owner_names = {item["id"]: item["text"] for item in variables.get("EJER", {}).get("values", [])}
        grouping_names = {item["id"]: item["text"] for item in variables.get("GRUPPERING", {}).get("values", [])}
        leaf_art = [code for code in art_names if code not in {"TOT", "UE", "I", "U"} and not code.startswith("S")]
        run_id = f"{source['id']}-{year}-v1"
        rows_read = rows_loaded = 0
        for municipality in municipalities:
            query = {"table": source["table"], "format": "BULK", "lang": "en", "variables": []}
            for variable in meta["variables"]:
                code = variable["id"]
                if code == region_code:
                    values = [municipality["id"]]
                elif code == "ART":
                    values = leaf_art
                elif code == "PRISENHED":
                    values = ["LOBM"]
                elif variable.get("time"):
                    values = [str(year)]
                else:
                    values = [item["id"] for item in variable["values"]]
                query["variables"].append({"code": code, "values": values})
            response = ctx.session.post(source["url"], json=query, timeout=180)
            response.raise_for_status()
            reader = csv.DictReader(io.StringIO(response.text), delimiter=";")
            for row_number, row in enumerate(reader, 2):
                rows_read += 1
                function, art = stable_code(row.get("FUNKTION")), stable_code(row.get("ART"))
                dranst = stable_code(row.get("DRANST"))
                owner = stable_code(row.get("EJER"))
                grouping = stable_code(row.get("GRUPPERING"))
                amount = reported_amount(row.get("INDHOLD"))
                if amount is None or not function or not art:
                    continue
                amount *= Decimal(1000)
                account = ".".join(part for part in (dranst, owner, grouping, art) if part)
                account_name = " · ".join(part for part in (owner_names.get(owner), grouping_names.get(grouping), art_names.get(art)) if part)
                side = "revenue" if art.startswith(("7", "8")) else ("financing" if dranst in {"5", "6", "7"} else "expenditure")
                ctx.bundle.node(node_row(country, function_class, side, function, function_names.get(function, function), year, ctx.loaded_at))
                ctx.bundle.node(node_row(country, economic_class, side, account, account_name or account, year, ctx.loaded_at))
                payload = dict(row)
                ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, source["table"], payload, ctx.loaded_at))
                ctx.bundle.write("municipal_budget_line_facts", fact_row(
                    country, f"DK:{municipality['id']}", year, source["kind"], side, function, account, amount,
                    currency, source, run_id, ctx.loaded_at, function_class, economic_class,
                    row_number=row_number, sheet=source["table"], item_type=dranst, quality_flags=["source_unit_dkk_thousands", "authorized_account_detail"],
                ))
                rows_loaded += 1
        write_run(ctx, run_id, source, year, None, rows_read, rows_loaded)
        ctx.source_row(None, source, country, cfg["coverage"])
    return {"entities": len(municipalities)}


def jsonstat_rows(dataset: dict[str, Any]) -> Iterator[dict[str, Any]]:
    dimensions = dataset["id"]
    categories = []
    for dimension in dimensions:
        category = dataset["dimension"][dimension]["category"]
        index = category["index"]
        codes = index if isinstance(index, list) else [code for code, _ in sorted(index.items(), key=lambda item: item[1])]
        labels = category.get("label", {})
        categories.append([(code, labels.get(code, code)) for code in codes])
    values = dataset["value"]
    for position, combination in enumerate(itertools.product(*categories)):
        value = values.get(str(position)) if isinstance(values, dict) else values[position]
        if value is None:
            continue
        row = {dimension: code_label[0] for dimension, code_label in zip(dimensions, combination)}
        row.update({f"{dimension}_label": code_label[1] for dimension, code_label in zip(dimensions, combination)})
        row["value"] = value
        yield row


def pxweb_query(ctx: Context, source: dict[str, Any], metadata: dict[str, Any], region: str, year: int) -> dict[str, Any]:
    query = []
    for variable in metadata["variables"]:
        if variable["code"] == "Region":
            values = [region]
        elif variable.get("time"):
            values = [str(year)]
        else:
            values = variable["values"]
        query.append({"code": variable["code"], "selection": {"filter": "item", "values": values}})
    return ctx.api_json("POST", source["url"], json={"query": query, "response": {"format": "json-stat2"}})


def run_sweden(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    year, currency = cfg["year"], cfg["currency"]
    metadata = {source["kind"]: ctx.api_json("GET", source["url"]) for source in cfg["sources"]}
    region_variable = metadata["expenditure"]["variables"][0]
    municipalities = [
        {"id": code, "text": text} for code, text in zip(region_variable["values"], region_variable["valueTexts"])
        if code != "00"
    ][:ctx.args.max_entities or None]
    for municipality in municipalities:
        ctx.bundle.entity(entity_row(country, municipality["id"], municipality["text"], currency, ctx.loaded_at, code_type="SE_KOMMUNKOD"))
    for source in cfg["sources"]:
        meta = metadata[source["kind"]]
        line_variable = next(item for item in meta["variables"] if item["code"] not in {"Region", "ContentsCode", "Tid"})
        line_names = dict(zip(line_variable["values"], line_variable["valueTexts"]))
        classification_id = f"SE_SCB_{source['table'].upper()}"
        side = source["kind"] if source["kind"] in {"revenue", "expenditure"} else "mixed"
        ctx.bundle.classification(classification_row(country, classification_id, side, meta["title"], source["url"], year, ctx.loaded_at))
        run_id = f"{source['id']}-{year}-v1"
        rows_read = rows_loaded = 0
        for municipality in municipalities:
            data = pxweb_query(ctx, source, meta, municipality["id"], year)
            for row_number, row in enumerate(jsonstat_rows(data), 1):
                rows_read += 1
                code = stable_code(row.get(line_variable["code"]))
                amount = reported_amount(row["value"])
                if amount is None:
                    continue
                amount *= Decimal(1000)
                ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, source["table"], row, ctx.loaded_at))
                if source["kind"] == "balance":
                    ctx.bundle.write("public_entity_balance_sheet_facts", {
                        "public_entity_id": f"SE:{municipality['id']}", "statement_date": f"{year}-12-31",
                        "reporting_scope": "standalone_municipality", "statement_line_code": code,
                        "account_code": code, "account_name": line_names.get(code), "balance_measure": "current_net",
                        "amount_local": numeric_json(amount), "currency_code": currency, "amount_eur": None, "fx_date": None,
                        "source_id": source["id"], "ingestion_run_id": run_id, "source_row_number": row_number,
                        "source_sheet": source["table"], "coverage_type": "census", "is_imputed": False,
                        "quality_flags": ["source_unit_sek_thousands", "preliminary_latest_year"], "loaded_at": ctx.loaded_at,
                    })
                else:
                    ctx.bundle.node(node_row(country, classification_id, side, code, line_names.get(code, code), year, ctx.loaded_at))
                    ctx.bundle.write("municipal_budget_line_facts", fact_row(
                        country, f"SE:{municipality['id']}", year, "actual", side, None, code, amount,
                        currency, source, run_id, ctx.loaded_at, None, classification_id,
                        row_number=row_number, sheet=source["table"],
                        summary="total" in line_names.get(code, "").lower(),
                        quality_flags=["source_unit_sek_thousands", "preliminary_latest_year"],
                    ))
                rows_loaded += 1
        write_run(ctx, run_id, source, year, None, rows_read, rows_loaded)
        ctx.source_row(None, source, country, cfg["coverage"])
    return {"entities": len(municipalities)}


def french_insee(row: dict[str, Any]) -> str:
    department = stable_code(row.get("NDEPT"))
    if len(department) == 3 and department.startswith("0"):
        department = department[1:]
    commune = stable_code(row.get("INSEE")).zfill(3)
    return department + commune


def french_csv_rows(path: Path) -> Iterator[tuple[int, str, dict[str, Any]]]:
    with zipfile.ZipFile(path) as archive:
        member = next(name for name in archive.namelist() if name.lower().endswith(".csv"))
        with archive.open(member) as raw:
            reader = csv.DictReader(io.TextIOWrapper(raw, encoding="cp1252", newline=""), delimiter=";")
            for row_number, row in enumerate(reader, 2):
                yield row_number, member, row


def run_france(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    year, currency = cfg["year"], cfg["currency"]
    sources = {source.get("detail"): source for source in cfg["sources"]}
    census_source, function_source = sources.get("census"), sources.get("function")
    if not census_source or not function_source:
        raise ValueError(f"France {year} requires census and function sources")
    census_path = ctx.download(country, census_source)
    function_path = ctx.download(country, function_source)
    function_class, economic_class = f"FR_DGFIP_FUNCTION_{year}", f"FR_DGFIP_ACCOUNT_{year}"
    ctx.bundle.classification(classification_row(country, function_class, "mixed", "DGFiP functional classification", function_source["url"], year, ctx.loaded_at))
    ctx.bundle.classification(classification_row(country, economic_class, "mixed", "DGFiP local chart of accounts", census_source["url"], year, ctx.loaded_at))
    selected: set[str] = set()
    functional_codes: set[str] = set()
    run_counts: dict[str, list[int]] = {source["id"]: [0, 0] for source in (function_source, census_source)}

    def load(source: dict[str, Any], path: Path, *, functional: bool) -> None:
        run_id = f"{source['id']}-v1"
        for row_number, member, row in french_csv_rows(path):
            if row.get("CATEG") != "Commune":
                continue
            code = french_insee(row)
            if code not in selected:
                if ctx.args.max_entities and len(selected) >= ctx.args.max_entities:
                    continue
                selected.add(code)
                ctx.bundle.entity(entity_row(country, code, stable_code(row.get("LBUDG")) or code, currency, ctx.loaded_at, code_type="FR_CODE_INSEE", region_code=row.get("CREGI"), district_code=row.get("NDEPT")))
            if functional:
                functional_codes.add(code)
            elif code in functional_codes:
                continue
            run_counts[source["id"]][0] += 1
            ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, member, row, ctx.loaded_at))
            function = stable_code(row.get("FONCTION")) or ("UNSPECIFIED" if functional else None)
            account = stable_code(row.get("COMPTE"))
            if not account:
                continue
            scope = "main_budget" if stable_code(row.get("CBUDG")) == "1" else "supplementary_budget"
            if function:
                ctx.bundle.node(node_row(country, function_class, "mixed", function, function, year, ctx.loaded_at))
            ctx.bundle.node(node_row(country, economic_class, "mixed", account, account, year, ctx.loaded_at))
            flags = [f"nomenclature:{stable_code(row.get('NOMEN'))}", "accounting_balance_execution"]
            if not functional:
                flags.append("functional_detail_not_published")
            for side, column in (("expenditure", "OBNETDEB"), ("revenue", "OBNETCRE")):
                amount = reported_amount(row.get(column))
                if amount is None:
                    continue
                ctx.bundle.write("municipal_budget_line_facts", fact_row(
                    country, f"FR:{code}", year, "actual", side, function, account, amount,
                    currency, source, run_id, ctx.loaded_at, function_class if function else None, economic_class,
                    row_number=row_number, sheet=member, scope=scope, quality_flags=flags,
                ))
                run_counts[source["id"]][1] += 1
            debit, credit = decimal_value(row.get("SD")), decimal_value(row.get("SC"))
            if debit or credit:
                # A blank counterpart column in a two-column DGFiP balance is a
                # nil side, not an unknown amount, so it nets out at zero here.
                # Which column was absent is recorded on the fact rather than
                # being lost, because "reported 0" and "reported nothing" are
                # different statements about the source.
                signed = (debit or Decimal(0)) - (credit or Decimal(0))
                balance_flags = [
                    "debit_minus_credit", *flags,
                    *(["debit_column_not_reported"] if debit is None else []),
                    *(["credit_column_not_reported"] if credit is None else []),
                ]
                ctx.bundle.write("public_entity_balance_sheet_facts", {
                    "public_entity_id": f"FR:{code}", "statement_date": f"{year}-12-31", "reporting_scope": scope,
                    "statement_line_code": function or account, "account_code": account, "account_name": None,
                    "balance_measure": "current_net_signed_debit", "amount_local": numeric_json(signed),
                    "currency_code": currency, "amount_eur": numeric_json(signed), "fx_date": f"{year}-12-31",
                    "source_id": source["id"], "ingestion_run_id": run_id, "source_row_number": row_number,
                    "source_sheet": member, "coverage_type": "census", "is_imputed": False,
                    "quality_flags": balance_flags, "loaded_at": ctx.loaded_at,
                })
                run_counts[source["id"]][1] += 1

    load(function_source, function_path, functional=True)
    load(census_source, census_path, functional=False)
    for source, path in ((function_source, function_path), (census_source, census_path)):
        rows_read, rows_loaded = run_counts[source["id"]]
        write_run(ctx, f"{source['id']}-v1", source, year, path, rows_read, rows_loaded)
        ctx.source_row(None, source, country, cfg["coverage"])
    return {"entities": len(selected), "functional_entities": len(functional_codes)}


def run_france_budgets(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    year, currency = cfg["year"], cfg["currency"]
    sources = [source for source in cfg["sources"] if source.get("collection") == "budget"]
    total_rows = total_facts = 0
    for source in sources:
        path = ctx.download(country, source); code = source["city_code"]; name = source["city_name"]
        entity_id = f"FR:{code}"
        ctx.bundle.entity(entity_row(country, code, name, currency, ctx.loaded_at, code_type="FR_CODE_INSEE"))
        function_class = f"FR_{code}_BUDGET_FUNCTION_{year}"; economic_class = f"FR_{code}_BUDGET_NATURE_{year}"
        ctx.bundle.classification(classification_row(country, function_class, "mixed", f"{name} enacted-budget function", source["url"], year, ctx.loaded_at))
        ctx.bundle.classification(classification_row(country, economic_class, "mixed", f"{name} enacted-budget nature/account", source["url"], year, ctx.loaded_at))
        run_id = f"{source['id']}-v1"; rows_read = rows_loaded = 0
        with path.open("r", encoding=source.get("encoding", "utf-8-sig"), newline="") as handle:
            for row_number, row in enumerate(csv.DictReader(handle, delimiter=";"), 2):
                parser = source["parser"]; measures: list[tuple[str, Any]] = []
                if parser == "toulouse":
                    side = "revenue" if stable_code(row.get("Nature AP/EPCP")).lower().startswith("rec") else "expenditure"
                    function = stable_code(row.get("Code fonction")); function_name = stable_code(row.get("Libellé fonction")) or function
                    economic = stable_code(row.get("Code article")); economic_name = stable_code(row.get("Libellé article")) or economic
                    measures = [(side, row.get("Mt Voté BP"))]
                elif parser == "blagnac":
                    function = stable_code(row.get("Code fonction")); function_name = stable_code(row.get("Libellé fonction")) or function
                    economic = stable_code(row.get("Article")); economic_name = stable_code(row.get("Libellé Article")) or economic
                    measures = [("expenditure", row.get("MT Voté BP Dépense")), ("revenue", row.get("MT Voté BP Recette"))]
                elif parser == "brest":
                    side = "revenue" if stable_code(row.get("Dépense / Recette")).upper().startswith("R") else "expenditure"
                    function = stable_code(row.get("Code ss-chap / ss-fonc régltr") or row.get("Code chap. / fonction")); function_name = stable_code(row.get("Lib. chap. / fonction")) or function
                    economic = stable_code(row.get("Code article / nature réglementaire")); economic_name = stable_code(row.get("Lib. article / nature")) or economic
                    measures = [(side, row.get("Budg. primitif imp. bud."))]
                elif parser == "nantes":
                    side = "revenue" if stable_code(row.get("Dépenses/Recettes")).upper().startswith("R") else "expenditure"
                    function = stable_code(row.get("Code sous-fonction")); function_name = stable_code(row.get("Libellé sous-fonction")) or function
                    economic = stable_code(row.get("Code article")); economic_name = stable_code(row.get("Libellé article")) or economic
                    measures = [(side, row.get("Montant"))]
                elif parser == "la_possession":
                    side = "revenue" if stable_code(row.get("BGT_CODRD")).lower().startswith("rec") else "expenditure"
                    function = stable_code(row.get("BGT_FONCTION")); function_name = stable_code(row.get("BGT_FONCTION_LABEL")) or function
                    economic = stable_code(row.get("BGT_NATURE")); economic_name = stable_code(row.get("BGT_NATURE_LABEL")) or economic
                    measures = [(side, row.get("BGT_MTPREV"))]
                elif parser == "sailly":
                    side = "revenue" if stable_code(row.get("sens")).lower().startswith("rec") else "expenditure"
                    function, function_name = _de_split_code(row.get("axe1")); economic, economic_name = _de_split_code(row.get("compte"))
                    measures = [(side, row.get(" prevu "))]
                else:
                    raise ValueError(f"Unsupported French budget parser: {parser}")
                rows_read += 1; ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, path.name, row, ctx.loaded_at))
                for side, raw_amount in measures:
                    amount = german_decimal(raw_amount)
                    if amount is None or not amount:
                        continue
                    function = function or "UNSPECIFIED"; economic = economic or "UNSPECIFIED"
                    function_code = f"{side}:{function}"; economic_code = f"{side}:{economic}"
                    ctx.bundle.node(node_row(country, function_class, side, function_code, function_name or function, year, ctx.loaded_at))
                    ctx.bundle.node(node_row(country, economic_class, side, economic_code, economic_name or economic, year, ctx.loaded_at))
                    ctx.bundle.write("municipal_budget_line_facts", fact_row(
                        country, entity_id, year, "enacted", side, function_code, economic_code, abs(amount),
                        currency, source, run_id, ctx.loaded_at, function_class, economic_class,
                        row_number=row_number, sheet=path.name, coverage_type="published_subset",
                        quality_flags=["official_municipal_enacted_budget", "data_gouv_catalog_discovery", "france_decentralized_budget_publication"],
                    )); rows_loaded += 1
        write_run(ctx, run_id, source, year, path, rows_read, rows_loaded); ctx.source_row(entity_id, source, country, cfg["coverage"])
        total_rows += rows_read; total_facts += rows_loaded
    return {"entities": len(sources), "rows": total_rows, "facts": total_facts, "fiscal_years": [year]}


def run_france_structured(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    results: dict[str, Any] = {}
    actual_sources = [source for source in cfg["sources"] if source.get("detail")]
    if actual_sources:
        results["national_actuals"] = run_france(ctx, country, {**cfg, "sources": actual_sources})
    budget_sources = [source for source in cfg["sources"] if source.get("collection") == "budget"]
    if budget_sources:
        results["enacted_budgets"] = run_france_budgets(ctx, country, {**cfg, "sources": budget_sources})
    results["entities"] = sum(int(result.get("entities", 0)) for result in results.values() if isinstance(result, dict))
    return results


def uk_side(variable: str) -> str:
    lower = variable.lower()
    if any(token in lower for token in ("income", "receipt", "grant", "sales", "fees")):
        return "revenue"
    if any(token in lower for token in ("reserve", "financing", "interest", "debt")):
        return "financing"
    return "expenditure"


def run_england(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    source = cfg["sources"][0]
    path = ctx.download(country, source)
    currency = cfg["currency"]
    classification_id = "GB_ENGLAND_MHCLG_RO"
    ctx.bundle.classification(classification_row(country, classification_id, "mixed", "England Revenue Outturn variables", source["url"], 2017, ctx.loaded_at))
    target_year = cfg["year"]
    run_id = f"{source['id']}-{target_year}-v1"
    identifier_columns = {"year_ending", "ONS_code", "LA_LGF_code", "LA_name", "status", "LA_class", "LA_subclass"}
    rows_read = rows_loaded = 0
    selected: set[str] = set()
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, 2):
            year_ending = int(stable_code(row.get("year_ending"))[:4])
            if year_ending != target_year:
                continue
            allowed_subclasses = {"Shire District", "Unitary Authority", "Met District", "London", "Shire County", "UA", "GLA"}
            if stable_code(row.get("status")) != "submitted" or stable_code(row.get("LA_subclass")) not in allowed_subclasses:
                continue
            code = stable_code(row.get("ONS_code") or row.get("LA_LGF_code"))
            if not code:
                continue
            if code not in selected:
                if ctx.args.max_entities and len(selected) >= ctx.args.max_entities:
                    continue
                selected.add(code)
                ctx.bundle.entity(entity_row(country, code, stable_code(row.get("LA_name")) or code, currency, ctx.loaded_at, code_type="GB_ONS_LOCAL_AUTHORITY", region_name="England", entity_type="local_authority"))
            rows_read += 1
            ctx.bundle.raw(raw_row(country, year_ending, source, run_id, row_number, "Revenue Outturn", row, ctx.loaded_at))
            for variable, value in row.items():
                if variable in identifier_columns:
                    continue
                amount = reported_amount(value)
                if amount is None:
                    continue
                amount *= Decimal(1000)
                side = uk_side(variable)
                ctx.bundle.node(node_row(country, classification_id, side, variable, variable, year_ending, ctx.loaded_at))
                ctx.bundle.write("municipal_budget_line_facts", fact_row(
                    country, f"GB:{code}", year_ending, "actual", side, None, variable, amount,
                    currency, source, run_id, ctx.loaded_at, None, classification_id,
                    row_number=row_number, sheet="Revenue Outturn", period=f"FY-{year_ending}",
                    scope="england_statistical_return",
                    summary=("_tot_" in variable.lower() or variable.lower().endswith(("_tot", "_total"))),
                    quality_flags=["source_unit_gbp_thousands", "wide_return_unpivoted"],
                ))
                rows_loaded += 1
    write_run(ctx, run_id, source, target_year, path, rows_read, rows_loaded)
    ctx.source_row(None, source, country, cfg["coverage"])
    return {"entities": len(selected)}


def run_england_budget(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    """Load England's individual-authority enacted Revenue Account budget.

    The MHCLG release is a wide ODS return. Each source variable is retained as
    the native economic item, its published section heading is retained as the
    function, and signed values are normalized to a positive amount on an
    explicit budget side. Only general-purpose councils are included; police,
    fire, waste, national-park and combined authorities are outside scope.
    """
    year, currency = cfg["year"], cfg["currency"]
    sources = [source for source in cfg["sources"] if source.get("collection") == "england_budget"]
    classification_id = f"GB_ENGLAND_MHCLG_RA_{year}"
    ctx.bundle.classification(classification_row(
        country, classification_id, "mixed", f"England Revenue Account budget variables {year}-{year + 1}",
        sources[0]["url"], year, ctx.loaded_at,
    ))
    selected: set[str] = set()
    total_rows = total_facts = 0
    allowed_gss_prefixes = ("E06", "E07", "E08", "E09", "E10")
    for source in sources:
        path = ctx.download(country, source)
        sheet_name = source.get("sheet", f"RA_LA_Data_{year}-{str(year + 1)[-2:]}")
        rows = list(iter_ods_rows(path, sheet_name))
        header_index = next((index for index, (_, row) in enumerate(rows) if row[:3] == ["E-code", "ONS Code", "Local authority"]), None)
        if header_index is None:
            raise ValueError(f"England RA source {source['id']} has no authority header row")
        header_number, headers = rows[header_index]
        metadata_rows = rows[:header_index]
        asset_ids = next((row for _, row in metadata_rows if stable_code(row[0] if row else None).lower().startswith("this row contains the data base asset id")), [])
        sections = next((row for _, row in metadata_rows if "section headings" in stable_code(row[0] if row else None).lower()), [])
        lines = next((row for _, row in metadata_rows if "line number" in stable_code(row[0] if row else None).lower()), [])
        run_id = f"{source['id']}-v1"
        rows_read = rows_loaded = 0
        current_sections: list[str] = []
        current = "Revenue Account"
        for column in range(len(headers)):
            if column < len(sections) and stable_code(sections[column]):
                current = stable_code(sections[column])
            current_sections.append(current)
        for row_number, values in rows[header_index + 1:]:
            ecode = stable_code(values[0] if values else None)
            gss_code = stable_code(values[1] if len(values) > 1 else None)
            name = stable_code(values[2] if len(values) > 2 else None)
            if ecode == "Eng" or gss_code == "E92000001":
                break
            if not gss_code.startswith(allowed_gss_prefixes):
                continue
            if gss_code not in selected:
                if ctx.args.max_entities and len(selected) >= ctx.args.max_entities:
                    continue
                selected.add(gss_code)
                ctx.bundle.entity(entity_row(
                    country, gss_code, name or gss_code, currency, ctx.loaded_at,
                    code_type="GB_GSS_LOCAL_AUTHORITY", region_name="England", entity_type="local_authority",
                ))
            rows_read += 1
            raw_payload = {
                stable_code(headers[index]) or f"column_{index + 1}": numeric_json(value) if isinstance(value, Decimal) else value
                for index, value in enumerate(values) if value not in (None, "") and index < len(headers)
            }
            ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, sheet_name, raw_payload, ctx.loaded_at))
            for column in range(6, min(len(values), len(headers))):
                raw_amount = values[column]
                if raw_amount in (None, ""):
                    continue
                try:
                    signed_amount = reported_amount(raw_amount)
                except ValueError:
                    continue
                if signed_amount is None:
                    continue
                asset_id = stable_code(asset_ids[column] if column < len(asset_ids) else None) or f"RA_{stable_code(lines[column] if column < len(lines) else column)}"
                item_name = stable_code(headers[column]) or asset_id
                section = current_sections[column] if column < len(current_sections) else "Revenue Account"
                inferred_side = uk_side(f"{asset_id} {item_name}")
                side = "revenue" if signed_amount < 0 and inferred_side == "expenditure" else inferred_side
                function = f"{source['kind']}:{side}:{section}"
                economic = f"{source['kind']}:{side}:{asset_id}"
                ctx.bundle.node(node_row(country, classification_id, side, function, section, year, ctx.loaded_at))
                ctx.bundle.node(node_row(country, classification_id, side, economic, item_name, year, ctx.loaded_at))
                ctx.bundle.write("municipal_budget_line_facts", fact_row(
                    country, f"GB:{gss_code}", year, "enacted", side, function, economic,
                    abs(signed_amount) * Decimal(1000), currency, source, run_id, ctx.loaded_at,
                    classification_id, classification_id, row_number=row_number, sheet=sheet_name,
                    period=f"FY-{year}", scope="england_statistical_return",
                    summary=item_name.upper().startswith("TOTAL") or asset_id.lower().endswith(("tot", "total")),
                    quality_flags=["official_mhclg_revenue_account_budget", "source_unit_gbp_thousands", "reported_sign_normalized_to_budget_side"],
                ))
                rows_loaded += 1
        write_run(ctx, run_id, source, year, path, rows_read, rows_loaded)
        ctx.source_row(None, source, country, cfg["coverage"])
        total_rows += rows_read
        total_facts += rows_loaded
    return {"entities": len(selected), "rows": total_rows, "facts": total_facts, "fiscal_years": [year]}


SCOTLAND_COUNCILS = {
    "Aberdeen City": ("S12000033", "Aberdeen City"),
    "Aberdeenshire": ("S12000034", "Aberdeenshire"),
    "Angus": ("S12000041", "Angus"),
    "Argyll & Bute": ("S12000035", "Argyll and Bute"),
    "City of Edinburgh": ("S12000036", "City of Edinburgh"),
    "Clackmannanshire": ("S12000005", "Clackmannanshire"),
    "Dumfries & Galloway": ("S12000006", "Dumfries and Galloway"),
    "Dundee City": ("S12000042", "Dundee City"),
    "East Ayrshire": ("S12000008", "East Ayrshire"),
    "East Dunbartonshire": ("S12000045", "East Dunbartonshire"),
    "East Lothian": ("S12000010", "East Lothian"),
    "East Renfrewshire": ("S12000011", "East Renfrewshire"),
    "Falkirk": ("S12000014", "Falkirk"),
    "Fife": ("S12000047", "Fife"),
    "Glasgow City": ("S12000049", "Glasgow City"),
    "Highland": ("S12000017", "Highland"),
    "Inverclyde": ("S12000018", "Inverclyde"),
    "Midlothian": ("S12000019", "Midlothian"),
    "Moray": ("S12000020", "Moray"),
    "Na h-Eileanan Siar": ("S12000013", "Na h-Eileanan Siar"),
    "North Ayrshire": ("S12000021", "North Ayrshire"),
    "North Lanarkshire": ("S12000050", "North Lanarkshire"),
    "Orkney Islands": ("S12000023", "Orkney Islands"),
    "Perth & Kinross": ("S12000048", "Perth and Kinross"),
    "Renfrewshire": ("S12000038", "Renfrewshire"),
    "Scottish Borders": ("S12000026", "Scottish Borders"),
    "Shetland Islands": ("S12000027", "Shetland Islands"),
    "South Ayrshire": ("S12000028", "South Ayrshire"),
    "South Lanarkshire": ("S12000029", "South Lanarkshire"),
    "Stirling": ("S12000030", "Stirling"),
    "West Dunbartonshire": ("S12000039", "West Dunbartonshire"),
    "West Lothian": ("S12000040", "West Lothian"),
}


def _scotland_numeric(value: Any) -> Decimal | None:
    """Return a reported POBE cell value, or ``None`` where none was reported.

    ``decimal_value`` already maps the publisher's missing-value placeholders
    to ``None``, so a dashed cell reads as "no value" instead of as a zero that
    would make a section heading look like a populated data row.
    """
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        return decimal_value(value)
    except ValueError:
        return None


def run_scotland(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    """Load Scotland's council-level POBE revenue and capital workbooks.

    The source publishes 2025-26 provisional outturn and 2026-27 budget
    estimates for all 32 councils. The capital workbook also includes budget
    estimates for 2027-28 and 2028-29. Workbook totals and check rows are
    excluded so published facts remain additive at the retained native grain.
    """
    currency = cfg["currency"]
    sources = {source["kind"]: source for source in cfg["sources"] if source.get("collection") == "scotland"}
    if set(sources) != {"revenue", "capital"}:
        raise ValueError("Scotland POBE requires revenue and capital workbooks")
    selected = list(SCOTLAND_COUNCILS.items())[:ctx.args.max_entities or None]
    for _, (code, name) in selected:
        ctx.bundle.entity(entity_row(
            country, code, name, currency, ctx.loaded_at,
            code_type="GB_GSS_LOCAL_AUTHORITY", region_name="Scotland", entity_type="local_authority",
        ))

    total_facts = 0
    fiscal_years: set[int] = set()
    for kind, source in sources.items():
        path = ctx.download(country, source)
        workbook = load_workbook(path, read_only=False, data_only=True)
        function_class = f"GB_SCOTLAND_POBE_{kind.upper()}_FUNCTION_2026"
        economic_class = f"GB_SCOTLAND_POBE_{kind.upper()}_ROW_2026"
        ctx.bundle.classification(classification_row(
            country, function_class, "mixed", f"Scotland POBE {kind} sections", source["url"], 2025, ctx.loaded_at,
        ))
        ctx.bundle.classification(classification_row(
            country, economic_class, "mixed", f"Scotland POBE {kind} return rows", source["url"], 2025, ctx.loaded_at,
        ))
        measures = [(2025, "actual", 8), (2026, "enacted", 9)]
        if kind == "capital":
            measures.extend([(2027, "enacted", 10), (2028, "enacted", 11)])
        run_id = f"{source['id']}-v1"
        rows_read = rows_loaded = 0
        for sheet_name, (code, _) in selected:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Scotland POBE workbook missing council sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            current_section_code = "GENERAL"
            current_section_name = "General"
            for row_number in range(1, sheet.max_row + 1):
                label_cell = sheet.cell(row_number, 2)
                label = stable_code(label_cell.value)
                row_code = stable_code(sheet.cell(row_number, 4).value)
                values = [(year, stage, _scotland_numeric(sheet.cell(row_number, column).value)) for year, stage, column in measures]
                numeric_values = [(year, stage, amount) for year, stage, amount in values if amount is not None]
                if label_cell.font.bold and label and not numeric_values:
                    current_section_code = row_code or f"ROW_{row_number}"
                    current_section_name = label
                    continue
                if not label or not row_code or label_cell.font.bold or label.lower().startswith("check:") or not numeric_values:
                    continue
                rows_read += 1
                ctx.bundle.raw(raw_row(
                    country, 2026, source, run_id, row_number, sheet_name,
                    {"label": label, "row_code": row_code, **{f"{year}_{stage}": numeric_json(amount) for year, stage, amount in numeric_values}},
                    ctx.loaded_at,
                ))
                for year, stage, signed_amount in numeric_values:
                    if not signed_amount:
                        continue
                    side = "revenue" if signed_amount < 0 else "expenditure"
                    amount = abs(signed_amount) * Decimal(1000)
                    function = f"{kind.upper()}:{side}:{current_section_code}"
                    economic = f"{kind.upper()}:{side}:{row_code}"
                    ctx.bundle.node(node_row(country, function_class, side, function, current_section_name, year, ctx.loaded_at))
                    ctx.bundle.node(node_row(country, economic_class, side, economic, label, year, ctx.loaded_at))
                    flags = ["official_scottish_government_pobe", "source_unit_gbp_thousands", "reported_sign_normalized_to_budget_side"]
                    if stage == "actual":
                        flags.append("provisional_outturn")
                    ctx.bundle.write("municipal_budget_line_facts", fact_row(
                        country, f"GB:{code}", year, stage, side, function, economic, amount,
                        currency, source, run_id, ctx.loaded_at, function_class, economic_class,
                        row_number=row_number, sheet=sheet_name, period=f"FY-{year}",
                        scope="scotland_statistical_return", coverage_type="census", quality_flags=flags,
                    ))
                    fiscal_years.add(year)
                    rows_loaded += 1
        workbook.close()
        write_run(ctx, run_id, source, 2026, path, rows_read, rows_loaded)
        ctx.source_row(None, source, country, cfg["coverage"])
        total_facts += rows_loaded
    return {"entities": len(selected), "facts": total_facts, "fiscal_years": sorted(fiscal_years)}


def run_wales(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    """Load enacted revenue budgets for Wales's 22 unitary authorities."""
    source = next(source for source in cfg["sources"] if source.get("collection") == "wales")
    path = ctx.download(country, source)
    currency = cfg["currency"]
    classification_id = "GB_WALES_RA_SERVICE_2025_2026"
    ctx.bundle.classification(classification_row(
        country, classification_id, "mixed", "StatsWales revenue-budget service classification",
        source["url"], 2025, ctx.loaded_at,
    ))
    run_id = f"{source['id']}-v1"
    entities: dict[str, str] = {}
    rows_read = rows_loaded = 0
    fiscal_years: set[int] = set()
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row_number, row in enumerate(csv.DictReader(handle), 2):
            if stable_code(row.get("Data description_reference")) != "1":
                continue
            code = stable_code(row.get("Authority_reference"))
            if not code.startswith("W060"):
                continue
            if code not in entities:
                if ctx.args.max_entities and len(entities) >= ctx.args.max_entities:
                    continue
                entities[code] = stable_code(row.get("Authority")) or code
                ctx.bundle.entity(entity_row(
                    country, code, entities[code], currency, ctx.loaded_at,
                    code_type="GB_GSS_LOCAL_AUTHORITY", region_name="Wales", entity_type="local_authority",
                ))
            rows_read += 1
            year_label = stable_code(row.get("Year"))
            try:
                year = int(year_label[:4])
            except ValueError:
                continue
            signed_amount = reported_amount(row.get("Data values"))
            if signed_amount is None:
                continue
            service = stable_code(row.get("Service_reference")) or "UNSPECIFIED"
            service_name = stable_code(row.get("Service")) or service
            hierarchy = stable_code(row.get("Service_hierarchy")) or "GENERAL"
            side = "revenue" if signed_amount < 0 else "expenditure"
            function = f"WALES:{side}:{hierarchy}"
            economic = f"WALES:{side}:{service}"
            ctx.bundle.node(node_row(country, classification_id, side, function, f"StatsWales hierarchy {hierarchy}", year, ctx.loaded_at))
            ctx.bundle.node(node_row(country, classification_id, side, economic, service_name, year, ctx.loaded_at))
            ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, "StatsWales", row, ctx.loaded_at))
            ctx.bundle.write("municipal_budget_line_facts", fact_row(
                country, f"GB:{code}", year, "enacted", side, function, economic,
                abs(signed_amount) * Decimal(1000), currency, source, run_id, ctx.loaded_at,
                classification_id, classification_id, row_number=row_number, sheet="StatsWales",
                period=f"FY-{year}", scope="wales_statistical_return", coverage_type="census",
                summary=service_name.lower().startswith(("total ", "revenue expenditure")),
                quality_flags=["official_statswales_ra", "source_unit_gbp_thousands", "reported_sign_normalized_to_budget_side"],
            ))
            fiscal_years.add(year)
            rows_loaded += 1
    write_run(ctx, run_id, source, 2026, path, rows_read, rows_loaded)
    ctx.source_row(None, source, country, cfg["coverage"])
    return {"entities": len(entities), "facts": rows_loaded, "fiscal_years": sorted(fiscal_years)}


def run_united_kingdom(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    collections = {source.get("collection") for source in cfg["sources"]}
    results: dict[str, Any] = {}
    if "england" in collections:
        england_cfg = dict(cfg)
        england_cfg["sources"] = [source for source in cfg["sources"] if source.get("collection") == "england"]
        results["england"] = run_england(ctx, country, england_cfg)
    if "england_budget" in collections:
        results["england_budget"] = run_england_budget(ctx, country, cfg)
    if "scotland" in collections:
        results["scotland"] = run_scotland(ctx, country, cfg)
    if "wales" in collections:
        results["wales"] = run_wales(ctx, country, cfg)
    results["entities"] = sum(int(result.get("entities", 0)) for result in results.values() if isinstance(result, dict))
    return results


def find_record_list(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list) and (not payload or isinstance(payload[0], dict)):
        return payload
    if isinstance(payload, dict):
        for key in ("data", "results", "items", "records", "content"):
            if key in payload:
                found = find_record_list(payload[key])
                if found:
                    return found
    return []


def first_value(row: dict[str, Any], aliases: Iterable[str]) -> Any:
    lowered = {key.lower(): value for key, value in row.items()}
    for alias in aliases:
        if alias.lower() in lowered and lowered[alias.lower()] not in (None, ""):
            return lowered[alias.lower()]
    return None


def run_ukraine(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    restore_raw_cache(ctx.args.cache_dir / country)
    year, currency = cfg["year"], cfg["currency"]
    source = cfg["sources"][0]
    path = ctx.download(country, source, optional=True)
    if path is None:
        message = f"Open Budget source unavailable; provide --openbudget-token or --source-file {source['id']}=/path/snapshot.json"
        ctx.bundle.write("ingestion_runs", {
            "ingestion_run_id": f"{source['id']}-v1", "source_id": source["id"], "started_at": ctx.loaded_at,
            "completed_at": ctx.loaded_at, "status": "blocked", "source_vintage": str(year), "source_sha256": None,
            "rows_read": 0, "rows_loaded": 0, "warning_count": 1, "error_message": message,
        })
        ctx.source_row(None, source, country, message)
        return {"entities": 0, "blocked": message}
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    records = find_record_list(payload)
    function_class, economic_class = f"UA_LOCAL_FUNCTION_{year}", f"UA_LOCAL_ECONOMIC_{year}"
    ctx.bundle.classification(classification_row(country, function_class, "mixed", "Ukraine local budget programme/function", source["url"], year, ctx.loaded_at))
    ctx.bundle.classification(classification_row(country, economic_class, "mixed", "Ukraine local budget economic classification", source["url"], year, ctx.loaded_at))
    run_id = f"{source['id']}-v1"
    selected: set[str] = set()
    rows_loaded = 0
    for row_number, row in enumerate(records, 1):
        code = stable_code(first_value(row, ("local_budget_code", "budget_code", "budgetCode", "kod_biudzhetu")))
        if not code:
            continue
        if code not in selected:
            if ctx.args.max_entities and len(selected) >= ctx.args.max_entities:
                continue
            selected.add(code)
            name = stable_code(first_value(row, ("local_budget_name", "budget_name", "budgetName", "name"))) or code
            ctx.bundle.entity(entity_row(country, code, name, currency, ctx.loaded_at, code_type="UA_LOCAL_BUDGET_CODE"))
        function = stable_code(first_value(row, ("program_code", "functional_code", "function_code", "kpkvk"))) or "UNSPECIFIED"
        economic = stable_code(first_value(row, ("economic_code", "kekv", "revenue_code", "classification_code"))) or "UNSPECIFIED"
        side_text = stable_code(first_value(row, ("budget_side", "type", "direction"))).lower()
        side = "revenue" if any(word in side_text for word in ("income", "revenue", "дох")) else "expenditure"
        ctx.bundle.node(node_row(country, function_class, side, function, stable_code(first_value(row, ("program_name", "function_name"))) or function, year, ctx.loaded_at))
        ctx.bundle.node(node_row(country, economic_class, side, economic, stable_code(first_value(row, ("economic_name", "classification_name"))) or economic, year, ctx.loaded_at))
        ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, "OpenBudget API", row, ctx.loaded_at))
        for stage, aliases in (
            ("enacted", ("initial_plan", "approved_plan", "plan")),
            ("revised", ("revised_plan", "adjusted_plan", "plan_updated")),
            ("actual", ("executed", "actual", "fact", "cash_execution")),
        ):
            amount = reported_amount(first_value(row, aliases))
            if amount is None:
                continue
            ctx.bundle.write("municipal_budget_line_facts", fact_row(
                country, f"UA:{code}", year, stage, side, function, economic, amount,
                currency, source, run_id, ctx.loaded_at, function_class, economic_class,
                row_number=row_number, sheet="OpenBudget API", quality_flags=["api_schema_alias_normalized"],
            ))
            rows_loaded += 1
    write_run(ctx, run_id, source, year, path, len(records), rows_loaded)
    ctx.source_row(None, source, country, cfg["coverage"])
    return {"entities": len(selected)}


def _ukraine_budget_directory(path: Path, year: int) -> list[dict[str, Any]]:
    records = json.loads(path.read_text(encoding="utf-8-sig"))
    start, end = f"{year}-01-01", f"{year}-12-31"
    latest: dict[str, dict[str, Any]] = {}
    for row in records:
        code = stable_code(row.get("codebudg"))
        if not code or row.get("details") != 1:
            continue
        if stable_code(row.get("beginDate")) > end:
            continue
        if row.get("endDate") and stable_code(row["endDate"]) < start:
            continue
        previous = latest.get(code)
        if previous is None or stable_code(row.get("beginDate")) > stable_code(previous.get("beginDate")):
            latest[code] = row
    # Territorial-community budgets are the comparable municipal tier. Kyiv is
    # a city with special status and does not use the phrase in its directory name.
    return sorted([
        row for code, row in latest.items()
        if "територіальної громади" in stable_code(row.get("namebudg")).lower() or code == "2600000000"
    ], key=lambda row: row["codebudg"])


def _ukraine_download_csv(ctx: Context, country: str, source: dict[str, Any], year: int, code: str) -> Path:
    cache = ctx.args.cache_dir / country / str(year) / source["kind"] / f"{code}.csv.gz"
    if cache.exists() and not ctx.args.refresh:
        return cache
    if ctx.args.offline:
        raise FileNotFoundError(f"Offline source missing: {cache}")
    params = {"budgetCode": code, "budgetItem": source["budget_item"], "period": "QUARTER", "year": year}
    if source.get("classification_type"):
        params["classificationType"] = source["classification_type"]
    response = None
    for attempt in range(5):
        response = ctx.session.get(source["url"], params=params, timeout=240)
        response.raise_for_status()
        if "text/csv" in response.headers.get("content-type", ""):
            break
        if attempt < 4:
            time.sleep(2 ** attempt)
    assert response is not None
    if "text/csv" not in response.headers.get("content-type", ""):
        raise ValueError(f"Ukraine API returned {response.headers.get('content-type')} for {code}: {response.text[:160]}")
    cache.parent.mkdir(parents=True, exist_ok=True)
    temporary = cache.with_suffix(cache.suffix + ".part")
    with gzip.open(temporary, "wb") as handle:
        handle.write(response.content)
    temporary.replace(cache)
    return cache


def run_ukraine_public_api(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    restore_raw_cache(ctx.args.cache_dir / country)
    year, currency = cfg["year"], cfg["currency"]
    sources = {source["kind"]: source for source in cfg["sources"]}
    directory_source = sources["entities"]
    directory_path = ctx.download(country, directory_source)
    municipalities = _ukraine_budget_directory(directory_path, year)[:ctx.args.max_entities or None]
    for municipality in municipalities:
        code = municipality["codebudg"]
        ctx.bundle.entity(entity_row(
            country, code, stable_code(municipality.get("namebudg")) or code, currency, ctx.loaded_at,
            code_type="UA_LOCAL_BUDGET_CODE", region_code=municipality.get("codeRegion"),
            geography_code=municipality.get("katottg"), geography_type="UA_KATOTTG",
        ))
    ctx.source_row(None, directory_source, country, cfg["coverage"])

    failures: list[str] = []
    total_loaded = 0
    for side in ("expenditure", "revenue"):
        source = sources[side]
        run_id = f"{source['id']}-{year}-v1"
        function_class = f"UA_LOCAL_FUNCTION_{year}" if side == "expenditure" else None
        economic_class = f"UA_LOCAL_{side.upper()}_{year}"
        if function_class:
            ctx.bundle.classification(classification_row(country, function_class, side, "Ukraine local budget programme and function", source["url"], year, ctx.loaded_at))
        ctx.bundle.classification(classification_row(country, economic_class, side, f"Ukraine local budget {side} classification", source["url"], year, ctx.loaded_at))
        rows_read = rows_loaded = 0
        with ThreadPoolExecutor(max_workers=ctx.args.api_workers) as executor:
            futures = {
                executor.submit(_ukraine_download_csv, ctx, country, source, year, municipality["codebudg"]): municipality
                for municipality in municipalities
            }
            for future in as_completed(futures):
                municipality = futures[future]
                code = municipality["codebudg"]
                try:
                    path = future.result()
                    with gzip.open(path, "rt", encoding="utf-8-sig", newline="") as handle:
                        reader = csv.DictReader(handle, delimiter=";")
                        for row_number, row in enumerate(reader, 2):
                            if stable_code(row.get("REP_PERIOD")) != f"12.{year}":
                                continue
                            rows_read += 1
                            function = stable_code(row.get("COD_CONS_MB_FK")) or None
                            economic = stable_code(row.get("COD_CONS_EK") or row.get("COD_INCO"))
                            if not economic:
                                continue
                            if function_class and function:
                                ctx.bundle.node(node_row(country, function_class, side, function, stable_code(row.get("COD_CONS_MB_FK_NAME")) or function, year, ctx.loaded_at))
                            ctx.bundle.node(node_row(
                                country, economic_class, side, economic,
                                stable_code(row.get("COD_CONS_EK_NAME") or row.get("NAME_INC")) or economic,
                                year, ctx.loaded_at,
                            ))
                            ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, f"{code}.csv", row, ctx.loaded_at))
                            for stage, column in (("enacted", "ZAT_AMT"), ("revised", "PLANS_AMT"), ("actual", "FAKT_AMT")):
                                amount = reported_amount(row.get(column))
                                if amount is None:
                                    continue
                                ctx.bundle.write("municipal_budget_line_facts", fact_row(
                                    country, f"UA:{code}", year, stage, side, function, economic, amount,
                                    currency, source, run_id, ctx.loaded_at, function_class, economic_class,
                                    row_number=row_number, sheet=f"{code}.csv", item_type=stable_code(row.get("FUND_TYP")),
                                    quality_flags=["year_end_cumulative_quarter", "official_openbudget_public_api"],
                                ))
                                rows_loaded += 1
                except Exception as exc:
                    failures.append(f"{code}:{side}:{type(exc).__name__}:{exc}")
        write_run(ctx, run_id, source, year, None, rows_read, rows_loaded)
        ctx.source_row(None, source, country, cfg["coverage"])
        total_loaded += rows_loaded
    if failures:
        raise RuntimeError(f"Ukraine API failed for {len(failures)} municipality/source requests; first: {failures[0]}")
    return {"entities": len(municipalities), "facts": total_loaded, "api_requests": len(municipalities) * 2}


def write_run(ctx: Context, run_id: str, source: dict[str, Any], year: int, path: Path | None, rows_read: int, rows_loaded: int) -> None:
    ctx.bundle.write("ingestion_runs", {
        "ingestion_run_id": run_id, "source_id": source["id"], "started_at": ctx.loaded_at,
        "completed_at": utc_now(), "status": "completed", "source_vintage": str(year),
        "source_sha256": sha256(path) if path and path.exists() else None,
        "rows_read": rows_read, "rows_loaded": rows_loaded, "warning_count": 0, "error_message": None,
    })


def _socrata_pages(ctx: Context, source: dict[str, Any], where: str | None = None) -> Iterator[dict[str, Any]]:
    offset, limit = 0, 50000
    while True:
        params: dict[str, Any] = {"$limit": limit, "$offset": offset}
        if where:
            params["$where"] = where
        rows = ctx.api_json("GET", source["url"], params=params)
        if not rows:
            break
        yield from rows
        if len(rows) < limit:
            break
        offset += limit


def run_us_socrata(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    currency = cfg["currency"]
    sources = cfg["sources"][:ctx.args.max_entities or None]
    total_rows = total_facts = 0
    for source in sources:
        city_code, city_name = source["city_code"], source["city_name"]
        entity_id = f"US:{city_code}"
        ctx.bundle.entity(entity_row(country, city_code, city_name, currency, ctx.loaded_at, code_type="US_CITY_PORTAL_CODE"))
        function_class = f"US_{city_code}_FUNCTION"
        economic_class = f"US_{city_code}_ECONOMIC"
        ctx.bundle.classification(classification_row(country, function_class, "mixed", f"{city_name} department/program classification", source["url"], 2026, ctx.loaded_at))
        ctx.bundle.classification(classification_row(country, economic_class, "mixed", f"{city_name} appropriation/account classification", source["url"], 2026, ctx.loaded_at))
        run_id = f"{source['id']}-v1"
        rows_read = rows_loaded = 0
        where = None
        if city_code == "NYC":
            # Financial-plan rows are republished as snapshots. Keep only the
            # newest publication for each fiscal year so a budget line is not
            # counted repeatedly merely because the plan was refreshed.
            publications = ctx.api_json("GET", source["url"], params={
                "$select": "fiscal_year,max(publication_date) as max_date",
                "$where": "fiscal_year >= 2026",
                "$group": "fiscal_year",
            })
            where = " OR ".join(
                f"(fiscal_year = {int(item['fiscal_year'])} AND publication_date = '{item['max_date']}')"
                for item in publications
            )
        elif city_code == "LA":
            where = "budget_fiscal_year >= '2026'"
        elif city_code == "SF":
            where = "fiscal_year >= '2026'"
        for row_number, row in enumerate(_socrata_pages(ctx, source, where), 1):
            rows_read += 1
            if city_code == "NYC":
                year = int(row.get("fiscal_year") or 2026)
                function = stable_code(row.get("agency_number")) or "UNSPECIFIED"
                function_name = stable_code(row.get("agency_name")) or function
                economic = stable_code(row.get("unit_appropriation_number")) or "UNSPECIFIED"
                economic_name = stable_code(row.get("unit_appropriation_name")) or economic
                measures = (("proposal", "total_financial_plan_amount"), ("enacted", "total_adopted_budget_amount"), ("revised", "total_current_budget_amount"))
                side = "expenditure"
            elif city_code == "CHI":
                year = 2026
                function = stable_code(row.get("department_number")) or "UNSPECIFIED"
                function_name = stable_code(row.get("department_description")) or function
                economic = stable_code(row.get("appropriation_account")) or "UNSPECIFIED"
                economic_name = stable_code(row.get("appropriation_account_description")) or economic
                measures = (("enacted", "_ordinance_amount_"),)
                side = "expenditure"
            elif city_code == "LA":
                year = int(row.get("budget_fiscal_year") or 2026)
                function = stable_code(row.get("department")) or "UNSPECIFIED"
                function_name = stable_code(row.get("department_name")) or function
                economic = stable_code(row.get("account")) or "UNSPECIFIED"
                economic_name = stable_code(row.get("account_name")) or economic
                measures = (("enacted", "adopted_budget_amount"), ("revised", "total_budget"), ("actual", "actual_expense"))
                side = "expenditure"
            else:
                year = int(row.get("fiscal_year") or 2026)
                function = ":".join(filter(None, (stable_code(row.get("department_code")), stable_code(row.get("program_code"))))) or "UNSPECIFIED"
                function_name = " · ".join(filter(None, (stable_code(row.get("department")), stable_code(row.get("program"))))) or function
                economic = stable_code(row.get("sub_object_code") or row.get("object_code")) or "UNSPECIFIED"
                economic_name = stable_code(row.get("sub_object") or row.get("object")) or economic
                measures = (("enacted", "budget"),)
                side = "revenue" if stable_code(row.get("revenue_or_spending")).lower() == "revenue" else "expenditure"
            ctx.bundle.node(node_row(country, function_class, side, function, function_name, year, ctx.loaded_at))
            ctx.bundle.node(node_row(country, economic_class, side, economic, economic_name, year, ctx.loaded_at))
            ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, source["dataset"], row, ctx.loaded_at))
            for stage, field in measures:
                amount = reported_amount(row.get(field))
                if amount is None:
                    continue
                ctx.bundle.write("municipal_budget_line_facts", fact_row(
                    country, entity_id, year, stage, side, function, economic, amount,
                    currency, source, run_id, ctx.loaded_at, function_class, economic_class,
                    row_number=row_number, sheet=source["dataset"], coverage_type="published_subset",
                    quality_flags=["official_city_open_data", "non_national_coverage"],
                ))
                rows_loaded += 1
        write_run(ctx, run_id, source, cfg["year"], None, rows_read, rows_loaded)
        ctx.source_row(entity_id, source, country, cfg["coverage"])
        total_rows += rows_read
        total_facts += rows_loaded
    return {"entities": len(sources), "rows": total_rows, "facts": total_facts}


def run_switzerland_lucerne(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    year, currency = cfg["year"], cfg["currency"]
    source = next(source for source in cfg["sources"] if source.get("collection") in (None, "lucerne"))
    path = ctx.download(country, source)
    function_class = f"CH_LU_HRM2_FUNCTION_{year}"
    economic_class = f"CH_LU_HRM2_ACCOUNT_{year}"
    ctx.bundle.classification(classification_row(country, function_class, "mixed", "Canton Lucerne HRM2 function", source["url"], year, ctx.loaded_at))
    ctx.bundle.classification(classification_row(country, economic_class, "mixed", "Canton Lucerne HRM2 account type", source["url"], year, ctx.loaded_at))
    run_id = f"{source['id']}-{year}-v1"
    entities: set[str] = set()
    rows_read = rows_loaded = 0
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row_number, row in enumerate(csv.DictReader(handle, delimiter=";"), 2):
            if stable_code(row.get("jahr")) != str(year):
                continue
            code = stable_code(row.get("gnr"))
            if not code:
                continue
            if ctx.args.max_entities and code not in entities and len(entities) >= ctx.args.max_entities:
                continue
            entities.add(code)
            entity_id = f"CH:{code}"
            ctx.bundle.entity(entity_row(country, code, stable_code(row.get("gemeinde")) or code, currency, ctx.loaded_at, code_type="CH_BFS_MUNICIPALITY_NUMBER", region_name="Luzern"))
            rows_read += 1
            function = stable_code(row.get("fkt_nr")) or "UNSPECIFIED"
            economic = stable_code(row.get("art_nr")) or "UNSPECIFIED"
            first = economic[:1]
            side = "expenditure" if first in {"3", "5"} else "revenue" if first in {"4", "6"} else None
            if side is None:
                continue
            amount = reported_amount(row.get("saldo"))
            if amount is None:
                continue
            amount = abs(amount)
            ctx.bundle.node(node_row(country, function_class, side, function, stable_code(row.get("fkt_name")) or function, year, ctx.loaded_at))
            ctx.bundle.node(node_row(country, economic_class, side, economic, stable_code(row.get("art_name")) or economic, year, ctx.loaded_at))
            ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, "gefis-lu-bg.csv", row, ctx.loaded_at))
            ctx.bundle.write("municipal_budget_line_facts", fact_row(
                country, entity_id, year, "enacted", side, function, economic, amount,
                currency, source, run_id, ctx.loaded_at, function_class, economic_class,
                row_number=row_number, sheet="gefis-lu-bg.csv", coverage_type="published_subset",
                quality_flags=["official_lustat_hrm2", "reported_sign_normalized_to_budget_side"],
            ))
            rows_loaded += 1
    write_run(ctx, run_id, source, year, path, rows_read, rows_loaded)
    ctx.source_row(None, source, country, cfg["coverage"])
    return {"entities": len(entities), "facts": rows_loaded}


def run_switzerland_zurich(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    year, currency = cfg["year"], cfg["currency"]
    source = next(source for source in cfg["sources"] if source.get("collection") == "zurich")
    path = ctx.cache_path(country, source)
    if path.exists() and not ctx.args.refresh:
        records = json.loads(path.read_text(encoding="utf-8"))
    elif ctx.args.offline:
        raise FileNotFoundError(f"Offline source missing: {path}")
    else:
        headers = {"api-key": source_api_key(source)}
        departments = ctx.api_json("GET", f"{source['url']}/departemente", headers=headers).get("value", [])
        records = []
        for department in departments:
            payload = ctx.api_json("GET", f"{source['url']}/budgetbuch", headers=headers, params={"orgKey": department["key"], "jahr": year})
            records.extend(payload.get("value", []))
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(records, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    code, entity_id = "0261", "CH:0261"
    ctx.bundle.entity(entity_row(country, code, "Zürich", currency, ctx.loaded_at, code_type="CH_BFS_MUNICIPALITY_NUMBER", region_name="Zürich"))
    function_class = f"CH_ZH_ORGANISATION_{year}"
    economic_class = f"CH_ZH_HRM2_ACCOUNT_{year}"
    ctx.bundle.classification(classification_row(country, function_class, "mixed", "Zürich department and institution", source["url"], year, ctx.loaded_at))
    ctx.bundle.classification(classification_row(country, economic_class, "mixed", "Zürich HRM2 account", source["url"], year, ctx.loaded_at))
    run_id = f"{source['id']}-v1"; rows_loaded = 0
    for row_number, row in enumerate(records, 1):
        account = row.get("konto") or {}; institution = account.get("institution") or {}; department = institution.get("departement") or {}
        economic = stable_code(account.get("kontoNr")).replace(" ", "")
        first = economic[:1]
        side = "expenditure" if first in {"3", "5"} else "revenue" if first in {"4", "6"} else None
        amount = reported_amount(row.get("budgetAktuell"))
        if not economic or side is None or amount is None:
            continue
        function = ":".join(filter(None, (stable_code(department.get("key")), stable_code(institution.get("key"))))) or "UNSPECIFIED"
        function_name = " · ".join(filter(None, (stable_code(department.get("bezeichnung")), stable_code(institution.get("bezeichnung"))))) or function
        ctx.bundle.node(node_row(country, function_class, side, f"{side}:{function}", function_name, year, ctx.loaded_at))
        ctx.bundle.node(node_row(country, economic_class, side, f"{side}:{economic}", stable_code(account.get("bezeichnung")) or economic, year, ctx.loaded_at))
        ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, "RPK budgetbuch API", row, ctx.loaded_at))
        ctx.bundle.write("municipal_budget_line_facts", fact_row(
            country, entity_id, year, "enacted", side, f"{side}:{function}", f"{side}:{economic}", abs(amount),
            currency, source, run_id, ctx.loaded_at, function_class, economic_class,
            row_number=row_number, sheet="RPK budgetbuch API", coverage_type="published_subset",
            quality_flags=["official_zurich_finance_api", "live_api_snapshot", "hrm2_account_detail"],
        ))
        rows_loaded += 1
    write_run(ctx, run_id, source, year, path, len(records), rows_loaded)
    ctx.source_row(entity_id, source, country, cfg["coverage"])
    return {"entities": 1, "facts": rows_loaded, "fiscal_years": [year]}


def run_switzerland_structured(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    lucerne = run_switzerland_lucerne(ctx, country, cfg)
    zurich = run_switzerland_zurich(ctx, country, cfg)
    return {"lucerne": lucerne, "zurich": zurich, "entities": int(lucerne.get("entities", 0)) + 1}


def run_germany_bremen(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    """Load the machine-readable enacted budget of the Stadtgemeinde Bremen.

    Germany has no single open municipal itemized-budget census. This adapter
    deliberately publishes one official city scatter and labels it as such.
    """
    currency = cfg["currency"]
    source = cfg["sources"][0]
    path = ctx.download(country, source)
    entity_code = "04011000"
    entity_id = f"DE:{entity_code}"
    ctx.bundle.entity(entity_row(
        country, entity_code, "Bremen", currency, ctx.loaded_at,
        code_type="DE_AGS", region_name="Bremen",
    ))
    function_class = "DE_HB_PRODUCT_GROUP_2026_2027"
    economic_class = "DE_HB_BUDGET_POSITION_2026_2027"
    ctx.bundle.classification(classification_row(country, function_class, "mixed", "Bremen product-group budget classification", source["url"], 2026, ctx.loaded_at))
    ctx.bundle.classification(classification_row(country, economic_class, "mixed", "Bremen kameral budget-position classification", source["url"], 2026, ctx.loaded_at))
    run_id = f"{source['id']}-v1"
    rows_read = rows_loaded = 0
    with path.open(encoding="latin-1", newline="") as handle:
        for row_number, row in enumerate(csv.DictReader(handle, delimiter=";"), 2):
            rows_read += 1
            aggregate = stable_code(row.get("Aggregat"))
            side = "expenditure" if aggregate.startswith("AUSG.") else "revenue" if aggregate.startswith("EINN.") else None
            if side is None:
                continue
            function = stable_code(row.get("PGR")) or "UNSPECIFIED"
            economic = stable_code(row.get("Haushaltsstelle")) or "UNSPECIFIED"
            purpose = stable_code(row.get("Zweckbestimmung")) or economic
            ctx.bundle.node(node_row(country, function_class, side, function, f"Produktgruppe {function}", 2026, ctx.loaded_at))
            ctx.bundle.node(node_row(country, economic_class, side, economic, purpose, 2026, ctx.loaded_at))
            for year in (2026, 2027):
                raw_amount = stable_code(row.get(f"Anschlag  {year}"))
                if not raw_amount:
                    continue
                amount = Decimal(raw_amount.replace(".", "").replace(",", "."))
                if not amount:
                    continue
                ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, "Stadtgemeinde Bremen", row, ctx.loaded_at))
                ctx.bundle.write("municipal_budget_line_facts", fact_row(
                    country, entity_id, year, "enacted", side, function, economic, amount,
                    currency, source, run_id, ctx.loaded_at, function_class, economic_class,
                    row_number=row_number, sheet="Stadtgemeinde Bremen", coverage_type="published_subset",
                    quality_flags=["official_bremen_city_budget", "non_national_coverage"],
                ))
                rows_loaded += 1
    write_run(ctx, run_id, source, 2026, path, rows_read, rows_loaded)
    ctx.source_row(entity_id, source, country, cfg["coverage"])
    return {"entities": 1, "rows": rows_read, "facts": rows_loaded, "fiscal_years": [2026, 2027]}


def german_decimal(value: Any) -> Decimal | None:
    """Parse a German/French municipal amount written in the de-DE convention.

    Returns ``None`` when the cell reports no value. ``[z]`` is the official
    "not available" marker in these publications, and a cell that cannot be
    parsed is an unknown amount too -- neither is a zero, so neither is
    reported as one.
    """
    text = stable_code(value).replace("€", "").replace("\u00a0", "").replace(" ", "")
    if not text or text in MISSING_VALUE_TOKENS or text in {"[z]", "[-]"}:
        return None
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return decimal_value(text)
    except ValueError:
        return None


def _de_split_code(value: Any) -> tuple[str, str]:
    text = stable_code(value)
    if " - " in text:
        code, name = text.split(" - ", 1)
        return code.strip(), name.strip()
    return text or "UNSPECIFIED", text or "UNSPECIFIED"


def run_germany_structured(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    """Load verified official machine-readable city-budget publications.

    German municipal publication is decentralized. This adapter intentionally
    enumerates official city sources and preserves their heterogeneous account,
    product, cost-centre and plan-year fields. It is a published subset, never a
    national census claim.
    """
    currency = cfg["currency"]
    sources = cfg["sources"][:ctx.args.max_entities or None]
    total_rows = total_facts = 0
    fiscal_years: set[int] = set()
    for source in sources:
        if source.get("parser") == "bremen":
            result = run_germany_bremen(ctx, country, {**cfg, "sources": [source]})
            total_rows += int(result.get("rows", 0)); total_facts += int(result.get("facts", 0))
            fiscal_years.update(result.get("fiscal_years", []))
            continue
        path = ctx.download(country, source)
        parser = source["parser"]
        code, name = source["city_code"], source["city_name"]
        entity_id = f"DE:{code}"
        ctx.bundle.entity(entity_row(
            country, code, name, currency, ctx.loaded_at,
            code_type="DE_AGS", region_name=source.get("region"), entity_type="municipality",
        ))
        function_class = f"DE_{code}_FUNCTION"
        economic_class = f"DE_{code}_ECONOMIC"
        ctx.bundle.classification(classification_row(country, function_class, "mixed", f"{name} product/function classification", source["url"], 2016, ctx.loaded_at))
        ctx.bundle.classification(classification_row(country, economic_class, "mixed", f"{name} account/plan-line classification", source["url"], 2016, ctx.loaded_at))
        run_id = f"{source['id']}-v1"
        rows_read = rows_loaded = 0

        def emit(row_number: int, sheet: str, payload: dict[str, Any], year: int, stage: str, side: str,
                 function: str, function_name: str, economic: str, economic_name: str,
                 amount: Decimal | None) -> None:
            nonlocal rows_loaded
            # ``None`` is a missing source value and ``Decimal(0)`` a reported
            # zero; neither becomes a fact, and neither is written as the other.
            if amount is None or not amount:
                return
            function = function or "UNSPECIFIED"; economic = economic or "UNSPECIFIED"
            ctx.bundle.node(node_row(country, function_class, side, f"{side}:{function}", function_name or function, year, ctx.loaded_at))
            ctx.bundle.node(node_row(country, economic_class, side, f"{side}:{economic}", economic_name or economic, year, ctx.loaded_at))
            ctx.bundle.write("municipal_budget_line_facts", fact_row(
                country, entity_id, year, stage, side, f"{side}:{function}", f"{side}:{economic}", abs(amount),
                currency, source, run_id, ctx.loaded_at, function_class, economic_class,
                row_number=row_number, sheet=sheet, coverage_type="published_subset",
                quality_flags=["official_municipal_budget_source", "germany_decentralized_publication", "reported_sign_normalized_to_budget_side"],
            ))
            fiscal_years.add(year); rows_loaded += 1

        if parser == "aachen":
            text = path.read_bytes().decode("cp1252")
            rows = list(csv.reader(text.splitlines(), delimiter=","))
            years = [int(stable_code(value)[:4]) for value in rows[4][2:] if stable_code(value)[:4].isdigit()]
            for row_number, values in enumerate(rows[6:], 7):
                if len(values) < 3:
                    continue
                function, function_name = _de_split_code(values[0]); economic, economic_name = _de_split_code(values[1])
                rows_read += 1
                payload = {"function": values[0], "account": values[1], **{str(year): values[index + 2] for index, year in enumerate(years) if index + 2 < len(values)}}
                ctx.bundle.raw(raw_row(country, years[0], source, run_id, row_number, "Ergebnisplanung", payload, ctx.loaded_at))
                for index, year in enumerate(years):
                    amount = german_decimal(values[index + 2] if index + 2 < len(values) else None)
                    side = "revenue" if economic.startswith("4") else "expenditure"
                    emit(row_number, "Ergebnisplanung", payload, year, "enacted" if year == years[0] else "proposal", side, function, function_name, economic, economic_name, amount)
        elif parser == "oldenburg":
            workbook = load_workbook(path, read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                headers = [stable_code(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
                for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                    row = dict(zip(headers, values)); function = stable_code(row.get("Teilhaushalt")); function_name = stable_code(row.get("Bezeichnung"))
                    if not function:
                        continue
                    rows_read += 1; ctx.bundle.raw(raw_row(country, 2025, source, run_id, row_number, sheet.title, row, ctx.loaded_at))
                    for field, side in ((headers[2], "revenue"), (headers[3], "expenditure")):
                        emit(row_number, sheet.title, row, 2025, "enacted", side, function, function_name, field, field, german_decimal(row.get(field)))
            workbook.close()
        else:
            encoding = "utf-8-sig" if parser in {"dortmund", "frankfurt"} else "cp1252"
            delimiter = ";"
            if parser == "bonn":
                archive = zipfile.ZipFile(path)
                member = next(item for item in archive.namelist() if item.endswith(source["member"]))
                stream = io.TextIOWrapper(archive.open(member), encoding=encoding, newline="")
                reader = csv.DictReader(stream, delimiter=delimiter)
                sheet = member
            else:
                stream = path.open("r", encoding=encoding, newline="")
                reader = csv.DictReader(stream, delimiter=delimiter)
                sheet = path.name
            try:
                for row_number, row in enumerate(reader, 2):
                    rows_read += 1
                    measures: list[tuple[int, str, Any]] = []
                    if parser == "bonn":
                        year = int(stable_code(row.get("Geschäftsjahr")) or 2024)
                        function = stable_code(row.get("Profitcenter") or row.get("Kostenstelle") or row.get("PSP-Element")); function_name = function
                        economic = stable_code(row.get("Kontonummer")); economic_name = stable_code(row.get("Bezeichnung")) or economic
                        side = "revenue" if economic.startswith("4") else "expenditure"
                        measures = [(year, "enacted", row.get("in Profit-Center-Hauswährung"))]
                    elif parser == "dortmund":
                        function = stable_code(row.get("Produktbereich") or row.get("Organisationseinheit")); function_name = stable_code(row.get("Bezeichnung Produktbereich") or row.get("Bezeichnung der Organisationseinheit")) or function
                        economic = stable_code(row.get("Zeile des Gesamt-/Teilplans")); economic_name = stable_code(row.get("Art")) or economic
                        descriptor = f"{row.get('Untergruppe', '')} {economic_name}".lower(); side = "revenue" if "ertr" in descriptor or "einzahl" in descriptor else "expenditure"
                        for field, value in row.items():
                            match = re.search(r"(Haushaltsansatz|Planung).*?(20\d{2})", field)
                            if match: measures.append((int(match.group(2)), "enacted" if match.group(1) == "Haushaltsansatz" else "proposal", value))
                    elif parser == "essen":
                        function, function_name = _de_split_code(row.get("Zeile_Ergebnisplan")); economic, economic_name = _de_split_code(row.get("Kostenart")); side = "revenue" if economic.startswith("4") else "expenditure"
                        for field, value in row.items():
                            match = re.match(r"(Ansatz|Planung)_(20\d{2})", field)
                            if match: measures.append((int(match.group(2)), "enacted" if match.group(1) == "Ansatz" else "proposal", value))
                    elif parser == "frankfurt":
                        function = stable_code(row.get("Produktgruppe") or row.get("Produktbereich")); function_name = stable_code(row.get("Produktgruppe Bezeichnung") or row.get("Produktbereich Bezeichnung")) or function
                        economic = stable_code(row.get("Nummer")); economic_name = stable_code(row.get("Gruppierung Bezeichnung")) or economic
                        side = "revenue" if stable_code(row.get("Ertrag / Aufwand")).upper().startswith("E") else "expenditure"
                        year = int(stable_code(row.get("Haushaltsjahr"))); measures = [(year, "enacted" if year in {2024, 2025} else "proposal", row.get("€"))]
                    elif parser == "karlsruhe":
                        function = stable_code(row.get("KOSTENSTELLE") or row.get("PSP-ELEMENT") or row.get("PROFITCENTER")); function_name = stable_code(row.get("KOSTENSTELLE BEZ.") or row.get("PSP-ELEMENT-BEZ.") or row.get("PROFITCENTER BEZ.")) or function
                        economic = stable_code(row.get("KOSTENART")); economic_name = stable_code(row.get("KOSTENART BEZ.")) or economic
                        side = "revenue" if stable_code(row.get("E/A")) == "E" or (not stable_code(row.get("E/A")) and economic.startswith("3")) else "expenditure"
                        base = int(stable_code(row.get("GESCH.JAHR"))); measures = [(base, "enacted", row.get("PLANZAHL LFD. JAHR"))] + [(base + offset, "proposal", row.get(f"PLANZAHL FOLGEJAHR {offset}")) for offset in range(1, 5)]
                    elif parser == "koeln":
                        function = stable_code(row.get("TPLAN") or row.get("AMT") or row.get("FISTL")); function_name = stable_code(row.get("TP_BEZ") or row.get("AMT_BEZ") or row.get("FISTL_BEZ")) or function
                        economic = stable_code(row.get("FIPOS") or row.get("FKONTO")); economic_name = stable_code(row.get("FIPOS_BEZ")) or economic
                        side = "revenue" if stable_code(row.get("GFP_VORZ") or row.get("TFP_VORZ")) == "-1" else "expenditure"
                        base = int(stable_code(row.get("GJAHR"))); measures = [(base, "enacted", row.get("PLAN_0"))] + [(base + offset, "proposal", row.get(f"PLAN_{offset}")) for offset in range(1, 5)]
                    elif parser == "moers":
                        function = stable_code(row.get("PSP-Element") or row.get("Profitcenter")); function_name = stable_code(row.get("Bezeichnung PSP-Element")) or function
                        economic = stable_code(row.get("Kostenart")); economic_name = stable_code(row.get("Kostenart Beschreibung")) or economic; side = "revenue" if economic.startswith("4") else "expenditure"
                        measures = [(2022, "enacted", row.get("Plan 2022")), (2023, "enacted", row.get("Plan 2023"))]
                    elif parser == "glueckstadt":
                        function = stable_code(row.get("Produkt-Kennung") or row.get("Produktgruppe - Kennung")); function_name = stable_code(row.get("Produkt-Bezeichnung") or row.get("Produktgruppe - Bezeichnung")) or function
                        economic = stable_code(row.get("Kontonummer") or row.get("Ertrags- und Aufwandsarten - Gliederung 1")); economic_name = stable_code(row.get("Kontobezeichnung") or row.get("Ertrags- und Aufwandsarten Beschreibung")) or economic; side = "revenue" if economic.startswith("4") else "expenditure"
                        measures = [(2022, "enacted", row.get("Haushaltsansatz 2022"))] + [(year, "proposal", row.get(f"Mittelfristige Finanzplanung {year}")) for year in (2023, 2024, 2025)]
                    else:
                        raise ValueError(f"Unsupported German parser: {parser}")
                    raw_year = measures[0][0] if measures else cfg["year"]
                    ctx.bundle.raw(raw_row(country, raw_year, source, run_id, row_number, sheet, row, ctx.loaded_at))
                    for year, stage, value in measures:
                        emit(row_number, sheet, row, year, stage, side, function, function_name, economic, economic_name, german_decimal(value))
            finally:
                stream.close()
                if parser == "bonn": archive.close()
        write_run(ctx, run_id, source, cfg["year"], path, rows_read, rows_loaded)
        ctx.source_row(entity_id, source, country, cfg["coverage"])
        total_rows += rows_read; total_facts += rows_loaded
    return {"entities": len(sources), "rows": total_rows, "facts": total_facts, "fiscal_years": sorted(fiscal_years)}


def run_paraguay_boost(ctx: Context, country: str, cfg: dict[str, Any]) -> dict[str, Any]:
    """Load the municipal worksheet of the endorsed Paraguay BOOST file."""
    source = cfg["sources"][0]; path = ctx.download(country, source); currency = cfg["currency"]
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[source.get("sheet", "Municipalidades")]
    rows = sheet.iter_rows(values_only=True)
    headers = [stable_code(value) for value in next(rows)]
    economic_class = "PY_BOOST_MUNICIPAL_ECONOMIC_2006_2022"
    ctx.bundle.classification(classification_row(country, economic_class, "expenditure", "Paraguay BOOST municipal economic classification", source.get("catalog_url", source["url"]), 2006, ctx.loaded_at))
    run_id = f"{source['id']}-v1"; entities: set[str] = set(); rows_read = rows_loaded = 0; fiscal_years: set[int] = set()
    for row_number, values in enumerate(rows, 2):
        row = dict(zip(headers, values)); year_value = row.get("YEAR")
        if year_value in (None, ""):
            continue
        year = int(year_value)
        admin = stable_code(row.get("ADMIN2")); code, name = _de_split_code(admin)
        if not code:
            continue
        normalized_code = code.replace(".", "")
        if normalized_code not in entities:
            if ctx.args.max_entities and len(entities) >= ctx.args.max_entities:
                continue
            entities.add(normalized_code)
            ctx.bundle.entity(entity_row(country, normalized_code, name or code, currency, ctx.loaded_at, code_type="PY_BOOST_MUNICIPAL_CODE"))
        rows_read += 1; fiscal_years.add(year)
        ctx.bundle.raw(raw_row(country, year, source, run_id, row_number, sheet.title, row, ctx.loaded_at))
        economic_raw = stable_code(row.get("ECON6"))
        if not economic_raw or economic_raw.lower() == "no disponible": economic_raw = stable_code(row.get("ECON5") or row.get("ECON4"))
        economic, economic_name = _de_split_code(economic_raw)
        node_code = f"expenditure:{economic}"
        ctx.bundle.node(node_row(country, economic_class, "expenditure", node_code, economic_name or economic, year, ctx.loaded_at))
        for stage, field in (("enacted", "approved"), ("revised", "MODIFIED"), ("actual", "PAID")):
            amount = reported_amount(row.get(field))
            if amount is None:
                continue
            ctx.bundle.write("municipal_budget_line_facts", fact_row(
                country, f"PY:{normalized_code}", year, stage, "expenditure", None, node_code, abs(amount),
                currency, source, run_id, ctx.loaded_at, None, economic_class,
                row_number=row_number, sheet=sheet.title, coverage_type="published_subset",
                quality_flags=["paraguay_boost_municipal_worksheet", "officially_endorsed_world_bank_distribution", "economic_classification_only"],
            ))
            rows_loaded += 1
    workbook.close()
    write_run(ctx, run_id, source, cfg["year"], path, rows_read, rows_loaded)
    ctx.source_row(None, source, country, cfg["coverage"])
    return {"entities": len(entities), "rows": rows_read, "facts": rows_loaded, "fiscal_years": sorted(fiscal_years)}


ADAPTERS = {
    "poland_dbf": run_poland,
    "denmark_statbank": run_denmark,
    "ukraine_openbudget": run_ukraine,
    "ukraine_openbudget_public_api": run_ukraine_public_api,
    "france_dgfip": run_france,
    "france_structured": run_france_structured,
    "sweden_pxweb": run_sweden,
    "england_mhclg": run_england,
    "united_kingdom_devolved": run_united_kingdom,
    "us_socrata_cities": run_us_socrata,
    "switzerland_lucerne": run_switzerland_lucerne,
    "switzerland_structured": run_switzerland_structured,
    "germany_bremen": run_germany_bremen,
    "germany_structured_cities": run_germany_structured,
    "paraguay_boost_municipal": run_paraguay_boost,
}


def bundle_path(output: Path, filename: str) -> Path:
    plain = output / filename
    compressed = output / f"{filename}.gz"
    return compressed if compressed.exists() else plain


def iter_jsonl(path: Path) -> Iterator[dict[str, Any]]:
    if not path.exists():
        return
    opener = gzip.open if path.suffix == ".gz" else open
    with opener(path, "rt", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, 1):
            if line.strip():
                try:
                    yield json.loads(line)
                except json.JSONDecodeError as exc:
                    raise ValueError(f"Invalid JSON in {path.name}:{line_number}: {exc}") from exc


def validate_bundle(output: Path) -> dict[str, Any]:
    errors: list[str] = []
    entities = set()
    for row in iter_jsonl(bundle_path(output, "public_entities.jsonl")):
        key = row["public_entity_id"]
        if key in entities:
            errors.append(f"duplicate public_entity_id: {key}")
        entities.add(key)
    classifications = {row["classification_id"] for row in iter_jsonl(bundle_path(output, "classification_versions.jsonl"))}
    nodes = set()
    for row in iter_jsonl(bundle_path(output, "budget_nodes.jsonl")):
        key = row["budget_node_id"]
        if key in nodes:
            errors.append(f"duplicate budget_node_id: {key}")
        nodes.add(key)
        if row["classification_id"] not in classifications:
            errors.append(f"node {key} has missing classification {row['classification_id']}")
    runs = {row["ingestion_run_id"] for row in iter_jsonl(bundle_path(output, "ingestion_runs.jsonl"))}
    fact_rows = 0
    for filename in ("municipal_budget_line_facts.jsonl", "public_entity_balance_sheet_facts.jsonl", "public_entity_cash_facts.jsonl"):
        for row in iter_jsonl(bundle_path(output, filename)):
            fact_rows += 1
            if row["public_entity_id"] not in entities:
                errors.append(f"{filename}: missing entity {row['public_entity_id']}")
            if row["ingestion_run_id"] not in runs:
                errors.append(f"{filename}: missing ingestion run {row['ingestion_run_id']}")
            for field in ("amount_local", "amount_eur"):
                value = row.get(field)
                if value is not None and not bigquery_numeric_compatible(value):
                    errors.append(f"{filename}: {field} exceeds BigQuery NUMERIC(38,9): {value}")
            if filename == "municipal_budget_line_facts.jsonl":
                if row["budget_stage"] not in {"proposal", "enacted", "revised", "actual"}:
                    errors.append(f"invalid budget stage: {row['budget_stage']}")
                if row["budget_side"] not in {"revenue", "expenditure", "financing"}:
                    errors.append(f"invalid budget side: {row['budget_side']}")
                for field in ("functional_classification_id", "economic_classification_id"):
                    value = row.get(field)
                    if value and value not in classifications:
                        errors.append(f"missing fact classification: {value}")
            if len(errors) >= 100:
                break
        if len(errors) >= 100:
            break
    return {"status": "passed" if not errors else "failed", "fact_rows_checked": fact_rows, "errors": errors}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--countries", default="POL,DNK,UKR,FRA,SWE,GBR,DEU,USA,CHE,PRY", help="Comma-separated ISO alpha-3 list")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--cache-dir", type=Path, default=DEFAULT_CACHE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--year", type=int, help="Optional filter for multi-year source rows")
    parser.add_argument("--years", help="Comma-separated fiscal years; for example 2024,2025")
    parser.add_argument("--max-entities", type=int, help="Deterministic smoke-test limit per country")
    parser.add_argument("--raw-mode", choices=("all", "sample", "none"), default="all")
    parser.add_argument("--raw-limit", type=int, default=10000)
    parser.add_argument("--gzip", action="store_true", help="Write .jsonl.gz bundle files")
    parser.add_argument("--offline", action="store_true")
    parser.add_argument("--refresh", action="store_true")
    parser.add_argument("--source-file", action="append", default=[], help="Override as SOURCE_ID=/absolute/path")
    parser.add_argument("--openbudget-token")
    parser.add_argument("--api-workers", type=int, default=4, help="Concurrent workers for paged/per-entity public APIs")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.year and args.years:
        raise SystemExit("Use either --year or --years, not both")
    config = json.loads(args.config.read_text(encoding="utf-8"))
    requested = [country.strip().upper() for country in args.countries.split(",") if country.strip()]
    unknown = sorted(set(requested) - set(config["countries"]))
    if unknown:
        raise SystemExit(f"Unknown countries: {', '.join(unknown)}")
    bundle = JsonlBundle(args.output_dir, args.raw_mode, args.raw_limit, args.gzip)
    context = Context(config, args, bundle)
    results: dict[str, Any] = {}
    failures: dict[str, str] = {}
    try:
        for country in requested:
            cfg = config["countries"][country]
            try:
                yearly_results = {}
                for year in configured_years(args, cfg):
                    year_cfg = config_for_year(cfg, year)
                    if not year_cfg["sources"]:
                        yearly_results[str(year)] = {"entities": 0, "unavailable": "No configured official source for this year"}
                        continue
                    yearly_results[str(year)] = ADAPTERS[cfg["adapter"]](context, country, year_cfg)
                results[country] = yearly_results
            except Exception as exc:
                failures[country] = f"{type(exc).__name__}: {exc}"
                if not args.max_entities:
                    raise
    finally:
        bundle.close()
    validation = validate_bundle(args.output_dir)
    if validation["status"] != "passed":
        failures["bundle_validation"] = "; ".join(validation["errors"][:5])
    manifest = {
        "schema_version": "1.0.0", "generated_at": context.loaded_at, "countries_requested": requested,
        "country_results": results, "failures": failures, "output_rows": dict(bundle.counts), "validation": validation,
        "raw_mode": args.raw_mode, "max_entities": args.max_entities, "gzip": args.gzip,
        "coverage_warning": "GBR currently covers England, Scotland and Wales; Northern Ireland remains outside the loaded itemized collection. Ukraine includes territorial-community budgets and Kyiv, not oblast or district budgets. DEU, USA and CHE are explicitly partial subnational collections.",
    }
    (args.output_dir / "international_municipal_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    if failures:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
