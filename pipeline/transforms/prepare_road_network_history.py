#!/usr/bin/env python3
"""Normalize official road-network stock history for the ten country profiles."""

from __future__ import annotations

import hashlib
import json
import os
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(os.environ.get("CZBUDGET_WORKSPACE_ROOT", Path(__file__).resolve().parents[3]))
RAW = ROOT / "data" / "sources" / "transport"
WEB_DATA = ROOT / "website" / "data"
KM_PER_MILE = 1.609344

COUNTRIES = {
    "CZE": ("CZ", "Česko", "Czechia"),
    "DEU": ("DE", "Německo", "Germany"),
    "DNK": ("DK", "Dánsko", "Denmark"),
    "FRA": ("FR", "Francie", "France"),
    "GBR": ("UK", "Spojené království", "United Kingdom"),
    "POL": ("PL", "Polsko", "Poland"),
    "SWE": ("SE", "Švédsko", "Sweden"),
    "CHE": ("CH", "Švýcarsko", "Switzerland"),
    "UKR": ("UA", "Ukrajina", "Ukraine"),
}

SOURCES = [
    {
        "id": "eurostat-road-if-roadsc",
        "file": "eurostat-road-if-roadsc.json",
        "url": "https://ec.europa.eu/eurostat/databrowser/view/road_if_roadsc/default/table",
        "title": "Eurostat — length of road network by category",
    },
    {
        "id": "eurostat-road-if-motorwa",
        "file": "eurostat-road-if-motorwa.json",
        "url": "https://ec.europa.eu/eurostat/databrowser/view/road_if_motorwa/default/table",
        "title": "Eurostat — length of motorways and E-roads",
    },
    {
        "id": "us-fhwa-hm220",
        "file": "us-fhwa-hm220-1980-2024.xlsx",
        "url": "https://www.fhwa.dot.gov/policyinformation/statistics/2024/hm220.cfm",
        "title": "FHWA Highway Statistics 2024 — HM-220",
    },
    {
        "id": "us-fhwa-hm10",
        "file": "us-fhwa-hm10-2024.xlsx",
        "url": "https://www.fhwa.dot.gov/policyinformation/statistics/2024/hm10.cfm",
        "title": "FHWA Highway Statistics 2024 — HM-10",
    },
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_stat_series(payload: dict, selections: dict[str, str]) -> list[dict]:
    dimensions = payload["id"]
    sizes = payload["size"]
    time_index = payload["dimension"]["time"]["category"]["index"]
    values = payload.get("value", {})
    statuses = payload.get("status", {})
    result = []
    for year, year_coordinate in time_index.items():
        coordinates = []
        for dimension in dimensions:
            if dimension == "time":
                coordinates.append(year_coordinate)
            else:
                key = selections[dimension]
                coordinates.append(payload["dimension"][dimension]["category"]["index"][key])
        flat_index = 0
        for offset, coordinate in enumerate(coordinates):
            stride = 1
            for size in sizes[offset + 1 :]:
                stride *= size
            flat_index += coordinate * stride
        value = values.get(str(flat_index), values.get(flat_index))
        if value is None:
            continue
        status = statuses.get(str(flat_index), statuses.get(flat_index))
        result.append({"year": int(year), "km": round(float(value), 3), **({"status": status} if status else {})})
    return result


def add_net_change(series: list[dict], break_threshold: float, minimum_break_km: float) -> list[dict]:
    previous = None
    result = []
    for point in series:
        item = dict(point)
        if previous and point["year"] == previous["year"] + 1:
            item["annual_net_change_km"] = round(point["km"] - previous["km"], 3)
            base = abs(previous["km"])
            if (
                base
                and abs(item["annual_net_change_km"]) >= minimum_break_km
                and abs(item["annual_net_change_km"]) / base >= break_threshold
            ):
                item["quality_flags"] = ["possible_series_break_or_reclassification"]
        result.append(item)
        previous = point
    return result


def metric(
    definition: str,
    source_id: str,
    series: list[dict],
    *,
    break_threshold: float,
    minimum_break_km: float,
) -> dict:
    normalized = add_net_change(series, break_threshold, minimum_break_km)
    latest_year = normalized[-1]["year"] if normalized else None
    return {
        "definition": definition,
        "source_id": source_id,
        "latest_year": latest_year,
        "freshness": "current" if latest_year and latest_year >= 2023 else "stale_or_incomplete",
        "series": normalized,
    }


def european_countries() -> list[dict]:
    road_payload = json.loads((RAW / "eurostat-road-if-roadsc.json").read_text(encoding="utf-8"))
    motorway_payload = json.loads((RAW / "eurostat-road-if-motorwa.json").read_text(encoding="utf-8"))
    result = []
    for code, (geo, name_cs, name_en) in COUNTRIES.items():
        roads = json_stat_series(road_payload, {"freq": "A", "unit": "KM", "tra_infr": "TOTAL", "geo": geo})
        motorways = json_stat_series(motorway_payload, {"freq": "A", "tra_infr": "MWAY", "unit": "KM", "geo": geo})
        result.append(
            {
                "code": code,
                "name_cs": name_cs,
                "name_en": name_en,
                "road_network": metric(
                    "Eurostat total road network",
                    "eurostat-road-if-roadsc",
                    roads,
                    break_threshold=0.02,
                    minimum_break_km=500,
                ),
                "motorways": metric(
                    "Eurostat motorways",
                    "eurostat-road-if-motorwa",
                    motorways,
                    break_threshold=0.15,
                    minimum_break_km=100,
                ),
            }
        )
    return result


def united_states() -> dict:
    workbook = load_workbook(RAW / "us-fhwa-hm220-1980-2024.xlsx", read_only=True, data_only=True)
    sheet = workbook["SUMMARY"]
    roads = []
    interstates = []
    for row in sheet.iter_rows(min_row=11, values_only=True):
        try:
            year = int(row[0])
        except (TypeError, ValueError):
            continue
        if not 1980 <= year <= 2024:
            continue
        total_miles = row[17]
        rural_interstate_miles = row[1]
        urban_interstate_miles = row[9]
        if isinstance(total_miles, (int, float)):
            roads.append({"year": year, "km": round(total_miles * KM_PER_MILE, 3)})
        if isinstance(rural_interstate_miles, (int, float)) and isinstance(urban_interstate_miles, (int, float)):
            interstates.append({"year": year, "km": round((rural_interstate_miles + urban_interstate_miles) * KM_PER_MILE, 3)})
    workbook.close()
    return {
        "code": "USA",
        "name_cs": "Spojené státy",
        "name_en": "United States",
        "road_network": metric(
            "FHWA public roads and streets",
            "us-fhwa-hm220",
            roads,
            break_threshold=0.02,
            minimum_break_km=500,
        ),
        "motorways": metric(
            "FHWA Interstate mileage; a conservative motorway proxy, not the full freeway network",
            "us-fhwa-hm220",
            interstates,
            break_threshold=0.15,
            minimum_break_km=100,
        ),
    }


def build() -> None:
    generated_at = datetime.now(timezone.utc).isoformat()
    source_rows = []
    for source in SOURCES:
        path = RAW / source["file"]
        if not path.is_file():
            raise FileNotFoundError(path)
        source_rows.append({**source, "bytes": path.stat().st_size, "sha256": sha256(path)})

    payload = {
        "schema_version": "1.0",
        "generated_at": generated_at,
        "unit": "km",
        "construction_history_status": "no_harmonized_openings_series; annual net stock change is provided as a proxy",
        "construction_measure": "annual net change in reported network stock",
        "methodology_cs": "Roční přírůstek je rozdíl mezi vykázanou délkou sítě ve dvou po sobě jdoucích letech. Není to čistý údaj o kilometrech nově otevřených staveb: zahrnuje také překlasifikování, opravy dat a vyřazení úseků. Evropské země používají Eurostat; USA používají FHWA a pro dálnice konzervativní proxy Interstate.",
        "methodology_en": "Annual addition is the difference between reported network length in consecutive years. It is not a pure measure of newly opened construction: reclassification, data revisions and retirements also affect it. European countries use Eurostat; the US uses FHWA and Interstate mileage as a conservative motorway proxy.",
        "countries": [*european_countries(), united_states()],
        "sources": [{k: v for k, v in source.items() if k != "sha256"} for source in source_rows],
    }
    WEB_DATA.mkdir(parents=True, exist_ok=True)
    (WEB_DATA / "road-network-history.v1.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (RAW / "manifest.json").write_text(
        json.dumps({"schema_version": "1.0", "generated_at": generated_at, "sources": source_rows}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print("Wrote road network history and source manifest")


if __name__ == "__main__":
    build()
