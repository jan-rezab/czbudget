#!/usr/bin/env python3
"""Build a comparable sovereign-fiscal benchmark from IMF WEO data.

The output is intentionally a general-government macro spine, not a substitute
for national budget data. National plan/outturn/program trees are registered in
data/international_fiscal_source_registry.json and will be layered on later.
"""

from __future__ import annotations

import os

import argparse
import csv
import json
import math
import re
import shutil
import statistics
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(os.environ.get("CZBUDGET_WORKSPACE_ROOT", Path(__file__).resolve().parents[3]))
SOURCE_URL = (
    "https://data.imf.org/-/media/iData/External-Storage/Documents/"
    "2F78EE59F79143A7921E5E203D3AAA80/en/WEOApr2026all.xlsx"
)
DEFAULT_SOURCE = ROOT / "data/sources/international_fiscal/WEOApr2026all.xlsx"
DEFAULT_REGISTRY = ROOT / "website/pipeline/config/international_fiscal_source_registry.json"
DEFAULT_SCOPE_REGISTRY = ROOT / "website/pipeline/config/fiscal_scope_registry.json"
DEFAULT_UNIVERSE = ROOT / "website/pipeline/config/sovereign_country_universe.json"
DEFAULT_JSON = ROOT / "data/international_fiscal_benchmarks_2005_2024.json"
DEFAULT_CSV = ROOT / "data/international_fiscal_benchmarks_2005_2024.csv"
DEFAULT_SUMMARY_CSV = ROOT / "data/international_fiscal_summary_2005_2024.csv"
DEFAULT_WEB_JSON = ROOT / "website/lib/data/sovereign-benchmark.v1.json"

START_YEAR = 2005
END_YEAR = 2024

FULL_PROFILE_ORDER = ["CZE", "UKR", "POL", "DEU", "GBR", "FRA", "USA", "CHE", "SWE", "DNK", "FIN", "BRA", "ESP", "JPN", "NLD", "NOR", "GRC"]

CURRENCY_CODES = {
    "Bangladesh taka": "BDT", "Lao kip": "LAK", "Mongolian tögrög": "MNT",
    "Papua New Guinea kina": "PGK", "Bosnian convertible marka": "BAM",
    "Eastern Caribbean dollar": "XCD", "Barbados dollar": "BBD", "Guyanese dollar": "GYD",
    "Trinidad and Tobago dollar": "TTD", "Azerbaijan manat": "AZN", "Bahrain dinar": "BHD",
    "Djibouti franc": "DJF", "Pakistan rupee": "PKR", "Tajik somoni": "TJS",
    "New Turkmen manat": "TMT", "U.A.E. dirham": "AED", "Uzbek som": "UZS",
    "Botswana pula": "BWP", "Burundi franc": "BIF", "Cabo Verdean escudo": "CVE",
    "São Tomé and Príncipe dobra": "STN", "Seychelles rupee": "SCR",
    "Zimbabwe gold": "ZWG", "United States dollar": "USD", "U.S. dollar": "USD",
    "Pound sterling": "GBP", "UK pound sterling": "GBP", "CFA franc (WAEMU)": "XOF",
    "CFA franc (CEMAC)": "XAF", "Chinese yuan": "CNY", "New Taiwan dollar": "TWD",
}

METRICS = {
    "revenue_pct_gdp": {
        "imf_indicator": "GGR_NGDP",
        "label_cs": "Příjmy vládních institucí",
        "unit": "pct_gdp",
        "polarity": "neutral",
    },
    "expenditure_pct_gdp": {
        "imf_indicator": "GGX_NGDP",
        "label_cs": "Výdaje vládních institucí",
        "unit": "pct_gdp",
        "polarity": "neutral",
    },
    "balance_pct_gdp": {
        "imf_indicator": "GGXCNL_NGDP",
        "label_cs": "Čisté půjčky (+) / výpůjčky (-)",
        "unit": "pct_gdp",
        "polarity": "positive",
    },
    "primary_balance_pct_gdp": {
        "imf_indicator": "GGXONLB_NGDP",
        "label_cs": "Primární saldo",
        "unit": "pct_gdp",
        "polarity": "positive",
    },
    "structural_balance_pct_potential_gdp": {
        "imf_indicator": "GGSB_NPGDP",
        "label_cs": "Strukturální saldo",
        "unit": "pct_potential_gdp",
        "polarity": "positive",
    },
    "gross_debt_pct_gdp": {
        "imf_indicator": "GGXWDG_NGDP",
        "label_cs": "Hrubý dluh vládních institucí",
        "unit": "pct_gdp",
        "polarity": "negative",
    },
    "nominal_gdp_usd_bn": {
        "imf_indicator": "NGDPD",
        "label_cs": "Nominální HDP",
        "label_en": "Nominal GDP",
        "unit": "usd_bn",
        "polarity": "neutral",
    },
    "nominal_gdp_local_bn": {
        "imf_indicator": "NGDP",
        "label_cs": "Nominální HDP v domácí měně",
        "label_en": "Nominal GDP in domestic currency",
        "unit": "local_currency_bn",
        "polarity": "neutral",
    },
    "gdp_per_capita_usd": {
        "imf_indicator": "NGDPDPC",
        "label_cs": "HDP na obyvatele",
        "label_en": "GDP per capita",
        "unit": "usd_per_capita",
        "polarity": "positive",
    },
    "gdp_per_capita_local": {
        "imf_indicator": "NGDPPC",
        "label_cs": "HDP na obyvatele v domácí měně",
        "label_en": "GDP per capita in domestic currency",
        "unit": "local_currency_per_capita",
        "polarity": "positive",
    },
    "gdp_per_capita_ppp": {
        "imf_indicator": "PPPPC",
        "label_cs": "HDP na obyvatele v paritě kupní síly",
        "label_en": "GDP per capita at purchasing power parity",
        "unit": "international_dollar_per_capita",
        "polarity": "positive",
    },
    "ppp_conversion_factor": {
        "imf_indicator": "PPPEX",
        "label_cs": "PPP převodní faktor",
        "label_en": "PPP conversion factor",
        "unit": "local_currency_per_international_dollar",
        "polarity": "neutral",
    },
    "real_gdp_growth_pct": {
        "imf_indicator": "NGDP_RPCH",
        "label_cs": "Reálný růst HDP",
        "unit": "percent",
        "polarity": "positive",
    },
    "inflation_pct": {
        "imf_indicator": "PCPIPCH",
        "label_cs": "Inflace CPI",
        "unit": "percent",
        "polarity": "target_2",
    },
    "unemployment_pct": {
        "imf_indicator": "LUR",
        "label_cs": "Nezaměstnanost",
        "unit": "percent",
        "polarity": "negative",
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--registry", type=Path, default=DEFAULT_REGISTRY)
    parser.add_argument("--scope-registry", type=Path, default=DEFAULT_SCOPE_REGISTRY)
    parser.add_argument("--universe", type=Path, default=DEFAULT_UNIVERSE)
    parser.add_argument("--output", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--csv-output", type=Path, default=DEFAULT_CSV)
    parser.add_argument("--summary-output", type=Path, default=DEFAULT_SUMMARY_CSV)
    parser.add_argument("--web-output", type=Path, default=DEFAULT_WEB_JSON)
    parser.add_argument("--download", action="store_true")
    return parser.parse_args()


def download_source(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with urllib.request.urlopen(SOURCE_URL) as response, path.open("wb") as target:
        shutil.copyfileobj(response, target)


def clean_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    number = float(value)
    if not math.isfinite(number):
        return None
    return round(number, 6)


def median(values: list[float]) -> float | None:
    return round(statistics.median(values), 3) if values else None


def mean(values: list[float]) -> float | None:
    return round(statistics.fmean(values), 3) if values else None


def stdev(values: list[float]) -> float | None:
    return round(statistics.pstdev(values), 3) if values else None


def value_map(series: dict[str, Any], metric_code: str) -> dict[int, float]:
    return {
        point["year"]: point["value"]
        for point in series[metric_code]["values"]
        if point["value"] is not None
    }


def build_summary(country_series: dict[str, Any]) -> dict[str, Any]:
    balance = value_map(country_series, "balance_pct_gdp")
    primary = value_map(country_series, "primary_balance_pct_gdp")
    debt = value_map(country_series, "gross_debt_pct_gdp")
    revenue = value_map(country_series, "revenue_pct_gdp")
    expenditure = value_map(country_series, "expenditure_pct_gdp")
    balance_values = list(balance.values())
    debt_start = debt.get(START_YEAR)
    debt_end = debt.get(END_YEAR)
    debt_2021 = debt.get(2021)

    observed_years = sorted({
        point["year"]
        for metric in country_series.values()
        for point in metric["values"]
        if point["value"] is not None
    })
    return {
        "period_start": observed_years[0] if observed_years else None,
        "period_end": observed_years[-1] if observed_years else None,
        "median_balance_pct_gdp": median(balance_values),
        "mean_balance_pct_gdp": mean(balance_values),
        "balance_volatility_pp": stdev(balance_values),
        "surplus_year_share": round(sum(value >= 0 for value in balance_values) / len(balance_values), 4)
        if balance_values
        else None,
        "worst_balance_pct_gdp": round(min(balance_values), 3) if balance_values else None,
        "median_primary_balance_pct_gdp": median(list(primary.values())),
        "mean_revenue_pct_gdp": mean(list(revenue.values())),
        "mean_expenditure_pct_gdp": mean(list(expenditure.values())),
        "gross_debt_start_pct_gdp": debt_start,
        "gross_debt_end_pct_gdp": debt_end,
        "gross_debt_change_pp": round(debt_end - debt_start, 3)
        if debt_start is not None and debt_end is not None
        else None,
        "gross_debt_peak_pct_gdp": round(max(debt.values()), 3) if debt else None,
        "post_2021_debt_change_pp": round(debt_end - debt_2021, 3)
        if debt_end is not None and debt_2021 is not None
        else None,
    }


def validate_extracted(extracted: dict[str, dict[str, Any]], full_profile_codes: set[str]) -> None:
    fiscal_metrics = (
        "revenue_pct_gdp",
        "expenditure_pct_gdp",
        "balance_pct_gdp",
        "primary_balance_pct_gdp",
        "gross_debt_pct_gdp",
    )
    for country_code, country_series in extracted.items():
        missing_metrics = set(METRICS) - set(country_series)
        if missing_metrics:
            raise ValueError(f"{country_code} is missing metrics: {sorted(missing_metrics)}")

        for metric_code in fiscal_metrics:
            points = country_series[metric_code]["values"]
            if len(points) != END_YEAR - START_YEAR + 1:
                raise ValueError(f"{country_code}/{metric_code} has an incomplete period")
            if country_code in full_profile_codes:
                for point in points:
                    if point["value"] is None:
                        raise ValueError(f"{country_code}/{metric_code}/{point['year']} is missing")
                    if point["status"] != "actual":
                        raise ValueError(f"{country_code}/{metric_code}/{point['year']} is not actual")

        revenue = value_map(country_series, "revenue_pct_gdp")
        expenditure = value_map(country_series, "expenditure_pct_gdp")
        balance = value_map(country_series, "balance_pct_gdp")
        for year in range(START_YEAR, END_YEAR + 1):
            if any(values.get(year) is None for values in (revenue, expenditure, balance)):
                continue
            identity_error = abs((revenue[year] - expenditure[year]) - balance[year])
            if identity_error > 0.02:
                raise ValueError(
                    f"{country_code}/{year} revenue - expenditure != balance "
                    f"(error {identity_error:.3f} pp)"
                )


def main() -> None:
    args = parse_args()
    if args.download or not args.source.exists():
        download_source(args.source)

    registry = json.loads(args.registry.read_text(encoding="utf-8"))
    registry_by_country = {country["country_code"]: country for country in registry["countries"]}
    missing_registry = set(FULL_PROFILE_ORDER) - set(registry_by_country)
    if missing_registry:
        raise ValueError(f"Missing countries in registry: {sorted(missing_registry)}")

    scope_registry = json.loads(args.scope_registry.read_text(encoding="utf-8"))
    scope_by_country = {country["country_code"]: country for country in scope_registry["countries"]}
    missing_scope = set(FULL_PROFILE_ORDER) - set(scope_by_country)
    if missing_scope:
        raise ValueError(f"Missing fiscal-scope definitions: {sorted(missing_scope)}")
    perimeter_codes = {
        perimeter["perimeter_code"] for perimeter in scope_registry["methodology"]["perimeters"]
    }
    if perimeter_codes != {"national_budget", "general_government", "public_sector"}:
        raise ValueError(f"Unexpected fiscal perimeters: {sorted(perimeter_codes)}")

    universe = json.loads(args.universe.read_text(encoding="utf-8"))
    universe_by_code = {country["iso3"]: country for country in universe["countries"]}
    weo_to_public = {country["weo_code"]: country["iso3"] for country in universe_by_code.values() if country.get("weo_code")}
    country_order = FULL_PROFILE_ORDER + sorted(
        (code for code in universe_by_code if code not in FULL_PROFILE_ORDER),
        key=lambda code: universe_by_code[code]["name_en"],
    )

    workbook = load_workbook(args.source, read_only=True, data_only=True)
    sheet = workbook["Countries"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    column = {name: index for index, name in enumerate(headers)}
    indicator_to_metric = {definition["imf_indicator"]: code for code, definition in METRICS.items()}

    extracted: dict[str, dict[str, Any]] = {country: {} for country in country_order}
    fiscal_metadata: dict[str, dict[str, Any]] = {country: {} for country in country_order}
    currency_names: dict[str, str | None] = {country: None for country in country_order}

    for row in rows:
        source_country_code = row[column["COUNTRY.ID"]]
        country_code = weo_to_public.get(source_country_code)
        indicator_code = row[column["INDICATOR.ID"]]
        if country_code not in extracted or indicator_code not in indicator_to_metric:
            continue

        metric_code = indicator_to_metric[indicator_code]
        currency_names[country_code] = row[column["PRIMARY_DOMESTIC_CURRENCY"]]
        latest_actual_raw = row[column["LATEST_ACTUAL_ANNUAL_DATA"]]
        latest_actual_match = re.search(r"(?:19|20)\d{2}", str(latest_actual_raw or ""))
        latest_actual = int(latest_actual_match.group()) if latest_actual_match else None
        values = []
        for year in range(START_YEAR, END_YEAR + 1):
            value = clean_number(row[column[year]])
            status = "actual" if latest_actual is not None and year <= latest_actual else "estimate"
            values.append({"year": year, "value": value, "status": status})

        extracted[country_code][metric_code] = {
            "latest_actual_year": latest_actual,
            "values": values,
        }

        if metric_code == "gross_debt_pct_gdp":
            fiscal_metadata[country_code] = {
                "general_government_composition": row[column["FISCAL_SECTOR_GENERAL_GOVERNMENT_COMPOSITION"]],
                "debt_valuation": row[column["FISCAL_SECTOR_VALUATION_OF_DEBT"]],
                "debt_instruments": row[
                    column["FISCAL_SECTOR_INSTRUMENTS_INCLUDED_IN_GROSS_AND_NET_DEBT"]
                ],
            }

    # WEO omits a small number of UN members. Keep explicit all-null series for
    # them (and for any metric WEO omits for a covered country) so absence is a
    # published status rather than a country silently disappearing.
    for country_code in country_order:
        for metric_code in METRICS:
            extracted[country_code].setdefault(metric_code, {
                "latest_actual_year": None,
                "values": [
                    {"year": year, "value": None, "status": "not_available"}
                    for year in range(START_YEAR, END_YEAR + 1)
                ],
            })

    validate_extracted(extracted, set(FULL_PROFILE_ORDER))

    countries = []
    summaries = []
    for country_code in country_order:
        country = registry_by_country.get(country_code)
        universe_country = universe_by_code[country_code]
        if country:
            country_metadata = {
                key: country[key]
                for key in (
                    "country_code", "role", "name_cs", "name_en", "currency_code",
                    "national_scope", "benchmark_reason", "benchmark_evidence_url",
                )
                if key in country
            }
            architecture = scope_by_country[country_code]
        else:
            currency_name = currency_names[country_code]
            # LCU is an explicit local-currency-unit marker when the WEO
            # workbook names the currency but does not publish an ISO code.
            currency_code = CURRENCY_CODES.get(currency_name or "", "LCU")
            country_metadata = {
                "country_code": country_code,
                "iso2": universe_country["iso2"],
                "weo_country_code": universe_country.get("weo_code"),
                "role": "global_macro_profile",
                "profile_tier": "macro_fiscal",
                "name_cs": universe_country["name_cs"],
                "name_en": universe_country["name_en"],
                "currency_code": currency_code,
                "currency_name": currency_name,
                "national_scope": "not_source_mapped",
            }
            architecture = {
                "country_code": country_code,
                "eu_status": "not_assessed",
                "national_budget_label_cs": f"Národní rozpočet — {universe_country['name_cs']}",
                "national_budget_label_en": f"National budget — {universe_country['name_en']}",
                "municipal_budget_relation": "not_assessed",
                "other_public_accounts_relation": "not_assessed",
                "public_corporation_treatment": "not_assessed",
                "architecture_cs": "Národní právní a institucionální uspořádání zatím není zmapováno. Publikovaná makrofiskální řada používá výhradně harmonizovaný sektor vládních institucí IMF.",
                "architecture_en": "The national legal and institutional architecture has not yet been source-mapped. The published macro-fiscal series uses only the IMF harmonised general-government sector.",
                "corporation_note_cs": "Údaje IMF za sektor vládních institucí nezahrnují tržní veřejné korporace; jejich obrat se k příjmům vlády nepřičítá.",
                "corporation_note_en": "IMF general-government data excludes market public corporations; their turnover is not added to government revenue.",
                "sources": [{
                    "source_name": "IMF World Economic Outlook",
                    "source_url": "https://data.imf.org/en/Datasets/WEO",
                    "purpose": "Harmonised general-government perimeter; national legal architecture remains explicitly not assessed.",
                }],
            }
        has_observation = any(
            point["value"] is not None
            for metric in extracted[country_code].values()
            for point in metric["values"]
        )
        countries.append(
            (country_metadata | {
                "iso2": universe_country["iso2"],
                "weo_country_code": universe_country.get("weo_code"),
                "data_status": "loaded" if has_observation else "not_loaded",
                "missing_dimensions": [] if has_observation else ["all WEO macro-fiscal metrics"],
                "imf_fiscal_metadata": fiscal_metadata[country_code],
                "fiscal_architecture": architecture,
            })
        )
        summaries.append({"country_code": country_code} | build_summary(extracted[country_code]))

    dataset = {
        "schema_version": "1.1.0",
        "dataset_id": "sovereign-fiscal-benchmark",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "period": {"start_year": START_YEAR, "end_year": END_YEAR, "year_count": END_YEAR - START_YEAR + 1},
        "scope": {
            "institutional_sector": "general_government",
            "accounting_note": "Comparable IMF WEO macro-fiscal series; national central/federal budget trees are separate layers.",
            "uk_note": "The requested England comparison is represented by the United Kingdom because England has no separate sovereign state budget.",
            "actuality_note": "Each point carries actual/estimate status from the source series. The selected fiscal period is actual through 2024 for all countries.",
        },
        "fiscal_perimeters": scope_registry["methodology"],
        "benchmark_policy": {
            "composite_score": False,
            "reason": "A single responsibility score would hide fiscal capacity, institutions, shocks and differences in government scope.",
            "descriptive_dimensions": [
                "balance level and volatility",
                "frequency of surpluses",
                "primary balance",
                "gross-debt level and change",
                "post-crisis debt reduction",
                "revenue and expenditure scale",
            ],
        },
        "source": {
            "provider": "International Monetary Fund",
            "dataset": "World Economic Outlook, April 2026",
            "url": SOURCE_URL,
            "download_page": "https://data.imf.org/en/Datasets/WEO",
            "source_file": args.source.name,
        },
        "countries": countries,
        "metrics": [
            {"metric_code": code, **definition}
            for code, definition in METRICS.items()
        ],
        "series": [
            {
                "country_code": country_code,
                "metrics": extracted[country_code],
            }
            for country_code in country_order
        ],
        "summaries": summaries,
        "national_source_registry": registry["countries"],
        "universe": {
            "definition": universe["universe"],
            "sovereign_state_count": len(universe["countries"]),
            "weo_profile_count": len(country_order),
            "missing_from_weo": [country["iso3"] for country in universe["countries"] if not country.get("weo_code") or country["weo_code"] not in weo_to_public],
        },
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(dataset, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    args.web_output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(args.output, args.web_output)

    args.csv_output.parent.mkdir(parents=True, exist_ok=True)
    with args.csv_output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["country_code", "year", "metric_code", "value", "status", "unit", "source_indicator"],
        )
        writer.writeheader()
        for country_code in country_order:
            for metric_code, metric_series in extracted[country_code].items():
                definition = METRICS[metric_code]
                for point in metric_series["values"]:
                    writer.writerow(
                        {
                            "country_code": country_code,
                            "year": point["year"],
                            "metric_code": metric_code,
                            "value": "" if point["value"] is None else point["value"],
                            "status": point["status"],
                            "unit": definition["unit"],
                            "source_indicator": definition["imf_indicator"],
                        }
                    )

    with args.summary_output.open("w", encoding="utf-8", newline="") as handle:
        fieldnames = list(summaries[0].keys())
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(summaries)

    print(f"Wrote {len(country_order)} countries × {len(METRICS)} metrics × {END_YEAR - START_YEAR + 1} years")
    print(args.output)
    print(args.csv_output)
    print(args.summary_output)
    print(args.web_output)


if __name__ == "__main__":
    main()
