#!/usr/bin/env python3
"""Build the EUR conversion series and a reproducible manifest of ministry-budget sources."""

from __future__ import annotations

import os

import csv
import hashlib
import json
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(os.environ.get("CZBUDGET_WORKSPACE_ROOT", Path(__file__).resolve().parents[3]))
RAW = ROOT / "data" / "sources" / "ministries"
WEB_DATA = ROOT / "website" / "data"
ECB_FILE = ROOT / "data" / "sources" / "ecb" / "exr-usd-eur-annual-2005-2024.csv"

SOURCES = [
    {"code":"CZE","year":2025,"stage":"execution","scope":"budget fulfilment of state-budget chapters and organisational units","file":"fin-1-12-oss-2025-12.zip","source_name":"MONITOR Státní pokladny / Ministerstvo financí ČR","source_url":"https://monitor.statnipokladna.gov.cz/datovy-katalog/transakcni-data","dimension":"organisation / commitment item / functional area / fund / source","tier":"A"},
    {"code":"UKR","year":2024,"stage":"budget law","scope":"state budget administrators and programmes","file":"state-budget-law-2024.zip","source_name":"data.gov.ua / Ministry of Finance of Ukraine","source_url":"https://data.gov.ua/dataset/3a6056bd-d775-4952-abd8-58383ff12c2f/resource/7ea5e5ae-0621-4ece-b86d-a757ab991d5e","dimension":"main spending unit / programme","tier":"B"},
    {"code":"POL","year":2025,"stage":"execution","scope":"state-budget execution, volumes I and II","file":"state-budget-execution-2025-tables.zip","source_name":"Ministerstwo Finansów","source_url":"https://www.gov.pl/web/finanse/sprawozdanie-roczne-za-2025-rok","dimension":"budget part / division / chapter / administrator / programme","tier":"A"},
    {"code":"DEU","year":2026,"stage":"budget plan","scope":"federal budget titles","file":"federal-budget-2026.csv","source_name":"Bundeshaushalt","source_url":"https://www.bundeshaushalt.de/DE/Download-Portal/download-portal.html","dimension":"department / chapter / title","tier":"A"},
    {"code":"GBR","year":2027,"stage":"2026-27 Main Supply Estimates","scope":"voted central-government departmental expenditure","file":"main-supply-estimates-2026-27.pdf","source_name":"HM Treasury","source_url":"https://www.gov.uk/government/publications/main-supply-estimates-2026-to-2027","dimension":"department / estimate ambit / programme subhead / control total","tier":"B"},
    {"code":"FRA","year":2024,"stage":"execution","scope":"state budget missions and programmes","file":"state-budget-execution-2024.csv","source_name":"Ministère de l'Économie et des Finances","source_url":"https://data.economie.gouv.fr/explore/dataset/plrg-2024/","dimension":"mission / programme","tier":"A"},
    {"code":"USA","year":2027,"stage":"presidential budget historical database","scope":"federal outlays by agency and account","file":"public-budget-database-outlays-fy2027.xlsx","source_name":"Office of Management and Budget","source_url":"https://www.whitehouse.gov/omb/information-resources/budget/supplemental-materials/","dimension":"agency / bureau / account","tier":"A"},
    {"code":"CHE","year":2026,"stage":"budget plan","scope":"Confederation expenditure by institution","file":"federal-finance-institutions-2026.csv","source_name":"Eidgenössische Finanzverwaltung","source_url":"https://www.efv.admin.ch/de/open-government-data-de","dimension":"institution / task group","tier":"A"},
    {"code":"SWE","year":2024,"stage":"execution","scope":"central-government expenditure","file":"central-government-expenditure-1997-2024.xlsx","source_name":"Statskontoret","source_url":"https://www.statskontoret.se/analys-och-statistik/oppna-data/arsutfall/","dimension":"agency / appropriation","tier":"A"},
    {"code":"DNK","year":2026,"stage":"finance act database","scope":"central-government appropriations","file":"finance-act-database.html","source_name":"Finansministeriet","source_url":"https://fm.dk/arbejdsomraader/finanslov-og-offentlige-finanser/arbejdet-med-finansloven/finanslovsdatabasen/","dimension":"ministry / main account","tier":"B"},
]

FISCAL_FRAMEWORKS = [
    {"code":"CHE","label_cs":"Dluhová brzda","label_en":"Debt brake","summary_cs":"Výdajový strop se odvozuje od cyklicky očištěných příjmů; odchylky se převádějí do kompenzačního účtu a přebytky snižují dluh.","summary_en":"The expenditure ceiling follows cyclically adjusted receipts; deviations flow to a compensation account and surpluses reduce debt.","source_url":"https://www.efv.admin.ch/en/debt-brake-fp"},
    {"code":"SWE","label_cs":"Přebytkový cíl a výdajový strop","label_en":"Surplus target and expenditure ceiling","summary_cs":"Víceletý výdajový strop, dluhová kotva a pravidlo pro místní vyrovnané rozpočty vážou každoroční rozpočet na střednědobý rámec.","summary_en":"A multi-year expenditure ceiling, debt anchor and local balanced-budget rule tie annual budgets to the medium-term framework.","source_url":"https://www.government.se/government-of-sweden/ministry-of-finance/central-government-budget/the-fiscal-policy-framework/"},
    {"code":"DNK","label_cs":"Rozpočtový zákon a čtyřleté stropy","label_en":"Budget law and four-year ceilings","summary_cs":"Limit strukturálního schodku a čtyřleté výdajové stropy pro stát, obce a regiony omezují růst výdajů před sestavením ročního rozpočtu.","summary_en":"A structural-deficit limit and four-year ceilings for central, local and regional government constrain spending before the annual budget is set.","source_url":"https://fm.dk/arbejdsomraader/finanslov-og-offentlige-finanser/budgetlov-og-udgiftslofter/"},
    {"code":"DEU","label_cs":"Dluhová brzda","label_en":"Debt brake","summary_cs":"Ústavní pravidlo omezuje strukturální zadlužení; konkrétní použití závisí na aktuálním znění článku 115 a výjimkách.","summary_en":"The constitutional rule limits structural borrowing; its application depends on the current Article 115 framework and exceptions.","source_url":"https://www.bundesfinanzministerium.de/Content/EN/Downloads/Public-Finances/germanys-federal-debt-rule.html"},
    {"code":"GRC","label_cs":"Střednědobý fiskálně-strukturální plán","label_en":"Medium-term fiscal-structural plan","summary_cs":"Plán 2025–2028 váže rozpočtovou trajektorii na víceletý strop růstu čistých výdajů a současně dokumentuje prioritní reformy a investice.","summary_en":"The 2025–2028 plan binds the fiscal path to a multi-year ceiling for net-expenditure growth while documenting priority reforms and investment.","source_url":"https://minfin.gov.gr/wp-content/uploads/2024/09/EN_Greece_MTFSP_2025_28_final.pdf"},
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def inspect_file(path: Path) -> dict:
    suffix = path.suffix.lower()
    if suffix == ".zip":
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
        return {"container_entries": len(names), "sample_entries": names[:12]}
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        result = {"sheets": workbook.sheetnames}
        workbook.close()
        return result
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
            header = handle.readline().strip()[:1000]
        return {"header_preview": header}
    return {}


def build_fx() -> None:
    values = []
    with ECB_FILE.open(encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            values.append({"year": int(row["TIME_PERIOD"][:4]), "usd_per_eur": round(float(row["OBS_VALUE"]), 8)})
    payload = {
        "schema_version": "1.0",
        "unit": "USD per EUR",
        "note_cs": "Roční referenční kurz ECB. Přepočet USD na EUR = USD / USD za EUR.",
        "note_en": "ECB annual reference rate. USD-to-EUR conversion = USD / USD per EUR.",
        "source": {"name": "European Central Bank Data Portal", "url": "https://data-api.ecb.europa.eu/service/data/EXR/A.USD.EUR.SP00.A?startPeriod=2005&endPeriod=2024&format=csvdata"},
        "values": values,
    }
    WEB_DATA.mkdir(parents=True, exist_ok=True)
    (WEB_DATA / "fx-eur-annual.v1.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def build_manifest() -> None:
    generated = datetime.now(timezone.utc).isoformat()
    rows = []
    for source in SOURCES:
        path = RAW / source["code"] / source["file"]
        entry = dict(source)
        entry["available"] = path.is_file()
        if path.is_file():
            entry.update({"bytes": path.stat().st_size, "sha256": sha256(path), **inspect_file(path)})
        rows.append(entry)
    manifest = {"schema_version":"1.0","generated_at":generated,"comparability":{"A":"machine-readable department or programme dimension; direct mapping candidate","B":"official source captured; extraction or mapping still required"},"countries":rows,"fiscal_frameworks":FISCAL_FRAMEWORKS}
    (RAW / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    web = dict(manifest)
    web["countries"] = [{k:v for k,v in row.items() if k not in {"sample_entries","header_preview","sha256"}} for row in rows]
    (WEB_DATA / "ministry-budget-sources.v1.json").write_text(json.dumps(web, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    build_fx()
    build_manifest()
    print("Wrote EUR FX series and ministry-budget source manifests")
