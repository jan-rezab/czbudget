import importlib.util
import csv
import io
import json
import struct
import tempfile
import unittest
import zipfile
from argparse import Namespace
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


MODULE_PATH = Path(__file__).resolve().parents[1] / "transforms/prepare_international_municipal_data.py"
SPEC = importlib.util.spec_from_file_location("international_municipal", MODULE_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC.loader
SPEC.loader.exec_module(MODULE)


def dbf_bytes():
    fields = [("CODE", "C", 4, 0), ("AMOUNT", "N", 8, 2)]
    header_length = 32 + 32 * len(fields) + 1
    record_length = 1 + sum(field[2] for field in fields)
    header = bytearray(32)
    header[0] = 3
    header[4:8] = struct.pack("<I", 1)
    header[8:10] = struct.pack("<H", header_length)
    header[10:12] = struct.pack("<H", record_length)
    descriptors = bytearray()
    for name, kind, length, decimals in fields:
        descriptor = bytearray(32)
        descriptor[:len(name)] = name.encode("ascii")
        descriptor[11] = ord(kind)
        descriptor[16] = length
        descriptor[17] = decimals
        descriptors.extend(descriptor)
    record = b" " + b"A123" + b"   12.50"
    return bytes(header) + bytes(descriptors) + b"\r" + record + b"\x1a"


class InternationalMunicipalHelpersTest(unittest.TestCase):
    def test_numeric_json_respects_bigquery_numeric_scale(self):
        self.assertEqual(MODULE.numeric_json(Decimal("211384.24024713458000")), "211384.240247135")
        self.assertEqual(MODULE.numeric_json(Decimal("12500.0")), "12500.0")
        self.assertTrue(MODULE.bigquery_numeric_compatible("211384.240247135"))
        self.assertFalse(MODULE.bigquery_numeric_compatible("211384.24024713458000"))

    def test_ods_reader_expands_repeated_cells_without_external_converter(self):
        content = '''<?xml version="1.0" encoding="UTF-8"?>
<office:document-content xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0" xmlns:table="urn:oasis:names:tc:opendocument:xmlns:table:1.0" xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0"><office:body><office:spreadsheet><table:table table:name="Data"><table:table-row><table:table-cell office:value-type="string"><text:p>Code</text:p></table:table-cell><table:table-cell table:number-columns-repeated="2" office:value-type="float" office:value="12.5"/></table:table-row></table:table></office:spreadsheet></office:body></office:document-content>'''
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "sample.ods"
            with zipfile.ZipFile(path, "w") as archive:
                archive.writestr("content.xml", content)
            self.assertEqual(list(MODULE.iter_ods_rows(path, "Data")), [(1, ["Code", Decimal("12.5"), Decimal("12.5")])])

    def test_dbf_reader_preserves_text_and_decimal(self):
        rows = list(MODULE.iter_dbf(io.BytesIO(dbf_bytes()), encoding="ascii"))
        self.assertEqual(rows, [(1, {"CODE": "A123", "AMOUNT": Decimal("12.50")})])

    def test_jsonstat_rows_respect_dimension_order(self):
        dataset = {
            "id": ["Region", "Item"],
            "dimension": {
                "Region": {"category": {"index": {"01": 0}, "label": {"01": "Town"}}},
                "Item": {"category": {"index": {"A": 0, "B": 1}, "label": {"A": "Alpha", "B": "Beta"}}},
            },
            "value": [10, 20],
        }
        self.assertEqual(list(MODULE.jsonstat_rows(dataset)), [
            {"Region": "01", "Item": "A", "Region_label": "Town", "Item_label": "Alpha", "value": 10},
            {"Region": "01", "Item": "B", "Region_label": "Town", "Item_label": "Beta", "value": 20},
        ])

    def test_french_insee_keeps_two_digit_department_padding(self):
        self.assertEqual(MODULE.french_insee({"NDEPT": "002", "INSEE": "123"}), "02123")
        self.assertEqual(MODULE.french_insee({"NDEPT": "034", "INSEE": "172"}), "34172")
        self.assertEqual(MODULE.french_insee({"NDEPT": "971", "INSEE": "05"}), "971005")

    def test_french_csv_rows_reads_semicolon_cp1252_archive(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "balance.zip"
            with zipfile.ZipFile(path, "w") as archive:
                archive.writestr("balance.csv", "CATEG;NDEPT;INSEE;LBUDG\r\nCommune;034;172;Béziers\r\n".encode("cp1252"))
            rows = list(MODULE.french_csv_rows(path))
            self.assertEqual(rows, [(2, "balance.csv", {"CATEG": "Commune", "NDEPT": "034", "INSEE": "172", "LBUDG": "Béziers"})])

    def test_france_census_fills_communes_without_duplicate_function_facts(self):
        fields = ["CATEG", "NDEPT", "INSEE", "LBUDG", "CREGI", "FONCTION", "COMPTE", "CBUDG", "NOMEN", "OBNETDEB", "OBNETCRE", "SD", "SC"]
        functional = [dict(zip(fields, ["Commune", "001", "001", "Alpha", "84", "01", "641", "1", "M57", "10", "0", "10", "0"]))]
        census = [
            dict(zip(fields, ["Commune", "001", "001", "Alpha", "84", "", "641", "1", "M57", "10", "0", "10", "0"])),
            dict(zip(fields, ["Commune", "001", "002", "Beta", "84", "", "641", "1", "M57", "20", "0", "20", "0"])),
        ]
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            paths = {}
            for detail, rows in (("function", functional), ("census", census)):
                path = root / f"{detail}.zip"
                stream = io.StringIO(newline="")
                writer = csv.DictWriter(stream, fieldnames=fields, delimiter=";", lineterminator="\r\n")
                writer.writeheader(); writer.writerows(rows)
                with zipfile.ZipFile(path, "w") as archive:
                    archive.writestr(f"{detail}.csv", stream.getvalue().encode("cp1252"))
                paths[detail] = path
            sources = [
                {"id": "census", "kind": "actual_and_balance", "detail": "census", "url": "https://example.test/census", "filename": "census.zip"},
                {"id": "function", "kind": "actual_and_balance", "detail": "function", "url": "https://example.test/function", "filename": "function.zip"},
            ]
            args = Namespace(source_file=[f"census={paths['census']}", f"function={paths['function']}"], refresh=False, offline=True, cache_dir=root, api_workers=1, openbudget_token=None, max_entities=None)
            bundle = MODULE.JsonlBundle(root / "output", "none", 0)
            context = MODULE.Context({}, args, bundle)
            result = MODULE.run_france(context, "FRA", {"year": 2025, "currency": "EUR", "sources": sources, "coverage": "test"})
            bundle.close()
            entities = [json.loads(line) for line in (root / "output/public_entities.jsonl").read_text().splitlines()]
            facts = [json.loads(line) for line in (root / "output/municipal_budget_line_facts.jsonl").read_text().splitlines()]
            self.assertEqual(result, {"entities": 2, "functional_entities": 1})
            self.assertEqual({row["public_entity_id"] for row in entities}, {"FR:01001", "FR:01002"})
            self.assertEqual([(row["public_entity_id"], row["source_id"]) for row in facts], [("FR:01001", "function"), ("FR:01002", "census")])

    def test_english_variable_side_is_deterministic(self):
        self.assertEqual(MODULE.uk_side("RO1_sales_fees_income"), "revenue")
        self.assertEqual(MODULE.uk_side("RO6_reserves_total"), "financing")
        self.assertEqual(MODULE.uk_side("RO2_highways_total_exp"), "expenditure")

    def test_source_year_lists_filter_devolved_uk_collections(self):
        cfg = {
            "year": 2026,
            "years": [2024, 2025, 2026],
            "sources": [
                {"id": "england", "years": [2024, 2025]},
                {"id": "scotland", "year": 2026},
            ],
        }
        self.assertEqual([row["id"] for row in MODULE.config_for_year(cfg, 2025)["sources"]], ["england"])
        self.assertEqual([row["id"] for row in MODULE.config_for_year(cfg, 2026)["sources"]], ["scotland"])

    def test_scotland_pobe_loads_leaf_rows_and_normalizes_source_signs(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            paths = {}
            for kind in ("revenue", "capital"):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Aberdeen City"
                sheet.cell(1, 2, "Service section").font = Font(bold=True)
                sheet.cell(1, 4, 10)
                sheet.cell(2, 2, "Leaf expenditure")
                sheet.cell(2, 4, 11)
                sheet.cell(2, 8, 3)
                sheet.cell(2, 9, -4)
                if kind == "capital":
                    sheet.cell(2, 10, 5)
                    sheet.cell(2, 11, 6)
                sheet.cell(3, 2, "Total service").font = Font(bold=True)
                sheet.cell(3, 4, 12)
                sheet.cell(3, 8, 100)
                sheet.cell(3, 9, 100)
                path = root / f"{kind}.xlsx"
                workbook.save(path)
                paths[kind] = path
            sources = [
                {"id": "revenue", "collection": "scotland", "kind": "revenue", "filename": "revenue.xlsx", "url": "https://example.test/revenue"},
                {"id": "capital", "collection": "scotland", "kind": "capital", "filename": "capital.xlsx", "url": "https://example.test/capital"},
            ]
            args = Namespace(
                source_file=[f"revenue={paths['revenue']}", f"capital={paths['capital']}"],
                refresh=False, offline=True, cache_dir=root, api_workers=1,
                openbudget_token=None, max_entities=1,
            )
            output = root / "output"
            bundle = MODULE.JsonlBundle(output, "none", 0)
            context = MODULE.Context({}, args, bundle)
            result = MODULE.run_scotland(context, "GBR", {
                "year": 2026, "currency": "GBP", "sources": sources, "coverage": "test",
            })
            bundle.close()
            facts = [json.loads(line) for line in (output / "municipal_budget_line_facts.jsonl").read_text().splitlines()]
            self.assertEqual(result, {"entities": 1, "facts": 6, "fiscal_years": [2025, 2026, 2027, 2028]})
            self.assertEqual({row["public_entity_id"] for row in facts}, {"GB:S12000033"})
            self.assertEqual({row["budget_side"] for row in facts}, {"revenue", "expenditure"})
            self.assertIn("4000", {row["amount_local"] for row in facts})
            self.assertEqual(MODULE.validate_bundle(output)["status"], "passed")

    def test_wales_statswales_keeps_unitary_authority_thousand_values(self):
        fields = [
            "Data values", "Data description_reference", "Year", "Authority",
            "Authority_reference", "Service", "Service_reference", "Service_hierarchy",
        ]
        rows = [
            {"Data values": "12.5", "Data description_reference": "1", "Year": "2026-27", "Authority": "Cardiff", "Authority_reference": "W06000015", "Service": "Libraries", "Service_reference": "45", "Service_hierarchy": "902"},
            {"Data values": "99", "Data description_reference": "2", "Year": "2026-27", "Authority": "Cardiff", "Authority_reference": "W06000015", "Service": "Libraries", "Service_reference": "45", "Service_hierarchy": "902"},
            {"Data values": "100", "Data description_reference": "1", "Year": "2026-27", "Authority": "Total Wales", "Authority_reference": "600", "Service": "Libraries", "Service_reference": "45", "Service_hierarchy": "902"},
        ]
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            path = root / "wales.csv"
            with path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=fields)
                writer.writeheader(); writer.writerows(rows)
            source = {"id": "wales", "collection": "wales", "kind": "revenue", "filename": "wales.csv", "url": "https://example.test/wales"}
            args = Namespace(
                source_file=[f"wales={path}"], refresh=False, offline=True,
                cache_dir=root, api_workers=1, openbudget_token=None, max_entities=None,
            )
            output = root / "output"
            bundle = MODULE.JsonlBundle(output, "none", 0)
            context = MODULE.Context({}, args, bundle)
            result = MODULE.run_wales(context, "GBR", {
                "year": 2026, "currency": "GBP", "sources": [source], "coverage": "test",
            })
            bundle.close()
            facts = [json.loads(line) for line in (output / "municipal_budget_line_facts.jsonl").read_text().splitlines()]
            self.assertEqual(result, {"entities": 1, "facts": 1, "fiscal_years": [2026]})
            self.assertEqual(facts[0]["public_entity_id"], "GB:W06000015")
            self.assertEqual(facts[0]["amount_local"], "12500.0")
            self.assertEqual(MODULE.validate_bundle(output)["status"], "passed")

    def test_nested_api_record_detection(self):
        rows = [{"budget_code": "1"}]
        self.assertIs(MODULE.find_record_list({"data": {"results": rows}}), rows)

    def test_ukraine_directory_selects_communities_and_kyiv(self):
        rows = [
            {"codebudg": "100", "details": 1, "beginDate": "2024-01-01", "endDate": None, "namebudg": "Бюджет Test територіальної громади"},
            {"codebudg": "2600000000", "details": 1, "beginDate": "2024-01-01", "endDate": None, "namebudg": "Бюджет міста Києва"},
            {"codebudg": "200", "details": 1, "beginDate": "2024-01-01", "endDate": None, "namebudg": "Обласний бюджет"},
        ]
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "budgets.json"
            path.write_text(json.dumps(rows), encoding="utf-8")
            self.assertEqual([row["codebudg"] for row in MODULE._ukraine_budget_directory(path, 2025)], ["100", "2600000000"])

    def test_paraguay_boost_keeps_only_municipal_sheet_and_three_stages(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); path = root / "boost.xlsx"
            workbook = Workbook(); sheet = workbook.active; sheet.title = "Municipalidades"
            sheet.append(["YEAR", "ADMIN1", "ADMIN2", "ECON4", "ECON5", "ECON6", "geo", "approved", "MODIFIED", "PAID"])
            sheet.append([2022, "30 - Municipalidades", "30.001 - Asunción", "100 - Servicios", "110 - Remuneraciones", "111 - Sueldos", None, 10, 12, 9])
            central = workbook.create_sheet("Central"); central.append(["YEAR", "ADMIN2", "PAID"]); central.append([2022, "Central", 999])
            workbook.save(path)
            source = {"id": "py", "kind": "budget_and_actual", "filename": "boost.xlsx", "sheet": "Municipalidades", "url": "https://example.test/boost"}
            args = Namespace(source_file=[f"py={path}"], refresh=False, offline=True, cache_dir=root, api_workers=1, openbudget_token=None, max_entities=None)
            output = root / "output"; bundle = MODULE.JsonlBundle(output, "none", 0); context = MODULE.Context({}, args, bundle)
            result = MODULE.run_paraguay_boost(context, "PRY", {"year": 2022, "currency": "PYG", "sources": [source], "coverage": "test"})
            bundle.close(); facts = [json.loads(line) for line in (output / "municipal_budget_line_facts.jsonl").read_text().splitlines()]
            self.assertEqual(result["entities"], 1); self.assertEqual(result["facts"], 3)
            self.assertEqual({row["budget_stage"] for row in facts}, {"enacted", "revised", "actual"})
            self.assertEqual({row["public_entity_id"] for row in facts}, {"PY:30001"})
            self.assertEqual(MODULE.validate_bundle(output)["status"], "passed")


if __name__ == "__main__":
    unittest.main()
