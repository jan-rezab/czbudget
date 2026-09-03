#!/usr/bin/env python3
"""Build the revenue deep-dive dataset from OECD and Eurostat sources."""

from __future__ import annotations

import csv
import io
import json
import tempfile
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "country-revenue.v1.json"
COUNTRIES = ["CZE", "DEU", "DNK", "FIN", "FRA", "GBR", "POL", "SWE", "CHE", "UKR", "USA", "BRA", "ESP", "JPN", "NLD", "NOR", "GRC"]
EUROSTAT_GEO = {
    "CZE": "CZ", "DEU": "DE", "DNK": "DK", "FRA": "FR", "GBR": "UK",
    "POL": "PL", "SWE": "SE", "CHE": "CH", "UKR": "UA", "FIN": "FI", "ESP": "ES", "NLD": "NL", "NOR": "NO", "GRC": "EL",
}
OECD_API = "https://sdmx.oecd.org/public/rest/data/OECD.CTP.TPS,DSD_REV_COMP_GLOBAL@DF_RSGLOBAL,2.1"
OECD_TRANSFER_URL = "https://www.oecd.org/content/dam/oecd/en/topics/policy-sub-issues/fiscal-federalism-network/table17_intergov_rev-tot_rev-gov-rev-by-sector.xlsx"
EUROSTAT_API = "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/env_ac_tax"
SOURCE_CATEGORIES = ["1000", "2000", "3000", "4000", "5000", "6000"]
DETAIL_CATEGORIES = ["1100", "1200", "5111", "5121"]
SECTORS = {"S1311": "central", "S1312": "state", "S1313": "local", "S1314": "social_security"}


def download(url: str, timeout: int = 240) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "PublicSpendingData/1.0"})
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read()


def number(value: object) -> float | None:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed


def rounded(value: float | None, digits: int = 2) -> float | None:
    return None if value is None else round(value, digits)


def fetch_oecd_rows() -> list[dict[str, str]]:
    areas = "+".join(COUNTRIES)
    sectors = "+".join(["S13", *SECTORS])
    categories = "+".join([*(f"T_{code}" for code in SOURCE_CATEGORIES + DETAIL_CATEGORIES), "_T"])
    selection = f"{areas}.TAX_REV.{sectors}.{categories}._T.XDC.A"
    query = urllib.parse.urlencode({
        "startPeriod": "2007", "endPeriod": "2023",
        "dimensionAtObservation": "AllDimensions", "format": "csvfilewithlabels",
    })
    payload = download(f"{OECD_API}/{selection}?{query}").decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(payload)))


def fetch_transfer_shares() -> dict[str, dict]:
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as handle:
        handle.write(download(OECD_TRANSFER_URL))
        handle.flush()
        book = load_workbook(handle.name, read_only=True, data_only=True)
        sheet = book["intergov_rev_%tot_rev-part_con"]
        rows = list(sheet.iter_rows(values_only=True))
    years = list(rows[1][3:])
    transfers: dict[str, dict] = {}
    for row in rows[2:]:
        code, level = row[0], row[2]
        if code not in COUNTRIES or level != "Local":
            continue
        available = [(int(year), number(value)) for year, value in zip(years, row[3:])]
        available = [(year, value) for year, value in available if year and value is not None]
        if available:
            year, value = available[-1]
            transfers[code] = {"year": year, "local_revenue_from_transfers_pct": rounded(value)}
    return transfers


def fetch_environmental_share(code: str) -> dict | None:
    geo = EUROSTAT_GEO.get(code)
    if not geo:
        return None
    query = urllib.parse.urlencode({"tax": "ENV", "unit": "PC_TSCO_X_ISCO", "geo": geo})
    try:
        payload = json.loads(download(f"{EUROSTAT_API}?{query}", timeout=90))
    except Exception:
        return None
    time_index = payload.get("dimension", {}).get("time", {}).get("category", {}).get("index", {})
    if isinstance(time_index, list):
        time_index = {year: index for index, year in enumerate(time_index)}
    values = payload.get("value", {})
    points = []
    for year, index in time_index.items():
        value = number(values.get(str(index), values.get(index)))
        if value is not None:
            points.append((int(year), value))
    if not points:
        return None
    year, value = max(points)
    return {"year": year, "share_of_tax_and_social_contributions_pct": rounded(value)}


def parse_oecd(rows: list[dict[str, str]]) -> dict[str, dict]:
    values: dict[str, dict] = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    labels: dict[str, str] = {}
    for row in rows:
        value = number(row.get("OBS_VALUE"))
        if value is None:
            continue
        code = row["REF_AREA"]
        year = int(row["TIME_PERIOD"])
        sector = row["SECTOR"]
        revenue_code = row.get("REVENUE_CODE") or row.get("STANDARD_REVENUE", "").removeprefix("T_")
        labels[revenue_code] = row.get("Revenue category", revenue_code)
        values[code][year][sector][revenue_code] = value

    result: dict[str, dict] = {}
    for code in COUNTRIES:
        timeline = []
        detail_timeline = []
        for year in sorted(values[code]):
            general = values[code][year].get("S13", {})
            total = general.get("TOTALTAX")
            if not total:
                continue
            categories = {key: general.get(key, 0.0) for key in SOURCE_CATEGORIES}
            if sum(categories.values()) <= 0:
                continue
            shares = {key: rounded(value / total * 100) for key, value in categories.items()}
            timeline.append({"year": year, "shares": shares})
            detail_codes = {
                "personal_income": "1100",
                "corporate_income": "1200",
                "vat": "5111",
                "excise": "5121",
                "social_security": "2000",
                "property": "4000",
            }
            if all(revenue_code in general for revenue_code in detail_codes.values()):
                detail_shares = {
                    key: rounded(general[revenue_code] / total * 100)
                    for key, revenue_code in detail_codes.items()
                }
                detail_shares["other"] = rounded(100 - sum(detail_shares.values()))
                detail_timeline.append({"year": year, "shares": detail_shares})
        if not timeline:
            continue
        latest = timeline[-1]
        latest_year = latest["year"]
        general = values[code][latest_year]["S13"]
        total = general["TOTALTAX"]
        levels = {
            key: values[code][latest_year].get(sector, {}).get("TOTALTAX")
            for sector, key in SECTORS.items()
        }
        level_total = sum(value for value in levels.values() if value is not None)
        level_shares = {
            key: rounded(value / level_total * 100) if value is not None and level_total else None
            for key, value in levels.items()
        }
        personal = general.get("1100", 0.0)
        corporate = general.get("1200", 0.0)
        labour = personal + general.get("2000", 0.0) + general.get("3000", 0.0)
        capital = corporate + general.get("4000", 0.0)
        consumption = general.get("5000", 0.0)
        proxy_other = max(0.0, total - labour - capital - consumption)
        proxy_total = labour + capital + consumption + proxy_other
        proxy = {
            "labour": rounded(labour / proxy_total * 100),
            "capital": rounded(capital / proxy_total * 100),
            "consumption": rounded(consumption / proxy_total * 100),
            "other": rounded(proxy_other / proxy_total * 100),
        }
        result[code] = {
            "latest_year": latest_year,
            "tax_mix": latest["shares"],
            "tax_detail": detail_timeline[-1]["shares"] if detail_timeline else None,
            "government_levels": level_shares,
            "economic_base_proxy": proxy,
            "timeline": timeline,
            "tax_detail_timeline": detail_timeline,
        }
    return result


def main() -> None:
    countries = parse_oecd(fetch_oecd_rows())
    transfers = fetch_transfer_shares()
    for code, country in countries.items():
        country["municipal_transfers"] = transfers.get(code)
        country["environmental_taxes"] = fetch_environmental_share(code)
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    payload = {
        "schema_version": "1.0.0",
        "contract": "country-revenue.v1",
        "generated_at": generated_at,
        "scope": {
            "tax_mix": "OECD tax revenue, general government, broad tax categories",
            "tax_detail": "OECD tax revenue, general government, selected standard subcategories as shares of total tax revenue",
            "government_levels": "OECD tax revenue by receiving government subsector, normalised to 100",
            "environmental_taxes": "Eurostat environmental taxes as a share of taxes and actual social contributions",
            "municipal_transfers": "OECD intergovernmental transfer revenue as a share of local-government total revenue",
        },
        "definitions": {
            "source_categories": {"1000": "income", "2000": "social_security", "3000": "payroll", "4000": "property", "5000": "consumption", "6000": "other"},
            "detail_categories": {"personal_income": "1100", "corporate_income": "1200", "vat": "5111", "excise": "5121", "social_security": "2000", "property": "4000", "other": "remainder_to_100"},
            "economic_base_proxy": "Analytical grouping: personal income + social contributions + payroll as labour; corporate income + property as capital; goods and services as consumption. Personal income and property taxes are not pure economic-base measures.",
            "non_additivity": "Environmental taxes overlap OECD goods-and-services and property categories and are shown as a memorandum item, not an additional slice.",
        },
        "sources": [
            {"id": "oecd_revenue", "title": "OECD Global Revenue Statistics — Comparative tax revenues", "url": "https://www.oecd.org/en/data/datasets/global-revenue-statistics-database.html", "api": OECD_API},
            {"id": "oecd_decentralisation", "title": "OECD Fiscal Decentralisation Database — Table 17", "url": "https://www.oecd.org/en/data/datasets/oecd-fiscal-decentralisation-database.html", "asset": OECD_TRANSFER_URL},
            {"id": "eurostat_environment", "title": "Eurostat environmental tax revenues (env_ac_tax)", "url": "https://ec.europa.eu/eurostat/en/web/products-datasets/-/ENV_AC_TAX", "api": EUROSTAT_API},
        ],
        "countries": countries,
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT.relative_to(ROOT)} with {len(countries)} country profiles")


if __name__ == "__main__":
    main()
