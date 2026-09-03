#!/usr/bin/env python3
"""Build the EU budget country-flow dataset from the Commission workbook."""

from __future__ import annotations

import argparse
import json
import math
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "eu-budget-flows.v1.json"
SOURCE_DIR = ROOT.parent / "data" / "source_cache" / "eu_budget_flows"
SOURCE_FILE = SOURCE_DIR / "eu-budget-spending-and-revenue-2000-2024.xlsx"
SOURCE_PAGE = "https://commission.europa.eu/strategy-and-policy/eu-budget/long-term-eu-budget/2021-2027/spending-and-revenue_en"
SOURCE_URL = "https://commission.europa.eu/document/download/45d0623a-529e-44d2-aae4-2ca9bac87ec3_en?filename=eu_budget_spending_and_revenue_2000-2023.xlsx"

MEMBERS = [
    ("AUT", "AT", "Rakousko", "Austria"), ("BEL", "BE", "Belgie", "Belgium"),
    ("BGR", "BG", "Bulharsko", "Bulgaria"), ("HRV", "HR", "Chorvatsko", "Croatia"),
    ("CYP", "CY", "Kypr", "Cyprus"), ("CZE", "CZ", "Česko", "Czechia"),
    ("DNK", "DK", "Dánsko", "Denmark"), ("EST", "EE", "Estonsko", "Estonia"),
    ("FIN", "FI", "Finsko", "Finland"), ("FRA", "FR", "Francie", "France"),
    ("DEU", "DE", "Německo", "Germany"), ("GRC", "EL", "Řecko", "Greece"),
    ("HUN", "HU", "Maďarsko", "Hungary"), ("IRL", "IE", "Irsko", "Ireland"),
    ("ITA", "IT", "Itálie", "Italy"), ("LVA", "LV", "Lotyšsko", "Latvia"),
    ("LTU", "LT", "Litva", "Lithuania"), ("LUX", "LU", "Lucembursko", "Luxembourg"),
    ("MLT", "MT", "Malta", "Malta"), ("NLD", "NL", "Nizozemsko", "Netherlands"),
    ("POL", "PL", "Polsko", "Poland"), ("PRT", "PT", "Portugalsko", "Portugal"),
    ("ROU", "RO", "Rumunsko", "Romania"), ("SVK", "SK", "Slovensko", "Slovakia"),
    ("SVN", "SI", "Slovinsko", "Slovenia"), ("ESP", "ES", "Španělsko", "Spain"),
    ("SWE", "SE", "Švédsko", "Sweden"),
]

HEADING_LABELS = {
    "1": ("Jednotný trh, inovace a digitální oblast", "Single Market, Innovation and Digital"),
    "2": ("Soudržnost, odolnost a hodnoty", "Cohesion, Resilience and Values"),
    "3": ("Přírodní zdroje a životní prostředí", "Natural Resources and Environment"),
    "4": ("Migrace a správa hranic", "Migration and Border Management"),
    "5": ("Bezpečnost a obrana", "Security and Defence"),
    "6": ("Sousedství a svět", "Neighbourhood and the World"),
    "7": ("Evropská veřejná správa", "European Public Administration"),
    "S": ("Nástroje solidarity", "Solidarity instruments"),
}

MEMBERSHIP_START = {
    "AUT": 1995, "BEL": 1958, "BGR": 2007, "HRV": 2013, "CYP": 2004, "CZE": 2004,
    "DNK": 1973, "EST": 2004, "FIN": 1995, "FRA": 1958, "DEU": 1958, "GRC": 1981,
    "HUN": 2004, "IRL": 1973, "ITA": 1958, "LVA": 2004, "LTU": 2004, "LUX": 1958,
    "MLT": 2004, "NLD": 1958, "POL": 2004, "PRT": 1986, "ROU": 2007, "SVK": 2004,
    "SVN": 2004, "ESP": 1986, "SWE": 1995,
}


def clean(value) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def number(value):
    if value is None or isinstance(value, bool):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(result):
        return None
    return round(result, 6)


def find_row(ws, phrase: str, exact: bool = False):
    target = clean(phrase).casefold()
    for row in ws.iter_rows():
        for cell in row[:4]:
            value = clean(cell.value).casefold()
            if (exact and value == target) or (not exact and target in value):
                return cell.row
    return None


def find_row_after(ws, phrase: str, start_row: int):
    target = clean(phrase).casefold()
    for row in ws.iter_rows(min_row=start_row):
        if any(target in clean(cell.value).casefold() for cell in row[:4]):
            return row[0].row
    return None


def header_map(ws, row: int):
    return {clean(cell.value): cell.column for cell in ws[row] if clean(cell.value)}


def nearest_header(ws, start_row: int, country_codes: set[str]):
    for row in range(start_row, 0, -1):
        mapping = header_map(ws, row)
        if len(country_codes.intersection(mapping)) >= 20:
            return mapping
    raise ValueError(f"No country header found near row {start_row} in {ws.title}")


def row_value(ws, row: int | None, column: int | None):
    if not row or not column:
        return None
    return number(ws.cell(row, column).value)


def heading_rows(ws, total_row: int, start_row: int = 1):
    headings = {}
    for row in range(start_row, total_row):
        code = clean(ws.cell(row, 1).value).rstrip(".")
        if code in HEADING_LABELS and code not in headings:
            headings[code] = row
    return headings


def download(force: bool = False):
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    if SOURCE_FILE.exists() and not force:
        return SOURCE_FILE
    request = urllib.request.Request(SOURCE_URL, headers={"User-Agent": "PublicSpendingData/1.0"})
    with urllib.request.urlopen(request, timeout=120) as response:
        SOURCE_FILE.write_bytes(response.read())
    return SOURCE_FILE


def build(source: Path):
    workbook = load_workbook(source, read_only=False, data_only=True)
    codes = {member[1] for member in MEMBERS}
    countries = {iso3: {"iso3": iso3, "eu_code": eu, "name_cs": cs, "name_en": en,
                        "member_since": MEMBERSHIP_START[iso3], "series": []}
                 for iso3, eu, cs, en in MEMBERS}

    available_years = sorted(int(name) for name in workbook.sheetnames if name.isdigit())
    for year in available_years:
        ws = workbook[str(year)]
        expenditure_row = find_row(ws, "TOTAL EXPENDITURE", exact=True)
        national_row = find_row(ws, "TOTAL National contribution")
        own_resources_row = find_row(ws, "TOTAL Own resources")
        customs_row = find_row(ws, "Customs duties", exact=True) or find_row_after(ws, "Customs duties", national_row)
        sugar_row = find_row(ws, "Sugar levies", exact=True) or find_row_after(ws, "Sugar levies", national_row)
        gni_row = find_row(ws, "Gross National Income")
        ngeu_row = find_row(ws, "TOTAL NGEU", exact=True)
        if not expenditure_row or not national_row:
            raise ValueError(f"Missing core total row in sheet {year}")

        spend_header = nearest_header(ws, expenditure_row, codes)
        revenue_header = nearest_header(ws, national_row, codes)
        ngeu_header = nearest_header(ws, ngeu_row, codes) if ngeu_row else {}
        main_headings = heading_rows(ws, expenditure_row) if year >= 2021 else {}

        for iso3, eu_code, *_ in MEMBERS:
            mff = row_value(ws, expenditure_row, spend_header.get(eu_code))
            ngeu = row_value(ws, ngeu_row, ngeu_header.get(eu_code)) or 0
            national = row_value(ws, national_row, revenue_header.get(eu_code))
            own = row_value(ws, own_resources_row, revenue_header.get(eu_code))
            customs = row_value(ws, customs_row, revenue_header.get(eu_code)) or 0
            sugar = row_value(ws, sugar_row, revenue_header.get(eu_code)) or 0
            gni = row_value(ws, gni_row, revenue_header.get(eu_code))
            if all(value is None for value in (mff, national, own, gni)):
                continue
            row = {
                "year": year,
                "allocated_spending_m_eur": round((mff or 0) + ngeu, 6),
                "mff_spending_m_eur": mff,
                "ngeu_spending_m_eur": ngeu or None,
                "national_contribution_m_eur": national,
                "total_own_resources_m_eur": own,
                "traditional_own_resources_m_eur": round(customs + sugar, 6),
                "gni_m_eur": gni,
                "accounting_difference_m_eur": round((mff or 0) + ngeu - national, 6) if national is not None else None,
            }
            if year >= 2021:
                breakdown = []
                ngeu_headings = heading_rows(ws, ngeu_row, expenditure_row + 1) if ngeu_row else {}
                for code, (label_cs, label_en) in HEADING_LABELS.items():
                    base = row_value(ws, main_headings.get(code), spend_header.get(eu_code)) or 0
                    extra = row_value(ws, ngeu_headings.get(code), ngeu_header.get(eu_code)) or 0
                    amount = base + extra
                    if amount:
                        breakdown.append({"code": code, "label_cs": label_cs, "label_en": label_en,
                                          "amount_m_eur": round(amount, 6)})
                remainder = round(row["allocated_spending_m_eur"] - sum(item["amount_m_eur"] for item in breakdown), 6)
                if abs(remainder) >= 0.01:
                    breakdown.append({"code": "O", "label_cs": "Jiné a nezařazené", "label_en": "Other and unallocated",
                                      "amount_m_eur": remainder})
                row["spending_breakdown"] = breakdown
            countries[iso3]["series"].append(row)

    result = {
        "contract": "eu-budget-flows.v1",
        "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "currency": "EUR",
        "unit": "million",
        "period": {"first": min(available_years), "last": max(available_years)},
        "coverage": {"current_member_states": len(MEMBERS), "country_year_rows": sum(len(c["series"]) for c in countries.values())},
        "definitions": {
            "allocated_spending_m_eur": "Country-attributed EU expenditure plus separately reported NextGenerationEU expenditure.",
            "national_contribution_m_eur": "VAT-, GNI- and plastics-based own resources plus balances and adjustments; excludes traditional own resources.",
            "total_own_resources_m_eur": "National contribution plus traditional own resources, principally customs duties.",
            "traditional_own_resources_m_eur": "Sugar levies plus customs duties reported for the country.",
            "accounting_difference_m_eur": "Allocated spending minus national contribution; a transparent arithmetic comparison, not the Commission's operating budgetary balance.",
        },
        "sources": {
            "publisher": "European Commission, Directorate-General for Budget",
            "page_url": SOURCE_PAGE,
            "download_url": SOURCE_URL,
            "source_file": str(SOURCE_FILE.relative_to(ROOT.parent)),
            "published": "2025-09-25",
            "title": "EU spending and revenue — Data 2000–2024",
        },
        "countries": list(countries.values()),
    }
    OUTPUT.write_text(json.dumps(result, ensure_ascii=False, separators=(",", ":")) + "\n")
    print(f"Wrote {OUTPUT.relative_to(ROOT)}: {result['coverage']['country_year_rows']} country-years")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path)
    parser.add_argument("--force-download", action="store_true")
    args = parser.parse_args()
    build(args.input or download(args.force_download))


if __name__ == "__main__":
    main()
