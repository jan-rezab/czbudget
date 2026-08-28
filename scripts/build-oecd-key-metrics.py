#!/usr/bin/env python3
"""Build the cross-topic OECD metric overlay used by country and comparison views.

The contract deliberately keeps each observation's reference year and definition.
It never fills a missing OECD observation with zero or with a non-OECD estimate.
"""

from __future__ import annotations

import csv
import io
import json
import tempfile
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "oecd-key-metrics.v1.json"
COUNTRIES = ["CZE", "DEU", "DNK", "FIN", "FRA", "GBR", "POL", "SWE", "CHE", "UKR", "USA", "BRA", "ESP", "JPN", "NLD", "NOR", "GRC"]
AREAS = "+".join(COUNTRIES)
BASE = "https://sdmx.oecd.org/public/rest/data"
AUTONOMY_URL = "https://www.oecd.org/content/dam/oecd/en/topics/policy-sub-issues/fiscal-federalism-network/table_taxpower-2025.xlsx"

FLOWS = {
    "revenue": "OECD.CTP.TPS,DSD_REV_COMP_GLOBAL@DF_RSGLOBAL,2.1",
    "labour": "OECD.CTP.TPS,DSD_TAX_WAGES_COMP@DF_TW_COMP,2.1",
    "corporate": "OECD.CTP.TPS,DSD_TAX_CIT@DF_CIT,1.0",
    "effective_corporate": "OECD.CTP.TPS,DSD_ETR@DF_ETR_BASELINE,2.1",
    "carbon": "OECD.CTP.TPS,DSD_NECR@DF_NECRS,1.1",
    "distribution": "OECD.WISE.INE,DSD_WISE_IDD@DF_IDD,1.0",
    "social": "OECD.ELS.SPD,DSD_SOCX_AGG@DF_SOCX_AGG,1.0",
    "pensions": "OECD.ELS.SPD,DSD_PAG@DF_PRR,1.0",
    "government_employment": "OECD.GOV.GIP,DSD_GOV@DF_GOV_EMPPS_REP_2025,1.0",
    "procurement": "OECD.GOV.GIP,DSD_GOV@DF_GOV_PPROC_2025,1.0",
    "wellbeing": "OECD.WISE.WDP,DSD_HSL@DF_HSL_CWB,1.1",
}

METRICS = {
    "tax_to_gdp": ("tax", "% HDP", "% of GDP", "Daňové příjmy včetně povinných příspěvků sociálního zabezpečení.", "Tax revenue including compulsory social-security contributions."),
    "labour_tax_wedge_single": ("tax", "%", "%", "Svobodný bez dětí, 100 % průměrné mzdy; daň z příjmu a příspěvky zaměstnance i zaměstnavatele vůči nákladům práce.", "Single person without children at 100% of the average wage; income tax plus employee and employer contributions as a share of labour cost."),
    "labour_tax_wedge_family": ("tax", "%", "%", "Manželský pár se dvěma dětmi, jeden příjem na 100 % průměrné mzdy; stejná definice daňového klínu.", "Married couple with two children and one earner at 100% of the average wage; same tax-wedge definition."),
    "corporate_statutory_rate": ("tax", "%", "%", "Kombinovaná zákonná sazba daně z příjmů právnických osob, ústřední a nižší úrovně vlády.", "Combined central and sub-central statutory corporate income-tax rate."),
    "corporate_eatr": ("tax", "%", "%", "Modelová efektivní průměrná sazba pro ziskovou investici v základním scénáři OECD.", "Modelled effective average tax rate for a profitable investment in the OECD baseline scenario."),
    "corporate_emtr": ("tax", "%", "%", "Modelová efektivní mezní sazba pro investici na hraně rentability v základním scénáři OECD.", "Modelled effective marginal tax rate for a break-even investment in the OECD baseline scenario."),
    "local_tax_autonomy": ("tax", "% místních daní", "% of local tax revenue", "Podíl místních daňových příjmů, u nichž místní vláda ovlivňuje sazbu nebo základ; kategorie a–c OECD.", "Share of local tax revenue over which local government controls the rate or base; OECD categories a–c."),
    "net_carbon_rate": ("tax", "EUR/t CO₂e", "EUR/t CO₂e", "Vážený průměr daní a obchodovatelných cen po odečtení fosilních subvencí, ve stálých cenách roku 2023.", "Weighted mean of taxes and tradable permit prices net of fossil-fuel subsidies, in constant 2023 prices."),
    "disposable_gini": ("distribution", "0–1", "0–1", "Giniho koeficient disponibilního příjmu po daních a transferech; 0 znamená rovnost.", "Gini coefficient of disposable income after taxes and transfers; 0 denotes equality."),
    "market_gini": ("distribution", "0–1", "0–1", "Giniho koeficient tržního příjmu před daněmi a transfery.", "Gini coefficient of market income before taxes and transfers."),
    "poverty_rate": ("distribution", "% populace", "% of population", "Podíl lidí pod 50 % mediánu disponibilního příjmu.", "Share of people below 50% of median disposable income."),
    "social_spending": ("social", "% HDP", "% of GDP", "Veřejné sociální výdaje v databázi SOCX.", "Public social expenditure in the SOCX database."),
    "pension_replacement_aw100": ("pensions", "% výdělku", "% of earnings", "Modelová čistá náhradová míra povinného důchodu pro pracovníka se 100 % průměrné mzdy.", "Modelled net mandatory pension replacement rate for a worker at 100% of average earnings."),
    "government_employment": ("government", "% zaměstnanosti", "% of employment", "Zaměstnanost v sektoru vládních institucí jako podíl celkové zaměstnanosti.", "Employment in general government as a share of total employment."),
    "procurement_gdp": ("government", "% HDP", "% of GDP", "Výdaje vládních institucí na veřejné zakázky.", "General-government procurement expenditure."),
    "housing_affordability": ("wellbeing", "% příjmu", "% of income", "Podíl hrubého upraveného disponibilního příjmu, který domácnostem zbývá po nákladech na bydlení; vyšší hodnota značí dostupnější bydlení.", "Share of gross adjusted disposable income remaining after housing costs; a higher value denotes more affordable housing."),
    "life_satisfaction": ("wellbeing", "0–10", "0–10", "Průměrná sebehodnocená životní spokojenost na škále 0 až 10.", "Average self-reported life satisfaction on a 0-to-10 scale."),
    "pisa_math": ("wellbeing", "bodů", "points", "Průměrný výsledek PISA v matematické gramotnosti.", "Average PISA mathematics performance."),
    "road_deaths": ("wellbeing", "na 100 tis.", "per 100,000", "Úmrtí při silničních nehodách na 100 000 obyvatel.", "Road deaths per 100,000 people."),
}


def download(url: str, timeout: int = 240) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "PublicSpendingData/1.0 (data build)"})
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read()


def rows(flow: str, key: str, *, start: int | None = None) -> list[dict[str, str]]:
    query = {"dimensionAtObservation": "AllDimensions", "format": "csvfilewithlabels"}
    if start is not None:
        query["startPeriod"] = str(start)
    url = f"{BASE}/{FLOWS[flow]}/{key}?{urllib.parse.urlencode(query)}"
    return list(csv.DictReader(io.StringIO(download(url).decode("utf-8-sig"))))


def value(row: dict[str, str]) -> float | None:
    try:
        return float(row.get("OBS_VALUE", ""))
    except (TypeError, ValueError):
        return None


def observation(row: dict[str, str] | None, source_id: str) -> dict | None:
    if not row or value(row) is None:
        return None
    return {"value": round(value(row), 3), "year": int(row["TIME_PERIOD"]), "source_id": source_id}


def latest(items: list[dict[str, str]]) -> dict[str, str] | None:
    valid = [row for row in items if value(row) is not None and str(row.get("TIME_PERIOD", "")).isdigit()]
    return max(valid, key=lambda row: int(row["TIME_PERIOD"])) if valid else None


def by_country(items: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    result: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in items:
        if row.get("REF_AREA") in COUNTRIES:
            result[row["REF_AREA"]].append(row)
    return result


def match(items: list[dict[str, str]], **criteria: str) -> list[dict[str, str]]:
    return [row for row in items if all(row.get(key) == expected for key, expected in criteria.items())]


def load_autonomy() -> dict[str, dict]:
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as handle:
        handle.write(download(AUTONOMY_URL))
        handle.flush()
        sheet = load_workbook(handle.name, read_only=True, data_only=True)["Flat_table"]
        records = list(sheet.iter_rows(values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in records[0]]
    grouped: dict[tuple[str, str, int], dict[str, float]] = defaultdict(dict)
    for raw in records[1:]:
        row = dict(zip(headers, raw))
        code, gov, year = row.get("COU"), row.get("GOV"), row.get("YEAR")
        if code not in COUNTRIES or gov not in {"LOCAL", "STATE"} or not isinstance(year, (int, float)):
            continue
        try:
            grouped[(code, gov.lower(), int(year))][str(row.get("TA_CODE"))] = float(row.get("VALUE"))
        except (TypeError, ValueError):
            continue
    result: dict[str, dict] = defaultdict(dict)
    autonomous = {"a1", "a2", "b1", "b2", "b3", "c"}
    shared = {"d1", "d2", "d3", "d4"}
    for code in COUNTRIES:
        for level in ("local", "state"):
            years = [year for (area, lev, year) in grouped if area == code and lev == level]
            if not years:
                continue
            year = max(years)
            categories = grouped[(code, level, year)]
            result[code][level] = {
                "year": year,
                "autonomous_share_pct": round(sum(categories.get(key, 0) for key in autonomous), 2),
                "shared_share_pct": round(sum(categories.get(key, 0) for key in shared), 2),
                "other_or_central_share_pct": round(sum(value for key, value in categories.items() if key not in autonomous | shared), 2),
                "categories": {key: round(val, 2) for key, val in sorted(categories.items())},
                "source_id": "oecd_tax_autonomy",
            }
    return result


def main() -> None:
    print("Fetching OECD key-metric slices…")
    fetched = {
        "revenue": rows("revenue", f"{AREAS}.TAX_REV.S13._T._T.PT_B1GQ.A", start=2019),
        "labour": rows("labour", f"{AREAS}......A", start=2024),
        "corporate": rows("corporate", f"{AREAS}.A......", start=2023),
        "effective_corporate": rows("effective_corporate", f"{AREAS}.A......", start=2024),
        "carbon": rows("carbon", f"{AREAS}._T._T.NETECR+ECRATE+SUBSID..EUR_TCO2E.MEANW.Q.A"),
        "distribution": rows("distribution", f"{AREAS}.A.......", start=2019),
        "social": rows("social", f"{AREAS}.A.SOCX.PT_B1GQ.ES10._T._T+TP01+TP31+TP41+TP51+TP71+TP82+TP91._Z", start=2019),
        "pensions": rows("pensions", f"{AREAS}.A.....", start=2024),
        "government_employment": rows("government_employment", f"A.{AREAS}....."),
        "procurement": rows("procurement", f"A.{AREAS}....."),
        "wellbeing": rows("wellbeing", f"{AREAS}......", start=2019),
    }
    grouped = {key: by_country(data) for key, data in fetched.items()}
    autonomy = load_autonomy()
    countries: dict[str, dict] = {}

    social_codes = {"_T": "total", "TP01": "old_age_survivors", "TP31": "incapacity", "TP41": "health", "TP51": "family", "TP71": "unemployment", "TP82": "housing", "TP91": "other"}
    wellbeing_codes = {"3_2": "housing_affordability", "3_3": "housing_cost_overburden", "11_1": "life_satisfaction", "6_1": "pisa_math", "6_2": "pisa_reading", "6_3": "pisa_science", "10_3": "road_deaths", "2_1": "employment_rate"}

    for code in COUNTRIES:
        record: dict = {"tax": {}, "distribution": {}, "social": {}, "pensions": {}, "government": {}, "wellbeing": {}, "comparison": {}}
        comparison = record["comparison"]

        obs = observation(latest(grouped["revenue"][code]), "oecd_revenue_statistics")
        record["tax"]["tax_to_gdp"] = obs
        if obs: comparison["tax_to_gdp"] = obs

        labour_rows = grouped["labour"][code]
        scenarios = []
        scenario_keys = sorted({(r.get("HOUSEHOLD_TYPE"), r.get("INCOME_PRINCIPAL"), r.get("INCOME_SPOUSE"), int(r["TIME_PERIOD"])) for r in labour_rows if r.get("TIME_PERIOD", "").isdigit()}, key=lambda item: item[3], reverse=True)
        newest_year = scenario_keys[0][3] if scenario_keys else None
        for household, principal, spouse, year in scenario_keys:
            if year != newest_year:
                continue
            subset = [r for r in labour_rows if r.get("HOUSEHOLD_TYPE") == household and r.get("INCOME_PRINCIPAL") == principal and r.get("INCOME_SPOUSE") == spouse and int(r["TIME_PERIOD"]) == year]
            metrics = {r["MEASURE"].lower(): round(value(r), 3) for r in subset if value(r) is not None and r.get("MEASURE") in {"AV_TW", "MR_TW_PE", "AV_R_EMPEE_SSC", "AV_R_EMPER_SSC", "AV_RITEESSC", "AV_ITR", "NPATR", "NPMTR_PE"}}
            if metrics:
                scenarios.append({"household_type": household, "principal_income": principal, "spouse_income": spouse, "year": year, "metrics": metrics, "source_id": "oecd_taxing_wages"})
        record["tax"]["labour"] = {"year": newest_year, "scenarios": scenarios}
        for metric_code, scenario in {
            "labour_tax_wedge_single": ("S_C0", "AW100", "_Z"),
            "labour_tax_wedge_family": ("C_C2", "AW100", "NOEARN_UNEMP"),
        }.items():
            hit = next((s for s in scenarios if (s["household_type"], s["principal_income"], s["spouse_income"]) == scenario and "av_tw" in s["metrics"]), None)
            if hit:
                comparison[metric_code] = {"value": hit["metrics"]["av_tw"], "year": hit["year"], "source_id": "oecd_taxing_wages"}

        cit_rows = grouped["corporate"][code]
        statutory = latest(match(cit_rows, MEASURE="CIT_C", TARGETING="ST", SECTOR="S13")) or latest(match(cit_rows, MEASURE="CIT", TARGETING="ST", SECTOR="S13"))
        small = latest(match(cit_rows, MEASURE="CIT_C", TARGETING="TA", SECTOR="S13"))
        record["tax"]["corporate"] = {
            "statutory_combined": observation(statutory, "oecd_corporate_tax_statistics"),
            "targeted_small_business": observation(small, "oecd_corporate_tax_statistics"),
        }
        if record["tax"]["corporate"]["statutory_combined"]:
            comparison["corporate_statutory_rate"] = record["tax"]["corporate"]["statutory_combined"]

        etr_rows = match(grouped["effective_corporate"][code], ETR_TAX_BASIS="BASELINE", ETR_SCENARIO="FIXED", ETR_TAX_TYPE="COMPOSITE", REGIME="_Z")
        for measure, target in (("EATR", "corporate_eatr"), ("EMTR", "corporate_emtr")):
            metric_obs = observation(latest(match(etr_rows, MEASURE=measure)), "oecd_corporate_effective_rates")
            record["tax"]["corporate"][measure.lower()] = metric_obs
            if metric_obs: comparison[target] = metric_obs

        carbon = {}
        for measure, target in (("NETECR", "net_effective_rate"), ("ECRATE", "effective_rate"), ("SUBSID", "subsidy")):
            carbon[target] = observation(latest(match(grouped["carbon"][code], MEASURE=measure)), "oecd_net_effective_carbon_rates")
        record["tax"]["carbon"] = carbon
        if carbon["net_effective_rate"]: comparison["net_carbon_rate"] = carbon["net_effective_rate"]
        record["tax"]["autonomy"] = autonomy.get(code, {})
        if autonomy.get(code, {}).get("local"):
            local = autonomy[code]["local"]
            comparison["local_tax_autonomy"] = {"value": local["autonomous_share_pct"], "year": local["year"], "source_id": "oecd_tax_autonomy"}

        idd = match(grouped["distribution"][code], AGE="_T", METHODOLOGY="METH2012", DEFINITION="D_CUR")
        distribution_filters = {
            "disposable_gini": {"MEASURE": "INC_DISP_GINI", "UNIT_MEASURE": "0_TO_1"},
            "market_gini": {"MEASURE": "INC_MRKT_GINI", "UNIT_MEASURE": "0_TO_1"},
            "poverty_rate": {"MEASURE": "PR_INC_DISP", "UNIT_MEASURE": "PT_POP", "POVERTY_LINE": "PL_50"},
            "poverty_gap": {"MEASURE": "PG_INC_DISP", "POVERTY_LINE": "PL_50"},
        }
        for target, criteria in distribution_filters.items():
            metric_obs = observation(latest(match(idd, **criteria)), "oecd_income_distribution")
            record["distribution"][target] = metric_obs
            if metric_obs and target in METRICS: comparison[target] = metric_obs

        for programme, target in social_codes.items():
            metric_obs = observation(latest(match(grouped["social"][code], PROGRAMME_TYPE=programme)), "oecd_socx")
            record["social"][target] = metric_obs
        if record["social"].get("total"): comparison["social_spending"] = record["social"]["total"]

        pension_rows = match(grouped["pensions"][code], SEX="M", OPTIONALITY="M")
        for measure, target in (("NPRR50", "net_replacement_aw50"), ("NPRR100", "net_replacement_aw100"), ("NPRR200", "net_replacement_aw200")):
            record["pensions"][target] = observation(latest(match(pension_rows, MEASURE=measure)), "oecd_pensions_at_a_glance")
        if record["pensions"].get("net_replacement_aw100"): comparison["pension_replacement_aw100"] = record["pensions"]["net_replacement_aw100"]

        emp = latest(match(grouped["government_employment"][code], MEASURE="EMPG", UNIT_MEASURE="PT_EMP", SECTOR="S13"))
        record["government"]["employment_share"] = observation(emp, "oecd_government_at_a_glance")
        if record["government"]["employment_share"]: comparison["government_employment"] = record["government"]["employment_share"]
        for unit, target in (("PT_B1GQ", "procurement_gdp"), ("PT_OTE_S13", "procurement_spending_share")):
            metric_obs = observation(latest(match(grouped["procurement"][code], MEASURE="GPROC", UNIT_MEASURE=unit, SECTOR="S13")), "oecd_government_at_a_glance")
            record["government"][target] = metric_obs
            if metric_obs and target == "procurement_gdp": comparison[target] = metric_obs

        wellbeing = [r for r in grouped["wellbeing"][code] if r.get("AGE") in {"_T", "_Z"} and r.get("SEX") in {"_T", "_Z"} and r.get("EDUCATION_LEV") in {"_T", "_Z"}]
        for measure, target in wellbeing_codes.items():
            metric_obs = observation(latest(match(wellbeing, MEASURE=measure)), "oecd_wellbeing")
            record["wellbeing"][target] = metric_obs
            if metric_obs and target in METRICS: comparison[target] = metric_obs

        countries[code] = record

    metric_contract = {
        code: {
            "topic": values[0], "unit_cs": values[1], "unit_en": values[2],
            "label_cs": code.replace("_", " ").capitalize(), "label_en": code.replace("_", " ").capitalize(),
            "boundary_cs": values[3], "boundary_en": values[4],
        }
        for code, values in METRICS.items()
    }
    labels = {
        "tax_to_gdp": ("Daňové příjmy", "Tax-to-GDP ratio"), "labour_tax_wedge_single": ("Daňový klín · jednotlivec", "Tax wedge · single"),
        "labour_tax_wedge_family": ("Daňový klín · rodina", "Tax wedge · family"), "corporate_statutory_rate": ("Zákonná sazba DPPO", "Statutory corporate rate"),
        "corporate_eatr": ("Efektivní průměrná sazba DPPO", "Corporate EATR"), "corporate_emtr": ("Efektivní mezní sazba DPPO", "Corporate EMTR"),
        "local_tax_autonomy": ("Místní daňová autonomie", "Local tax autonomy"), "net_carbon_rate": ("Čistá efektivní cena uhlíku", "Net effective carbon rate"),
        "disposable_gini": ("Nerovnost po přerozdělení", "Inequality after redistribution"), "market_gini": ("Tržní nerovnost", "Market-income inequality"),
        "poverty_rate": ("Relativní chudoba", "Relative poverty"), "social_spending": ("Veřejné sociální výdaje", "Public social spending"),
        "pension_replacement_aw100": ("Čistá náhradová míra důchodu", "Net pension replacement rate"), "government_employment": ("Zaměstnanost ve vládním sektoru", "Government employment"),
        "procurement_gdp": ("Veřejné zakázky", "Public procurement"), "housing_affordability": ("Náklady na bydlení", "Housing costs"),
        "life_satisfaction": ("Životní spokojenost", "Life satisfaction"), "pisa_math": ("PISA · matematika", "PISA · mathematics"), "road_deaths": ("Úmrtí na silnicích", "Road deaths"),
    }
    for code, (cs, en) in labels.items():
        metric_contract[code]["label_cs"] = cs
        metric_contract[code]["label_en"] = en

    source_titles = {
        "oecd_revenue_statistics": "Global Revenue Statistics", "oecd_taxing_wages": "Taxing Wages 2026", "oecd_corporate_tax_statistics": "Corporate Tax Statistics",
        "oecd_corporate_effective_rates": "Corporate effective tax rates", "oecd_tax_autonomy": "Fiscal Decentralisation Database · tax autonomy",
        "oecd_net_effective_carbon_rates": "Net Effective Carbon Rates", "oecd_income_distribution": "Income Distribution Database", "oecd_socx": "Social Expenditure Database (SOCX)",
        "oecd_pensions_at_a_glance": "Pensions at a Glance", "oecd_government_at_a_glance": "Government at a Glance 2025", "oecd_wellbeing": "OECD Well-being Data Monitor",
    }
    result = {
        "schema_version": "1.0.0", "dataset_id": "OECD_KEY_METRICS_V1", "generated_at": datetime.now(timezone.utc).isoformat(),
        "country_universe": COUNTRIES, "missing_values": "not_imputed", "metrics": metric_contract,
        "sources": {key: {"title": title, "publisher": "OECD", "url": AUTONOMY_URL if key == "oecd_tax_autonomy" else "https://data-explorer.oecd.org/"} for key, title in source_titles.items()},
        "coverage": {key: {"countries_with_rows": sum(bool(grouped[key].get(code)) for code in COUNTRIES), "rows": len(data)} for key, data in fetched.items()},
        "countries": countries,
    }
    OUTPUT.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT.relative_to(ROOT)} with {len(countries)} countries and {len(metric_contract)} comparison metrics")


if __name__ == "__main__":
    main()
