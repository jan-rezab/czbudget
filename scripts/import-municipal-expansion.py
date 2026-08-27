#!/usr/bin/env python3
"""Build compact public profiles from official national municipal sources.

The importer keeps a small browser directory plus one JSON detail file per
municipality. Source caches are intentionally outside the deployable website.
"""

from __future__ import annotations

import argparse
import csv
import fcntl
import gzip
import html
import io
import json
import re
import subprocess
import time
import urllib.parse
import urllib.request
import zipfile
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import openpyxl


WEB = Path(__file__).resolve().parents[1]
WORKSPACE = WEB.parent
CACHE = WORKSPACE / "data/source_cache/municipal-expansion"
OUTPUT = WORKSPACE / "outputs/municipal-expansion"
USER_AGENT = "publicspendingdata.org municipal importer/1.0"


def clean(value: object) -> str:
    return str(value or "").strip()


def number(value: object) -> float | None:
    text = clean(value).replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def slugify(value: str) -> str:
    import unicodedata
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", "-", value).strip("-") or "municipality"


def mdb_rows(database: Path, table: str) -> Iterable[dict[str, str]]:
    process = subprocess.Popen(["mdb-export", str(database), table], stdout=subprocess.PIPE, text=True, encoding="utf-8", errors="replace")
    assert process.stdout is not None
    yield from csv.DictReader(process.stdout)
    if process.wait() != 0:
        raise RuntimeError(f"mdb-export failed for {database.name}:{table}")


def trim_detail(rows: list[dict], limit: int = 240) -> list[dict]:
    rows.sort(key=lambda row: (len(row.get("code", "")), row.get("code", ""), row.get("stage", "")))
    return rows[:limit]


def import_spain() -> dict:
    budget_db = CACHE / "ESP/Presupuestos2026.accdb"
    actual_db = CACHE / "ESP/Liquidaciones2025.accdb"
    if not budget_db.exists() or not actual_db.exists():
        raise FileNotFoundError("Download and unzip the two official CONPREL Access files first")
    inventory: dict[str, dict] = {}
    internal_to_code: dict[str, str] = {}
    for database in (actual_db, budget_db):
        for row in mdb_rows(database, "tb_inventario"):
            code = clean(row.get("codbdgel"))
            entity_code = clean(row.get("codente"))
            # AA is the municipality tier in the national local-entity code.
            if len(code) != 10 or code[5:7] != "AA" or code != entity_code or clean(row.get("nsec")) != "1":
                continue
            name = clean(row.get("nombreppal")) or clean(row.get("nombreente"))
            inventory[code] = {"code": code, "name": name, "population": int(number(row.get("poblacion")) or 0)}
            internal_to_code[clean(row.get("idente"))] = code
    profiles = {code: {**entity, "country": "ESP", "currency": "EUR", "years": [2025, 2026], "history": {2025: {}, 2026: {}}, "detail": []} for code, entity in inventory.items()}

    for row in mdb_rows(budget_db, "tb_economica"):
        code = internal_to_code.get(clean(row.get("idente")))
        account = clean(row.get("cdcta"))
        side = "expenditure" if clean(row.get("tipreig")) == "G" else "revenue"
        amount = number(row.get("importe"))
        if not code or amount is None or not account:
            continue
        if len(account) == 1:
            profiles[code]["history"][2026][side] = profiles[code]["history"][2026].get(side, 0.0) + amount
        if len(account) <= 4 and amount:
            profiles[code]["detail"].append({"year": 2026, "stage": "enacted", "side": side, "code": account, "amount": amount})

    for row in mdb_rows(actual_db, "tb_economica"):
        code = internal_to_code.get(clean(row.get("idente")))
        account = clean(row.get("cdcta"))
        side = "expenditure" if clean(row.get("tipreig")) == "G" else "revenue"
        if not code or not account:
            continue
        measures = (("enacted", "imported"), ("revised", "importer"), ("actual", "importel"), ("cash", "importec"))
        for stage, field in measures:
            amount = number(row.get(field))
            if amount is None:
                continue
            if len(account) == 1 and stage == "actual":
                profiles[code]["history"][2025][side] = profiles[code]["history"][2025].get(side, 0.0) + amount
            if len(account) <= 4 and amount:
                profiles[code]["detail"].append({"year": 2025, "stage": stage, "side": side, "code": account, "amount": amount})

    for profile in profiles.values():
        for year in (2025, 2026):
            row = profile["history"][year]
            if row.get("revenue") is not None and row.get("expenditure") is not None:
                row["balance"] = row["revenue"] - row["expenditure"]
        profile["history"] = [{"year": year, **profile["history"][year]} for year in (2025, 2026)]
        profile["detail"] = trim_detail(profile["detail"])
        profile["url"] = f"/municipalities/spain/{slugify(profile['name'])}-{profile['code'].lower()}/"
    return {"country": "ESP", "generated_at": datetime.now(timezone.utc).isoformat(), "source": "https://serviciostelematicosext.hacienda.gob.es/sgfal/conprel", "entities": sorted(profiles.values(), key=lambda row: row["name"].casefold())}


def japan_catalog() -> list[dict[str, str]]:
    source = (WORKSPACE / "outputs/municipal-transparency-crawl/JPN/official-source.html").read_text(encoding="utf-8")
    result = []
    for block in re.findall(r'<article class="stat-dataset_list-item">(.*?)</article>', source, re.S):
        download = re.search(r"file-download\?statInfId=(\d+)&(?:amp;)?fileKind=1", block)
        title = re.search(r'class="stat-link_text[^>]*>\s*(.*?)\s*</a>', block, re.S)
        table = re.search(r'表番号.*?>(\d+)</span>', block, re.S)
        if download:
            result.append({"id": download.group(1), "table": table.group(1) if table else "cover", "title": re.sub(r"<.*?>", "", html.unescape(title.group(1) if title else "")).strip()})
    return result


def download(url: str, path: Path) -> None:
    if path.exists() and path.stat().st_size:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=300) as response, path.open("wb") as handle:
        while chunk := response.read(1024 * 1024):
            handle.write(chunk)


def import_japan() -> dict:
    # Core settlement, revenue/expenditure, debt/funds, and ageing-related
    # health/care tables. The complete official catalogue remains linked.
    wanted_tables = {"2", "4", "7", "8", "9", "10", "11", "12", "13", "14", "29", "33", "63", "64", "94", "95"}
    catalog = [row for row in japan_catalog() if row["table"] in wanted_tables]
    raw_dir = CACHE / "JPN"
    for row in catalog:
        download(f"https://www.e-stat.go.jp/stat-search/file-download?statInfId={row['id']}&fileKind=1", raw_dir / f"{row['id']}.csv")
    profiles: dict[str, dict] = {}
    for source in catalog:
        path = raw_dir / f"{source['id']}.csv"
        with path.open(encoding="cp932", errors="replace", newline="") as handle:
            reader = csv.reader(handle)
            header = next(reader, [])
            if len(header) < 10:
                continue
            for values in reader:
                if len(values) < 10:
                    continue
                code, prefecture, name = clean(values[2]), clean(values[3]), clean(values[4])
                entity_type = clean(values[5])
                # Types 6 and 7 are inter-municipal administrative unions and
                # special-purpose accounts. Types 1-5, 8 and 9 are Japan's
                # cities, Tokyo wards, towns and villages (1,741 bodies in FY2024).
                if not re.fullmatch(r"\d{6}", code) or not name or entity_type in {"6", "7"}:
                    continue
                profile = profiles.setdefault(code, {"code": code, "name": name, "region": prefecture, "country": "JPN", "currency": "JPY", "years": [2024], "history": [{"year": 2024}], "detail": []})
                row_code, row_label = clean(values[8]), clean(values[9])
                for column, raw in zip(header[10:], values[10:]):
                    amount = number(raw)
                    if amount in (None, 0):
                        continue
                    column_code = clean(column).split(":", 1)[0]
                    column_label = clean(column).split(":", 1)[-1]
                    native_code = ".".join(part for part in (row_code, column_code) if part)
                    label = " · ".join(part for part in (row_label, column_label) if part)
                    profile["detail"].append({"year": 2024, "stage": "actual", "table": source["table"], "table_title": source["title"], "code": native_code, "name": label, "amount": amount * 1000})
        # Table 2 is the settlement-summary table. Its first numeric fields are
        # retained as native detail; explicit totals are derived by label below.
    for profile in profiles.values():
        latest = profile["history"][0]
        for row in profile["detail"]:
            label = row.get("name", "")
            if row.get("table") == "2" and row.get("code", "").startswith("01."):
                if "歳入総額" in label and latest.get("revenue") is None: latest["revenue"] = row["amount"]
                elif "歳出総額" in label and latest.get("expenditure") is None: latest["expenditure"] = row["amount"]
                elif "実質収支" in label and latest.get("balance") is None: latest["balance"] = row["amount"]
            elif row.get("table") == "33" and row.get("code") == "97.009":
                latest["debt"] = row["amount"]
        profile["detail"] = trim_detail(profile["detail"], 360)
        profile["url"] = f"/municipalities/japan/{slugify(profile['name'])}-{profile['code'].lower()}/"
    return {"country": "JPN", "generated_at": datetime.now(timezone.utc).isoformat(), "source": "https://www.e-stat.go.jp/stat-search/files?toukei=00200251&tstat=000001077755", "tables": catalog, "entities": sorted(profiles.values(), key=lambda row: (row.get("region", ""), row["name"]))}


def request_json(url: str) -> dict:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT, "Accept": "application/json"})
    last_error: Exception | None = None
    for attempt in range(1):
        try:
            with urllib.request.urlopen(request, timeout=8) as response:
                return json.load(response)
        except Exception as error:
            last_error = error
            if attempt == 0:
                break
            time.sleep(2 ** attempt)
    raise RuntimeError(f"Official API failed after retries: {last_error}")


def paced_request_json(url: str) -> dict:
    """Start at most one SICONFI request per second across all cache workers."""
    lock_path = CACHE / "BRA/.request-rate.lock"
    stamp_path = CACHE / "BRA/.request-rate.timestamp"
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    with lock_path.open("a+") as lock:
        fcntl.flock(lock, fcntl.LOCK_EX)
        try:
            last_started = float(stamp_path.read_text(encoding="utf-8")) if stamp_path.exists() else 0.0
            time.sleep(max(0, 1.02 - (time.time() - last_started)))
            stamp_path.write_text(str(time.time()), encoding="utf-8")
        finally:
            fcntl.flock(lock, fcntl.LOCK_UN)
    return request_json(url)


def import_brazil(resume: bool = True, limit: int | None = None, shard_index: int = 0, shard_count: int = 1) -> dict:
    base = "https://apidatalake.tesouro.gov.br/ords/cdwhprd/siconfi/tt"
    entities = [row for row in paced_request_json(f"{base}/entes").get("items", []) if clean(row.get("esfera")) == "M"]
    if limit: entities = entities[:limit]
    if shard_count > 1: entities = entities[shard_index::shard_count]
    raw_dir = CACHE / "BRA/RREO-2025-P6"
    raw_dir.mkdir(parents=True, exist_ok=True)
    profiles = []
    errors = []

    def rreo_report(code: str, report_type: str, cache: Path) -> dict:
        if cache.exists() and resume:
            with gzip.open(cache, "rt", encoding="utf-8") as handle:
                return json.load(handle)
        params = urllib.parse.urlencode({"an_exercicio": 2025, "nr_periodo": 6, "co_tipo_demonstrativo": report_type, "no_anexo": "RREO-Anexo 01", "co_esfera": "M", "id_ente": code})
        payload = paced_request_json(f"{base}/rreo?{params}")
        cache.parent.mkdir(parents=True, exist_ok=True)
        with gzip.open(cache, "wt", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False)
        return payload

    def dca_report(code: str, annex: str, cache: Path) -> dict:
        if cache.exists() and resume:
            with gzip.open(cache, "rt", encoding="utf-8") as handle:
                return json.load(handle)
        params = urllib.parse.urlencode({"an_exercicio": 2024, "no_anexo": annex, "id_ente": code})
        payload = paced_request_json(f"{base}/dca?{params}")
        cache.parent.mkdir(parents=True, exist_ok=True)
        with gzip.open(cache, "wt", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False)
        return payload

    for index, entity in enumerate(entities, 1):
        code = str(entity["cod_ibge"])
        cache = raw_dir / f"{code}.json.gz"
        simplified_cache = raw_dir / "RREO-Simplificado" / f"{code}.json.gz"
        eligible_for_simplified = 0 < int(number(entity.get("populacao")) or 0) < 50_000
        payload = None
        reporting_basis = "RREO"
        try:
            # Small municipalities are the only bodies eligible to opt into
            # the simplified filing. Ask for that form first when no regular
            # response is already cached, avoiding a predictable empty call.
            if eligible_for_simplified and not cache.exists():
                simplified_payload = rreo_report(code, "RREO Simplificado", simplified_cache)
                if simplified_payload.get("items"):
                    payload = simplified_payload
                    reporting_basis = "RREO Simplificado"
            if payload is None:
                payload = rreo_report(code, "RREO", cache)
            if not payload.get("items") and eligible_for_simplified:
                payload = rreo_report(code, "RREO Simplificado", simplified_cache)
                reporting_basis = "RREO Simplificado" if payload.get("items") else "DCA fallback"
        except Exception as error:
            errors.append({"code": code, "name": clean(entity.get("ente")), "error": f"RREO: {error}"})
            print(f"Brazil RREO retry needed {code}: {error}", flush=True)
            continue
        # Municipalities below 50,000 inhabitants may opt for the official
        # simplified RREO. Only a miss in both 2025 forms reaches the annual
        # account fallback below.
        summary = {"year": 2025}
        detail = []
        items = payload.get("items", [])
        side_by_account: dict[tuple[str, str], str] = {}
        for row in items:
            account = clean(row.get("cod_conta"))
            label = clean(row.get("conta"))
            upper_label = label.upper()
            upper_column = clean(row.get("coluna")).upper()
            key = (account, label)
            if "PREVISÃO" in upper_column or "REALIZADAS" in upper_column or "RECEITAS (EXCETO" in upper_label:
                side_by_account[key] = "revenue"
            if any(term in upper_column for term in ("DOTAÇÃO", "EMPENHADAS", "LIQUIDADAS", "PAGAS")) or "DESPESAS (EXCETO" in upper_label:
                side_by_account[key] = "expenditure"
        for row in items:
            account = clean(row.get("cod_conta"))
            label = clean(row.get("conta"))
            column = clean(row.get("coluna"))
            amount = number(row.get("valor"))
            if amount is None or "%" in column: continue
            upper_column = column.upper()
            if "INICIAL" in upper_column:
                stage = "enacted"
            elif "ATUALIZADA" in upper_column:
                stage = "revised"
            elif "NO BIMESTRE" in upper_column:
                stage = "period"
            elif "SALDO" in upper_column:
                stage = "remaining"
            elif "PAGAS" in upper_column:
                stage = "cash"
            elif "EMPENHADAS" in upper_column:
                stage = "committed"
            elif any(term in upper_column for term in ("REALIZADAS", "LIQUIDADAS")) or upper_column.startswith("ATÉ O BIMESTRE"):
                stage = "actual"
            else:
                stage = "execution"
            side = side_by_account.get((account, label)) or ("revenue" if "RECEITA" in clean(row.get("anexo")).upper() or "RECEITA" in label.upper() else "expenditure")
            if len(detail) < 360: detail.append({"year": 2025, "stage": stage, "side": side, "code": account, "name": label, "column": column, "amount": amount})
            upper = label.upper()
            is_revenue_total = "RECEITAS (EXCETO INTRA-ORÇAMENTÁRIAS)" in upper or "RECEITAS (EXCETO INTRAORÇAMENTÁRIAS)" in upper
            is_expenditure_total = "DESPESAS (EXCETO INTRA-ORÇAMENTÁRIAS)" in upper or "DESPESAS (EXCETO INTRAORÇAMENTÁRIAS)" in upper
            if is_revenue_total and upper_column.startswith("ATÉ O BIMESTRE"):
                summary.setdefault("revenue", amount)
            if is_expenditure_total and "LIQUIDADAS ATÉ O BIMESTRE" in upper_column:
                summary.setdefault("expenditure", amount)
        years = [2025]
        # Some directory municipalities have no RREO Annex 01 row. Preserve
        # that gap, but use their official 2024 DCA I-C/I-D annual accounts where
        # available instead of publishing an empty profile.
        if not payload.get("items"):
            reporting_basis = "DCA fallback"
            try:
                dca_revenue = dca_report(code, "DCA-Anexo I-C", raw_dir / "DCA-2024-I-C" / f"{code}.json.gz")
                dca_expenditure = dca_report(code, "DCA-Anexo I-D", raw_dir / "DCA-2024-I-D" / f"{code}.json.gz")
            except Exception as error:
                errors.append({"code": code, "name": clean(entity.get("ente")), "error": f"DCA fallback: {error}"})
                print(f"Brazil DCA retry needed {code}: {error}", flush=True)
                continue
            dca_items = [*dca_revenue.get("items", []), *dca_expenditure.get("items", [])]
            if dca_items:
                summary = {"year": 2024}
                years = [2024]
                for side, rows in (("revenue", dca_revenue.get("items", [])), ("expenditure", dca_expenditure.get("items", []))):
                    for row in rows:
                        account = clean(row.get("cod_conta")); label = clean(row.get("conta")); column = clean(row.get("coluna")); amount = number(row.get("valor"))
                        if amount is None: continue
                        upper_label, upper_column = label.upper(), column.upper()
                        stage = "cash" if "PAGAS" in upper_column else "execution" if "EMPENHADAS" in upper_column or "RESTOS A PAGAR" in upper_column else "actual"
                        if len(detail) < 360: detail.append({"year": 2024, "stage": stage, "side": side, "code": account, "name": label, "column": column, "amount": amount})
                        if side == "revenue" and "RECEITAS (EXCETO INTRA-ORÇAMENTÁRIAS)" in upper_label and "BRUTAS REALIZADAS" in upper_column:
                            summary.setdefault("revenue", amount)
                        if side == "expenditure" and account == "TotalDespesas" and "LIQUIDADAS" in upper_column:
                            summary.setdefault("expenditure", amount)
        if summary.get("revenue") is not None and summary.get("expenditure") is not None: summary["balance"] = summary["revenue"] - summary["expenditure"]
        profiles.append({"code": code, "name": clean(entity.get("ente")), "region": clean(entity.get("uf")), "population": entity.get("populacao"), "country": "BRA", "currency": "BRL", "reporting_basis": reporting_basis, "years": years, "history": [summary], "detail": detail, "url": f"/municipalities/brazil/{slugify(clean(entity.get('ente')))}-{code}/"})
        if index % 100 == 0: print(f"Brazil {index}/{len(entities)}", flush=True)
    return {"country": "BRA", "generated_at": datetime.now(timezone.utc).isoformat(), "source": "https://apidatalake.tesouro.gov.br/docs/siconfi", "entities": profiles, "errors": errors}


def cached_json(url: str, path: Path) -> list[dict]:
    """Read a public JSON response through a persistent, auditable cache."""
    if not path.exists():
        download(url, path)
    return json.loads(path.read_text(encoding="utf-8"))


def socrata_url(dataset: str, select: str, where: str, *, group: str | None = None) -> str:
    params = {"$select": select, "$where": where, "$order": "codigo_entidad", "$limit": 50000}
    if group:
        params["$group"] = group
    return f"https://www.datos.gov.co/resource/{dataset}.json?{urllib.parse.urlencode(params)}"


def import_colombia() -> dict:
    """Import the complete 2025 municipal CUIPO plan and execution layer."""
    period = "20251201"
    cache = CACHE / "COL" / period
    scope = f"periodo='{period}' AND ambito_nombre='Municipios'"
    datasets = {
        "revenue_actual": (
            "9axr-9gnb",
            "codigo_entidad,nombre_entidad,sum(total_recaudo) as amount",
            f"{scope} AND cuenta='1'",
            "codigo_entidad,nombre_entidad",
        ),
        "expenditure_actual": (
            "4f7r-epif",
            "codigo_entidad,nombre_entidad,sum(pagos) as amount",
            f"{scope} AND cuenta='2'",
            "codigo_entidad,nombre_entidad",
        ),
        "expenditure_plan": (
            "d9mu-h6ar",
            "codigo_entidad,nombre_entidad,sum(apropiacion_inicial) as initial,sum(apropiacion_definitiva) as definitive",
            f"{scope} AND cuenta='2'",
            "codigo_entidad,nombre_entidad",
        ),
    }
    rows: dict[str, list[dict]] = {}
    for label, (dataset, select, where, group) in datasets.items():
        rows[label] = cached_json(socrata_url(dataset, select, where, group=group), cache / f"{label}.json")

    # The official revenue-programming dataset currently exposes its last four
    # columns under shifted metadata names. The top-level account row remains
    # unambiguous: ambito_codigo=1, nombre_cuenta=Municipios.
    revenue_plan_url = socrata_url(
        "22ah-ddsj",
        "codigo_entidad,nombre_entidad,cod_detalle_sectorial as initial,nom_detalle_sectorial as definitive",
        f"periodo='{period}' AND nombre_cuenta='Municipios' AND ambito_codigo='1'",
    )
    rows["revenue_plan"] = cached_json(revenue_plan_url, cache / "revenue_plan.json")

    profiles: dict[str, dict] = {}
    for group_rows in rows.values():
        for row in group_rows:
            code = clean(row.get("codigo_entidad"))
            if not code:
                continue
            profiles.setdefault(code, {
                "code": code,
                "name": clean(row.get("nombre_entidad")).title(),
                "region": None,
                "country": "COL",
                "currency": "COP",
                "years": [2025],
                "history": [{"year": 2025}],
                "detail": [],
            })

    by_code = {label: {clean(row.get("codigo_entidad")): row for row in group_rows} for label, group_rows in rows.items()}
    for code, profile in profiles.items():
        summary = profile["history"][0]
        detail = profile["detail"]
        for side in ("revenue", "expenditure"):
            plan = by_code[f"{side}_plan"].get(code, {})
            actual = by_code[f"{side}_actual"].get(code, {})
            enacted = number(plan.get("initial"))
            revised = number(plan.get("definitive"))
            outcome = number(actual.get("amount"))
            for stage, amount in (("enacted", enacted), ("revised", revised), ("actual", outcome)):
                if amount is not None:
                    detail.append({"year": 2025, "stage": stage, "side": side, "code": "1" if side == "revenue" else "2", "name": "INGRESOS" if side == "revenue" else "GASTOS", "amount": amount})
            summary[side] = outcome
        if summary.get("revenue") is not None and summary.get("expenditure") is not None:
            summary["balance"] = summary["revenue"] - summary["expenditure"]
        profile["url"] = f"/municipalities/colombia/{slugify(profile['name'])}-{code.lower()}/"

    complete = [profile for profile in profiles.values() if profile["history"][0].get("revenue") is not None and profile["history"][0].get("expenditure") is not None]
    errors = [{"code": code, "error": "missing 2025 revenue or expenditure execution"} for code, profile in profiles.items() if profile not in complete]
    return {
        "country": "COL",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "https://www.datos.gov.co/browse?q=OVCF%20CUIPO",
        "source_datasets": {"revenue_plan": "22ah-ddsj", "revenue_actual": "9axr-9gnb", "expenditure_plan": "d9mu-h6ar", "expenditure_actual": "4f7r-epif"},
        "period": period,
        "entities": sorted(complete, key=lambda row: row["name"].casefold()),
        "errors": errors,
    }


def workbook_value_map(sheet) -> tuple[dict[int, tuple[int, str]], dict[str, int]]:
    years: dict[int, tuple[int, str]] = {}
    for column in range(1, sheet.max_column + 1):
        label = clean(sheet.cell(4, column).value)
        match = re.search(r"(20\d{2})", label)
        if not match:
            continue
        stage = "enacted" if "გეგმა" in label else "actual"
        # A workbook can contain both the approved annual plan and a partial
        # outturn for the same year. The annual column comes first.
        years.setdefault(int(match.group(1)), (column, stage))
    labels = {clean(sheet.cell(row, 4).value): row for row in range(1, sheet.max_row + 1) if clean(sheet.cell(row, 4).value)}
    return years, labels


def import_georgia() -> dict:
    """Import Georgia's national municipality workbooks and 2025 functions."""
    base = "https://www.mof.ge"
    list_url = f"{base}/ka/FileList/List?page=1&id=79"
    catalog = CACHE / "GEO" / "municipal-budget-catalog.html"
    if not catalog.exists():
        download(list_url, catalog)
    source = catalog.read_text(encoding="utf-8")
    links = list(dict.fromkeys(html.unescape(link) for link in re.findall(r'href="([^"]+\.xlsx/[^"]+)"', source)))
    if not links:
        raise RuntimeError("Georgia Ministry of Finance catalog returned no municipal workbooks")

    functional_path = WORKSPACE / "outputs/municipal-transparency-crawl/GEO/municipal-functional-2025.xlsx"
    functional = openpyxl.load_workbook(functional_path, data_only=True).active
    functional_by_name: dict[str, tuple[str, list[dict]]] = {}
    function_columns = [(column, clean(functional.cell(3, column).value), clean(functional.cell(4, column).value)) for column in range(3, functional.max_column + 1)]
    for row in range(5, functional.max_row + 1):
        name = clean(functional.cell(row, 2).value)
        if "მუნიციპალიტეტი" not in name:
            continue
        source_index = int(functional.cell(row, 1).value)
        detail = []
        for column, code, label in function_columns:
            amount = number(functional.cell(row, column).value)
            if amount not in (None, 0):
                detail.append({"year": 2025, "stage": "actual", "side": "expenditure", "code": code, "name": label, "amount": amount * 1000})
        functional_by_name[name] = (f"MOF-{source_index:03d}", detail)

    profiles = []
    errors = []
    raw_dir = CACHE / "GEO" / "municipal-workbooks"
    for index, link in enumerate(links, 1):
        url = urllib.parse.urljoin(base, link)
        path = raw_dir / f"{index:03d}.xlsx"
        try:
            if not path.exists():
                download(url, path)
            sheet = openpyxl.load_workbook(path, data_only=True).active
            name = clean(sheet.cell(2, 4).value)
            if "მუნიციპალიტეტი" not in name:
                continue
            year_columns, labels = workbook_value_map(sheet)
            # These are total receipts/payments. The narrower current-account
            # rows exclude asset transactions and do not reconcile to totals.
            revenue_row = labels.get("შემოსულობები")
            expenditure_row = labels.get("გადასახდელები")
            if not revenue_row or not expenditure_row:
                raise ValueError("headline receipts/payments rows not found")
            if name not in functional_by_name:
                raise ValueError("municipality missing from 2025 functional workbook")
            code, function_detail = functional_by_name[name]
            history = []
            detail = list(function_detail)
            for year in (2025, 2026):
                if year not in year_columns:
                    continue
                column, stage = year_columns[year]
                revenue = number(sheet.cell(revenue_row, column).value)
                expenditure = number(sheet.cell(expenditure_row, column).value)
                if revenue is not None: revenue *= 1000
                if expenditure is not None: expenditure *= 1000
                if year == 2025:
                    functional_total = next((row["amount"] for row in function_detail if row["code"] == "7"), None)
                    if functional_total is None or expenditure is None or abs(functional_total - expenditure) > 1:
                        raise ValueError(f"2025 payments do not reconcile to functional total ({expenditure} vs {functional_total})")
                history.append({"year": year, "revenue": revenue, "expenditure": expenditure, "balance": revenue - expenditure if revenue is not None and expenditure is not None else None})
                for side, amount, native_code, label in (("revenue", revenue, "R", "შემოსულობები"), ("expenditure", expenditure, "E", "გადასახდელები")):
                    if amount is not None:
                        detail.append({"year": year, "stage": stage, "side": side, "code": native_code, "name": label, "amount": amount})
            if not history:
                raise ValueError("2025/2026 values not found")
            profiles.append({
                "code": code,
                "name": name,
                "region": None,
                "country": "GEO",
                "currency": "GEL",
                "years": [row["year"] for row in history],
                "history": history,
                "detail": detail,
                "source_url": url,
                "url": f"/municipalities/georgia/{slugify(name)}-{code.lower()}/",
            })
        except Exception as error:
            errors.append({"source": url, "error": f"{type(error).__name__}: {error}"})
    return {
        "country": "GEO",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "https://www.mof.ge/ka/page/budget-of-autonomous-republics-and-municipalities",
        "entities": sorted(profiles, key=lambda row: row["name"]),
        "errors": errors,
    }


ITALY_TOP_LEVEL = {
    "revenue": {
        "0": "Incassi da regolarizzare",
        "1": "Entrate correnti tributarie, contributive e perequative",
        "2": "Trasferimenti correnti", "3": "Entrate extratributarie",
        "4": "Entrate in conto capitale", "5": "Riduzione di attività finanziarie",
        "6": "Accensione di prestiti", "7": "Anticipazioni di tesoreria",
        "9": "Entrate per conto terzi e partite di giro",
    },
    "expenditure": {
        "0": "Pagamenti da regolarizzare", "1": "Spese correnti",
        "2": "Spese in conto capitale", "3": "Incremento di attività finanziarie",
        "4": "Rimborso di prestiti", "5": "Chiusura anticipazioni di tesoreria",
        "7": "Spese per conto terzi e partite di giro",
    },
}


def zip_csv_rows(path: Path, prefix: str):
    with zipfile.ZipFile(path) as archive:
        member = next(name for name in archive.namelist() if name.startswith(prefix))
        with archive.open(member) as raw:
            yield from csv.reader(io.TextIOWrapper(raw, encoding="latin-1", newline=""))


def italian_municipality_name(value: str) -> str:
    value = re.sub(r"^COMUNE DI\s+", "", clean(value), flags=re.IGNORECASE)
    return value.title().replace(" D'", " d'").replace(" Dell'", " dell'")


def import_italy() -> dict:
    """Import 2025 municipal cash receipts/payments from national SIOPE ZIPs."""
    root = CACHE / "ITA"
    sources = {
        "registry": ("SIOPE_ANAGRAFICHE.zip", "https://www.siope.it/Siope/documenti/siope2/open/last/SIOPE_ANAGRAFICHE.zip"),
        "revenue": ("SIOPE_ENTRATE.2025.zip", "https://www.siope.it/Siope/documenti/siope2/open/last/SIOPE_ENTRATE.2025.zip"),
        "expenditure": ("SIOPE_USCITE.2025.zip", "https://www.siope.it/Siope/documenti/siope2/open/last/SIOPE_USCITE.2025.zip"),
    }
    for filename, url in sources.values():
        if not (root / filename).exists():
            download(url, root / filename)

    province_to_region = {}
    for row in zip_csv_rows(root / sources["registry"][0], "ANAG_REG_PROV"):
        if len(row) >= 5:
            province_to_region[row[4]] = row[2]
    entities = {}
    for row in zip_csv_rows(root / sources["registry"][0], "ANAG_ENTI_SIOPE"):
        if len(row) < 9 or row[8] != "COMUNE" or row[1] > "2025-12-31" or row[2] < "2025-01-01":
            continue
        entities[row[0]] = {"code": row[0], "name": italian_municipality_name(row[4]), "region": province_to_region.get(row[6]), "population": number(row[7])}

    totals = {side: defaultdict(float) for side in ("revenue", "expenditure")}
    detail = {side: defaultdict(lambda: defaultdict(float)) for side in ("revenue", "expenditure")}
    for side, prefix in (("revenue", "ENTRATE_2025"), ("expenditure", "USCITE_2025")):
        for row in zip_csv_rows(root / sources[side][0], prefix):
            if len(row) < 5 or row[0] not in entities:
                continue
            amount = int(row[4]) / 100
            totals[side][row[0]] += amount
            detail[side][row[0]][row[3].split(".", 1)[0]] += amount

    reporters = set(totals["revenue"]) & set(totals["expenditure"])
    profiles = []
    for code in sorted(reporters, key=lambda item: entities[item]["name"].casefold()):
        entity = entities[code]
        revenue = round(totals["revenue"][code], 2)
        expenditure = round(totals["expenditure"][code], 2)
        native_detail = []
        for side in ("revenue", "expenditure"):
            for account, amount in sorted(detail[side][code].items()):
                native_detail.append({"year": 2025, "stage": "cash", "side": side, "code": account, "name": ITALY_TOP_LEVEL[side].get(account, f"Titolo {account}"), "amount": round(amount, 2)})
            reconciled = sum(row["amount"] for row in native_detail if row["side"] == side)
            if abs(reconciled - totals[side][code]) > 0.02:
                raise ValueError(f"SIOPE {side} detail does not reconcile for {code}")
        profiles.append({
            **entity, "country": "ITA", "currency": "EUR", "years": [2025],
            "history": [{"year": 2025, "revenue": revenue, "expenditure": expenditure, "balance": round(revenue - expenditure, 2)}],
            "detail": native_detail, "url": f"/municipalities/italy/{slugify(entity['name'])}-{code}/",
        })

    nonreporting = sorted(set(entities) - reporters)
    return {
        "country": "ITA", "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "https://www.siope.it/Siope/", "period": 2025, "entities": profiles,
        "excluded_nonreporting_registry_aliases": [{"code": code, "name": entities[code]["name"]} for code in nonreporting],
        "errors": [],
    }


def bolivia_local_name(value: str) -> str:
    value = re.sub(r"^Gobierno Autónomo Municipal De\s+", "", clean(value), flags=re.IGNORECASE)
    return value


def import_bolivia() -> dict:
    """Import 2025 budgets and execution for Bolivia's active local governments."""
    root = CACHE / "BOL"
    source_root = "https://abierto.economiayfinanzas.gob.bo/presupuesto/descargas"
    paths = {
        "registry": root / "entidad-clasificador.json",
        "expenditure": root / "gestion_2025_gasto.zip",
        "revenue": root / "gestion_2025_ingreso.zip",
    }
    if not paths["registry"].exists(): download("https://abierto.economiayfinanzas.gob.bo/api/entidad-clasificador", paths["registry"])
    for side in ("expenditure", "revenue"):
        native = "gasto" if side == "expenditure" else "ingreso"
        if not paths[side].exists(): download(f"{source_root}/{native}/gestiones/gestion_2025.zip", paths[side])

    allowed = {"Gobiernos Autonomos Municipales", "Gobiernos Autónomos Indígenas Originarios Campesinos"}
    registry = {
        str(row["entidad"]): row for row in json.loads(paths["registry"].read_text(encoding="utf-8"))
        if row.get("desc_subarea") in allowed
    }
    stage_fields = {
        "revenue": (("enacted", "inicial"), ("revised", "vigente"), ("actual", "devengado"), ("cash", "percibido")),
        "expenditure": (("enacted", "inicial"), ("revised", "vigente"), ("actual", "devengado"), ("cash", "pagado")),
    }
    totals = {side: defaultdict(lambda: defaultdict(float)) for side in stage_fields}
    details = {side: defaultdict(lambda: defaultdict(float)) for side in stage_fields}
    regions = {}
    for side, prefix in (("revenue", "presupuesto_ingreso"), ("expenditure", "presupuesto_gasto")):
        with zipfile.ZipFile(paths[side]) as archive:
            member = next(name for name in archive.namelist() if name.startswith(prefix) and name.endswith(".csv"))
            with archive.open(member) as raw:
                for row in csv.DictReader(io.TextIOWrapper(raw, encoding="utf-8", newline="")):
                    code = row["entidad"]
                    if code not in registry:
                        continue
                    region = clean(row.get("desc_departamento"))
                    if region:
                        regions.setdefault(code, region)
                    account = clean(row["rubro_tipo"] if side == "revenue" else row["objeto_grupo"])
                    label = clean(row["rubro_desc_tipo"] if side == "revenue" else row["objeto_desc_grupo"])
                    for stage, field in stage_fields[side]:
                        amount = number(row.get(field)) or 0
                        totals[side][code][stage] += amount
                        details[side][code][(stage, account, label)] += amount

    reporters = set(totals["revenue"]) & set(totals["expenditure"])
    profiles = []
    for code in sorted(reporters, key=lambda item: registry[item]["desc_entidad"].casefold()):
        source = registry[code]
        revenue = round(totals["revenue"][code]["actual"], 2)
        expenditure = round(totals["expenditure"][code]["actual"], 2)
        native_detail = []
        for side in ("revenue", "expenditure"):
            for (stage, account, label), amount in sorted(details[side][code].items()):
                native_detail.append({"year": 2025, "stage": stage, "side": side, "code": account, "name": label, "amount": round(amount, 2)})
            for stage, _ in stage_fields[side]:
                reconciled = sum(row["amount"] for row in native_detail if row["side"] == side and row["stage"] == stage)
                if abs(reconciled - totals[side][code][stage]) > 0.1:
                    raise ValueError(f"Bolivia {side}/{stage} detail does not reconcile for {code}")
        name = bolivia_local_name(source["desc_entidad"])
        profiles.append({
            "code": code, "name": name, "region": regions.get(code), "country": "BOL", "currency": "BOB", "years": [2025],
            "government_type": source["desc_subarea"],
            "history": [{"year": 2025, "revenue": revenue, "expenditure": expenditure, "balance": round(revenue - expenditure, 2)}],
            "detail": native_detail, "url": f"/municipalities/bolivia/{slugify(name)}-{code}/",
        })

    dormant = sorted(set(registry) - reporters)
    return {
        "country": "BOL", "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "https://abierto.economiayfinanzas.gob.bo/descargas", "period": 2025,
        "entities": profiles,
        "excluded_superseded_municipal_governments": [{"code": code, "name": registry[code]["desc_entidad"]} for code in dormant],
        "errors": [],
    }


SALVADOR_DEPARTMENTS = {
    "81":"Ahuachapán", "82":"Santa Ana", "83":"Sonsonate", "84":"Chalatenango",
    "85":"La Libertad", "86":"San Salvador", "87":"Cuscatlán", "88":"La Paz",
    "89":"Cabañas", "90":"San Vicente", "91":"Usulután", "92":"San Miguel",
    "93":"Morazán", "94":"La Unión",
}


def import_el_salvador() -> dict:
    """Import the last full-year national SAFIM municipal execution file."""
    root = CACHE / "SLV"
    archive_path = root / "700-DGCG-DA-2023-PRES.zip"
    registry_path = root / "municipalities.json"
    if not archive_path.exists():
        download("https://www.transparenciafiscal.gob.sv/downloads/zip/700-DGCG-DA-2023-PRES.zip", archive_path)
    if not registry_path.exists():
        municipalities = []
        for department in range(81, 95):
            request = urllib.request.Request(
                f"https://www.transparenciafiscal.gob.sv/ptf-services/municipality/municipalities?depto={department}",
                headers={"User-Agent": "Mozilla/5.0", "Referer": "https://www.transparenciafiscal.gob.sv/ptf/es/PTF2-Municipios.html"},
            )
            with urllib.request.urlopen(request, timeout=60) as response:
                municipalities.extend(json.load(response))
        registry_path.parent.mkdir(parents=True, exist_ok=True)
        registry_path.write_text(json.dumps(municipalities, ensure_ascii=False), encoding="utf-8")
    registry = {str(row["CODIGO"]): row["NOMBRE"] for row in json.loads(registry_path.read_text(encoding="utf-8"))}

    totals = {side: defaultdict(lambda: defaultdict(float)) for side in ("revenue", "expenditure")}
    details = {side: defaultdict(lambda: defaultdict(float)) for side in ("revenue", "expenditure")}
    with zipfile.ZipFile(archive_path) as archive:
        member = next(name for name in archive.namelist() if name.endswith(".csv"))
        with archive.open(member) as raw:
            for row in csv.DictReader(io.TextIOWrapper(raw, encoding="utf-8-sig", newline=""), delimiter=";"):
                if row["MES"] != "12" or len(row["OE"]) != 2 or row["TIPOPRES"] not in {"1", "2"}:
                    continue
                code = row["ALCALDIA"]
                side = "revenue" if row["TIPOPRES"] == "1" else "expenditure"
                for stage, field in (("enacted", "VOTADO"), ("revised", "MODIFICADO"), ("actual", "DEVENGADO")):
                    amount = number(row[field]) or 0
                    totals[side][code][stage] += amount
                    details[side][code][(stage, row["OE"], row["OE_DESC"])] += amount

    reporters = set(totals["revenue"]) & set(totals["expenditure"])
    profiles = []
    for code in sorted(reporters, key=lambda item: registry.get(item, item).casefold()):
        name = re.sub(r"^ALCALDIA MUNICIPAL DE\s+", "", registry.get(code, code), flags=re.IGNORECASE).rstrip(".").title()
        revenue = round(totals["revenue"][code]["actual"], 2)
        expenditure = round(totals["expenditure"][code]["actual"], 2)
        native_detail = []
        for side in ("revenue", "expenditure"):
            for (stage, account, label), amount in sorted(details[side][code].items()):
                native_detail.append({"year": 2023, "stage": stage, "side": side, "code": account, "name": label, "amount": round(amount, 2)})
            for stage in ("enacted", "revised", "actual"):
                reconciled = sum(row["amount"] for row in native_detail if row["side"] == side and row["stage"] == stage)
                if abs(reconciled - totals[side][code][stage]) > 0.05:
                    raise ValueError(f"El Salvador {side}/{stage} detail does not reconcile for {code}")
        profiles.append({
            "code": code, "name": name, "region": SALVADOR_DEPARTMENTS.get(code[:2]), "country": "SLV", "currency": "USD", "years": [2023],
            "history": [{"year": 2023, "revenue": revenue, "expenditure": expenditure, "balance": round(revenue - expenditure, 2)}],
            "detail": native_detail, "url": f"/municipalities/el-salvador/{slugify(name)}-{code}/",
        })
    missing = sorted(set(registry) - reporters)
    return {
        "country": "SLV", "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "https://www.transparenciafiscal.gob.sv/ptf/es/PTF2-Datos_Abiertos.html", "period": 2023,
        "status": "partial", "entities": profiles,
        "nonreporting_municipalities": [{"code": code, "name": registry[code]} for code in missing], "errors": [],
    }


MEXICO_STATES = {
    "01":"Aguascalientes", "02":"Baja California", "03":"Baja California Sur", "04":"Campeche",
    "05":"Coahuila", "06":"Colima", "07":"Chiapas", "08":"Chihuahua", "09":"Ciudad de México",
    "10":"Durango", "11":"Guanajuato", "12":"Guerrero", "13":"Hidalgo", "14":"Jalisco",
    "15":"México", "16":"Michoacán", "17":"Morelos", "18":"Nayarit", "19":"Nuevo León",
    "20":"Oaxaca", "21":"Puebla", "22":"Querétaro", "23":"Quintana Roo", "24":"San Luis Potosí",
    "25":"Sinaloa", "26":"Sonora", "27":"Tabasco", "28":"Tamaulipas", "29":"Tlaxcala",
    "30":"Veracruz", "31":"Yucatán", "32":"Zacatecas",
}


def import_mexico() -> dict:
    """Import Mexico's definitive 2024 EFIPEM municipal finance returns."""
    root = CACHE / "MEX"
    archive_path = root / "conjunto_de_datos_efipem_municipal_csv.zip"
    if not archive_path.exists():
        download(
            "https://www.inegi.org.mx/contenidos/programas/finanzas/datosabiertos/conjunto_de_datos_efipem_municipal_csv.zip",
            archive_path,
        )
    totals: dict[str, dict[str, float]] = defaultdict(dict)
    detail: dict[str, list[dict]] = defaultdict(list)
    names: dict[str, str] = {}
    with zipfile.ZipFile(archive_path) as archive:
        with archive.open("catalogos/tc_municipio.csv") as raw:
            for row in csv.DictReader(io.TextIOWrapper(raw, encoding="utf-8-sig", newline="")):
                names[row["CVEGEO"]] = clean(row["NOM_MUN"])
        member = "conjunto_de_datos/efipem_municipal_anual_tr_cifra_2024.csv"
        with archive.open(member) as raw:
            for row in csv.DictReader(io.TextIOWrapper(raw, encoding="utf-8-sig", newline="")):
                code = row["CVEGEO"]
                side = "revenue" if row["TEMA"] == "Ingresos" else "expenditure"
                amount = number(row["VALOR"])
                if amount is None:
                    continue
                if row["CATEGORIA"] == "Tema":
                    totals[code][side] = amount
                elif row["CATEGORIA"] == "Capítulo":
                    detail[code].append({
                        "year": 2024, "stage": "actual", "side": side,
                        "code": f"{side[0].upper()}-{len([item for item in detail[code] if item['side'] == side]) + 1:02d}",
                        "name": clean(row["DESCRIPCION_CATEGORIA"]), "amount": amount,
                    })
    profiles = []
    errors = []
    for code, values in totals.items():
        if values.get("revenue") is None or values.get("expenditure") is None:
            errors.append({"code": code, "error": "missing definitive revenue or expenditure total"})
            continue
        for side in ("revenue", "expenditure"):
            reconciled = sum(row["amount"] for row in detail[code] if row["side"] == side)
            if abs(reconciled - values[side]) > 1:
                raise ValueError(f"Mexico {side} chapters do not reconcile for {code}")
        name = names.get(code, code)
        revenue, expenditure = values["revenue"], values["expenditure"]
        profiles.append({
            "code": code, "name": name, "region": MEXICO_STATES.get(code[:2]),
            "country": "MEX", "currency": "MXN", "years": [2024],
            "history": [{"year": 2024, "revenue": revenue, "expenditure": expenditure, "balance": revenue - expenditure}],
            "detail": detail[code], "url": f"/municipalities/mexico/{slugify(name)}-{code}/",
        })
    return {
        "country": "MEX", "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "https://www.inegi.org.mx/programas/finanzas/", "period": 2024,
        "status": "partial", "entities": sorted(profiles, key=lambda row: (row.get("region", ""), row["name"].casefold())),
        "reporting_note": "2,380 municipal governments published definitive 2024 returns; the national municipal universe is larger.",
        "errors": errors,
    }


def spanish_number(value: object) -> float | None:
    text = clean(value).replace(".", "").replace(",", ".")
    if not text or text == "-":
        return None
    return float(text)


def import_costa_rica() -> dict:
    """Import the complete 2025 municipality result exported from CGR SIPP."""
    source_path = CACHE / "CRI/Instituciones-2025.csv"
    if not source_path.exists():
        raise FileNotFoundError("Export the 2025 SIPP institutions report filtered to municipalities first")
    profiles = []
    with source_path.open(encoding="latin-1", newline="") as handle:
        for row in csv.DictReader(handle, delimiter=";"):
            native_name = clean(row.get("Nombre"))
            if not native_name.startswith("MUNICIPALIDAD DE "):
                continue
            name = re.sub(r"^MUNICIPALIDAD DE\s+", "", native_name).title()
            revenue = spanish_number(row.get("Ingreso recibido *"))
            expenditure = spanish_number(row.get("Gastado *"))
            balance = spanish_number(row.get("Superávit o déficit *"))
            if revenue is None or expenditure is None:
                continue
            # SIPP explicitly labels the report in millions of colones.
            revenue *= 1_000_000
            expenditure *= 1_000_000
            balance = balance * 1_000_000 if balance is not None else revenue - expenditure
            code = f"SIPP-{slugify(name).upper()}"
            detail = [
                {"year": 2025, "stage": "actual", "side": "revenue", "code": "TOTAL", "name": "Ingreso recibido", "amount": revenue},
                {"year": 2025, "stage": "actual", "side": "expenditure", "code": "TOTAL", "name": "Gastado", "amount": expenditure},
            ]
            profiles.append({
                "code": code, "name": name, "region": None, "country": "CRI", "currency": "CRC", "years": [2025],
                "history": [{"year": 2025, "revenue": revenue, "expenditure": expenditure, "balance": balance}],
                "detail": detail, "url": f"/municipalities/costa-rica/{slugify(name)}-{slugify(code)}/",
            })
    if len(profiles) != 84:
        raise ValueError(f"Costa Rica SIPP export yielded {len(profiles)} municipalities, expected 84")
    return {
        "country": "CRI", "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "https://cgrweb.cgr.go.cr/apex/f?p=150220:2", "period": 2025,
        "entities": sorted(profiles, key=lambda row: row["name"].casefold()), "errors": [],
    }


GUATEMALA_EXPENSE_FILES = (
    "552f21ff-af02-48ad-b813-1c48d2b4737e.xlsx", "93944d9f-5f03-45e7-b889-ec9d9cf140d9.xlsx",
    "95ff9cc8-e0bf-4871-add2-d25be168103a.xlsx", "c703380d-dfe9-4611-9525-ef755fc3fa00.xlsx",
    "a14cec3b-0504-4d0d-a347-10f121a8e684.xlsx", "e8afabcd-21bb-4eb2-8d46-0521c6de17e6.xlsx",
    "5d27ddd2-6045-4e75-98a4-a03232ea1b82.xlsx", "87c29651-fc56-44f9-972a-50d1ae77f3b8.xlsx",
)


def import_guatemala() -> dict:
    """Import 2025 SICOINGL/SICOINDES returns for all 340 municipalities."""
    root = CACHE / "GTM"
    revenue_path = root / "b33a3951-408e-4d1b-abd5-0d63c280ec38.xlsx"
    required = [revenue_path, *(root / name for name in GUATEMALA_EXPENSE_FILES)]
    if any(not path.exists() for path in required):
        raise FileNotFoundError("Download the nine official MINFIN municipal workbooks first")
    stage_fields = {
        "revenue": (("enacted", "asignado"), ("revised", "vigente"), ("actual", "devengado"), ("cash", "percibido")),
        "expenditure": (("enacted", "asignado"), ("revised", "vigente"), ("actual", "devengado"), ("cash", "pagado")),
    }
    totals = {side: defaultdict(lambda: defaultdict(float)) for side in stage_fields}
    details = {side: defaultdict(lambda: defaultdict(float)) for side in stage_fields}
    registry: dict[str, dict] = {}

    def consume(path: Path, side: str) -> None:
        sheet = openpyxl.load_workbook(path, read_only=True, data_only=True).active
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        index = {clean(value): column for column, value in enumerate(header)}
        account_field = "codigoClase" if side == "revenue" else "codigoGrupoGasto"
        label_field = "clase" if side == "revenue" else "grupoGasto"
        for values in rows:
            code = clean(values[index["codigoEntidad"]])
            if not code:
                continue
            registry.setdefault(code, {
                "name": re.sub(r"^MUNICIPALIDAD DE\s+", "", clean(values[index["entidad"]]), flags=re.IGNORECASE).title(),
                "region": clean(values[index["region"]]).title(),
                "department": clean(values[index["departamento"]]).title(),
            })
            account = clean(values[index[account_field]])
            label = clean(values[index[label_field]])
            for stage, field in stage_fields[side]:
                amount = number(values[index[field]]) or 0
                totals[side][code][stage] += amount
                details[side][code][(stage, account, label)] += amount

    consume(revenue_path, "revenue")
    for filename in GUATEMALA_EXPENSE_FILES:
        consume(root / filename, "expenditure")
    reporters = set(totals["revenue"]) & set(totals["expenditure"])
    profiles = []
    for code in reporters:
        entity = registry[code]
        native_detail = []
        for side in ("revenue", "expenditure"):
            for (stage, account, label), amount in sorted(details[side][code].items()):
                native_detail.append({"year": 2025, "stage": stage, "side": side, "code": account, "name": label, "amount": round(amount, 2)})
            for stage, _ in stage_fields[side]:
                reconciled = sum(row["amount"] for row in native_detail if row["side"] == side and row["stage"] == stage)
                if abs(reconciled - totals[side][code][stage]) > 0.2:
                    raise ValueError(f"Guatemala {side}/{stage} detail does not reconcile for {code}")
        revenue = round(totals["revenue"][code]["actual"], 2)
        expenditure = round(totals["expenditure"][code]["actual"], 2)
        profiles.append({
            "code": code, "name": entity["name"], "region": entity["department"], "macro_region": entity["region"],
            "country": "GTM", "currency": "GTQ", "years": [2025],
            "history": [{"year": 2025, "revenue": revenue, "expenditure": expenditure, "balance": round(revenue - expenditure, 2)}],
            "detail": native_detail, "url": f"/municipalities/guatemala/{slugify(entity['name'])}-{code}/",
        })
    if len(profiles) != 340:
        raise ValueError(f"Guatemala import yielded {len(profiles)} paired municipalities, expected 340")
    return {
        "country": "GTM", "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "https://datos.minfin.gob.gt/es/dataset/informacion-presupuestaria-municipal-2025",
        "period": 2025, "entities": sorted(profiles, key=lambda row: (row["region"], row["name"])), "errors": [],
    }


def peru_is_municipality(name: str) -> bool:
    return name.startswith("MUNICIPALIDAD PROVINCIAL") or name.startswith("MUNICIPALIDAD DISTRITAL") or name == "MUNICIPALIDAD METROPOLITANA DE LIMA"


def import_peru() -> dict:
    """Import 2024 Consulta Amigable revenue and spending for 1,891 municipalities."""
    root = CACHE / "PER"
    paths = {"revenue": root / "2024-Ingreso.zip", "expenditure": root / "2024-Gasto.zip"}
    if any(not path.exists() for path in paths.values()):
        raise FileNotFoundError("Download the official 2024 MEF revenue and expenditure ZIPs first")
    stage_fields = {
        "revenue": (("enacted", "MONTO_PIA"), ("revised", "MONTO_PIM"), ("actual", "MONTO_RECAUDADO")),
        "expenditure": (("enacted", "MONTO_PIA"), ("revised", "MONTO_PIM"), ("actual", "MONTO_DEVENGADO"), ("cash", "MONTO_GIRADO")),
    }
    totals = {side: defaultdict(lambda: defaultdict(float)) for side in stage_fields}
    details = {side: defaultdict(lambda: defaultdict(float)) for side in stage_fields}
    registry: dict[str, dict] = {}
    for side, path in paths.items():
        with zipfile.ZipFile(path) as archive, archive.open(archive.namelist()[0]) as raw:
            rows = csv.reader(io.TextIOWrapper(raw, encoding="utf-8-sig", newline=""))
            header = next(rows)
            index = {value: column for column, value in enumerate(header)}
            account_field = "GENERICA" if side == "revenue" else "FUNCION"
            label_field = "GENERICA_NOMBRE" if side == "revenue" else "FUNCION_NOMBRE"
            for row_count, values in enumerate(rows, 1):
                if values[index["NIVEL_GOBIERNO"]] != "M":
                    continue
                native_name = clean(values[index["EJECUTORA_NOMBRE"]])
                if not peru_is_municipality(native_name):
                    continue
                code = clean(values[index["SEC_EJEC"]])
                registry.setdefault(code, {
                    "name": re.sub(r"^MUNICIPALIDAD (?:PROVINCIAL|DISTRITAL|METROPOLITANA) (?:DE )?", "", native_name).title(),
                    "region": clean(values[index["DEPARTAMENTO_EJECUTORA_NOMBRE"]]).title(),
                    "government_type": native_name.split(" ", 2)[1].title(),
                })
                account = clean(values[index[account_field]])
                label = clean(values[index[label_field]])
                for stage, field in stage_fields[side]:
                    amount = number(values[index[field]]) or 0
                    totals[side][code][stage] += amount
                    details[side][code][(stage, account, label)] += amount
                if row_count % 5_000_000 == 0:
                    print(f"Peru {side}: {row_count:,} source rows", flush=True)
    reporters = set(totals["revenue"]) & set(totals["expenditure"])
    profiles = []
    for code in reporters:
        entity = registry[code]
        native_detail = []
        for side in ("revenue", "expenditure"):
            for (stage, account, label), amount in sorted(details[side][code].items()):
                native_detail.append({"year": 2024, "stage": stage, "side": side, "code": account, "name": label, "amount": round(amount, 2)})
            for stage, _ in stage_fields[side]:
                reconciled = sum(row["amount"] for row in native_detail if row["side"] == side and row["stage"] == stage)
                if abs(reconciled - totals[side][code][stage]) > 0.5:
                    raise ValueError(f"Peru {side}/{stage} detail does not reconcile for {code}")
        revenue = round(totals["revenue"][code]["actual"], 2)
        expenditure = round(totals["expenditure"][code]["actual"], 2)
        profiles.append({
            "code": code, "name": entity["name"], "region": entity["region"], "government_type": entity["government_type"],
            "country": "PER", "currency": "PEN", "years": [2024],
            "history": [{"year": 2024, "revenue": revenue, "expenditure": expenditure, "balance": round(revenue - expenditure, 2)}],
            "detail": native_detail, "url": f"/municipalities/peru/{slugify(entity['name'])}-{code}/",
        })
    if len(profiles) != 1891:
        raise ValueError(f"Peru import yielded {len(profiles)} paired municipalities, expected 1,891")
    return {
        "country": "PER", "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "https://datosabiertos.mef.gob.pe/dataset/presupuesto-y-ejecucion-de-gasto",
        "period": 2024, "entities": sorted(profiles, key=lambda row: (row["region"], row["name"])), "errors": [],
    }


def import_south_korea() -> dict:
    """Import 2024 Local Finance 365 settlements for all 243 local governments."""
    root = CACHE / "KOR"
    paths = {
        "revenue": sorted(root.glob("revenue-*.json")),
        "expenditure": sorted(root.glob("expense-*.json")),
    }
    if any(len(group) != 3 for group in paths.values()):
        raise FileNotFoundError("Download the three official Local Finance 365 JSON pages for each side first")
    totals: dict[str, dict[str, float]] = {"revenue": {}, "expenditure": {}}
    registry: dict[str, dict[str, str]] = {}
    detail: dict[str, list[dict]] = defaultdict(list)
    account_names = {
        "1": "General account",
        "2": "Public-enterprise special accounts",
        "3": "Other special accounts",
        "4": "Funds",
    }
    for side, files in paths.items():
        total_field = "total" if side == "revenue" else "totPfaAmt"
        for path in files:
            payload = json.loads(path.read_text(encoding="utf-8"))
            for row in payload.get("resultList", []):
                code = clean(row.get("lafCd"))
                if not code or clean(row.get("fyr")) != "2024":
                    continue
                registry.setdefault(code, {
                    "name": clean(row.get("lafHgNm")),
                    "region": clean(row.get("waLafHgNm")),
                    "government_type": "upper-tier" if clean(row.get("lafHgNm")).endswith("본청") else "basic local government",
                })
                total = float(row.get(total_field) or 0)
                components = [float(row.get(f"pfaAmt{index}") or 0) for index in range(1, 5)]
                if abs(total - sum(components)) > 1:
                    raise ValueError(f"South Korea {side} accounts do not reconcile for {code}")
                totals[side][code] = total
                for index, amount in enumerate(components, 1):
                    detail[code].append({
                        "year": 2024, "stage": "actual", "side": side, "code": str(index),
                        "name": account_names[str(index)], "amount": amount,
                    })
    reporters = set(totals["revenue"]) & set(totals["expenditure"])
    profiles = []
    for code in reporters:
        entity = registry[code]
        revenue, expenditure = totals["revenue"][code], totals["expenditure"][code]
        profiles.append({
            "code": code, "name": entity["name"], "region": entity["region"],
            "government_type": entity["government_type"], "country": "KOR", "currency": "KRW", "years": [2024],
            "history": [{"year": 2024, "revenue": revenue, "expenditure": expenditure, "balance": revenue - expenditure}],
            "detail": detail[code], "url": f"/municipalities/south-korea/{slugify(entity['name'])}-{code}/",
        })
    if len(profiles) != 243:
        raise ValueError(f"South Korea import yielded {len(profiles)} paired local governments, expected 243")
    if sum(profile["government_type"] == "upper-tier" for profile in profiles) != 17:
        raise ValueError("South Korea import did not identify the expected 17 upper-tier governments")
    return {
        "country": "KOR", "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "https://www.lofin365.go.kr/portal/LF5100000.do", "period": 2024,
        "entities": sorted(profiles, key=lambda row: (row["region"], row["name"])), "errors": [],
    }


def post_json(url: str, data: dict) -> object:
    body = urllib.parse.urlencode(data, doseq=True).encode()
    request = urllib.request.Request(url, data=body, headers={
        "User-Agent": USER_AGENT,
        "Referer": "https://datos.sinim.gov.cl/evolucion_presupuestaria.php",
        "X-Requested-With": "XMLHttpRequest",
    })
    with urllib.request.urlopen(request, timeout=120) as response:
        return json.load(response)


def ensure_chile_source() -> tuple[Path, Path]:
    """Cache SINIM's official 2025 roster and municipal evolution responses."""
    root = CACHE / "CHL"
    data_root = root / "2025"
    roster_path = root / "municipalities.json"
    root.mkdir(parents=True, exist_ok=True)
    data_root.mkdir(parents=True, exist_ok=True)
    if not roster_path.exists():
        page = urllib.request.urlopen(urllib.request.Request(
            "https://datos.sinim.gov.cl/evolucion_presupuestaria.php", headers={"User-Agent": USER_AGENT}
        ), timeout=120).read().decode("utf-8", errors="replace")
        select = re.search(r'<select id="regiones".*?</select>', page, re.S)
        if not select:
            raise ValueError("Could not resolve the SINIM region list")
        regions = [(code, re.sub(r"\s+", " ", html.unescape(name)).strip()) for code, name in re.findall(r'<option value="(\d+)">(.*?)</option>', select.group())]
        roster = []
        endpoint = "https://datos.sinim.gov.cl/evolucion_presupuestaria/obtener_datos_filtros.php"
        for region_code, region_name in regions:
            provinces = post_json(endpoint, {"region": region_code})
            for province in provinces:
                municipalities = post_json(endpoint, {"region": region_code, "provincia": province["id"]})
                for municipality in municipalities:
                    roster.append({
                        "code": clean(municipality["id_legal"]), "source_id": clean(municipality["id"]),
                        "name": clean(municipality["municipio"]).title(), "region": region_name.title(),
                        "region_id": region_code, "province": clean(province["provincia"]).title(),
                        "province_id": clean(province["id"]),
                    })
        if len(roster) != 345 or len({row["code"] for row in roster}) != 345:
            raise ValueError(f"SINIM roster yielded {len(roster)} municipalities, expected 345 unique codes")
        roster_path.write_text(json.dumps(roster, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    roster = json.loads(roster_path.read_text(encoding="utf-8"))
    endpoint = "https://datos.sinim.gov.cl/evolucion_presupuestaria/obtener_datos.php"

    def fetch(entity: dict) -> str:
        target = data_root / f"{entity['code']}.json"
        if target.exists() and target.stat().st_size:
            return entity["code"]
        payload = post_json(endpoint, {
            "sector": "9", "periodos[]": ["26"], "regiones": entity["region_id"],
            "provincias[]": [entity["province_id"]], "municipios[]": [entity["source_id"]], "corrmon": "0",
        })
        target.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
        return entity["code"]

    missing = [row for row in roster if not (data_root / f"{row['code']}.json").exists()]
    if missing:
        with ThreadPoolExecutor(max_workers=12) as executor:
            futures = {executor.submit(fetch, row): row["code"] for row in missing}
            for completed, future in enumerate(as_completed(futures), 1):
                future.result()
                if completed % 50 == 0:
                    print(f"Chile: cached {completed}/{len(missing)} municipal responses", flush=True)
    return roster_path, data_root


def chile_amount(value: object) -> float | None:
    text = clean(value).replace(".", "").replace(",", ".")
    if not text or text in {"-", "—"}:
        return None
    try:
        # SINIM labels the evolution table M$: thousands of Chilean pesos.
        return float(text) * 1_000
    except ValueError:
        return None


def import_chile() -> dict:
    """Import 2025 SINIM municipal revenue and expenditure for all 345 communes."""
    roster_path, data_root = ensure_chile_source()
    roster = json.loads(roster_path.read_text(encoding="utf-8"))
    profiles = []
    missing_finance = []
    for entity in roster:
        payload = json.loads((data_root / f"{entity['code']}.json").read_text(encoding="utf-8"))
        values = {int(row["id_area"]): chile_amount(row.get("valor")) for row in payload.get("valores", {}).get("26", []) if row.get("id_area")}
        revenue, expenditure = values.get(21), values.get(47)
        if revenue is None or expenditure is None:
            missing_finance.append(entity["code"])
            continue
        reported_balance = values.get(2)
        if reported_balance is not None and abs(reported_balance - (revenue - expenditure)) > 1_000:
            raise ValueError(f"Chile reported balance does not reconcile for {entity['code']}")
        native_detail = []
        for section_key, side in (("1", "revenue"), ("2", "expenditure")):
            for area in payload.get("evolucion", {}).get(section_key, {}).get("areas", []):
                amount = values.get(int(area["id_area"]))
                if amount is None:
                    continue
                native_detail.append({
                    "year": 2025, "stage": "actual", "side": side, "code": clean(area["id_area"]),
                    "name": clean(area["nombre"]), "level": clean(area.get("tipo")), "amount": amount,
                })
        profiles.append({
            "code": entity["code"], "name": entity["name"], "region": entity["region"], "province": entity["province"],
            "country": "CHL", "currency": "CLP", "years": [2025],
            "history": [{"year": 2025, "revenue": revenue, "expenditure": expenditure, "balance": revenue - expenditure}],
            "detail": native_detail, "url": f"/municipalities/chile/{slugify(entity['name'])}-{entity['code']}/",
        })
    status = "complete" if not missing_finance else "partial"
    if len(profiles) + len(missing_finance) != 345:
        raise ValueError("Chile entity-count validation failed")
    return {
        "country": "CHL", "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "https://datos.sinim.gov.cl/evolucion_presupuestaria.php", "period": 2025,
        "status": status, "missing_finance": missing_finance,
        "entities": sorted(profiles, key=lambda row: (row["region"], row["name"])), "errors": [],
    }


def write_bundle(bundle: dict, filename: str | None = None) -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    country = bundle["country"]
    target = OUTPUT / (filename or f"{country}.json")
    target.write_text(json.dumps(bundle, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    print(json.dumps({"country": country, "entities": len(bundle["entities"]), "errors": len(bundle.get("errors", [])), "output": str(target)}, ensure_ascii=False))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("country", choices=("ESP", "JPN", "BRA", "COL", "GEO", "ITA", "BOL", "SLV", "MEX", "CRI", "GTM", "PER", "KOR", "CHL"))
    parser.add_argument("--limit", type=int)
    parser.add_argument("--shard-index", type=int, default=0)
    parser.add_argument("--shard-count", type=int, default=1)
    args = parser.parse_args()
    if args.country == "ESP": write_bundle(import_spain())
    elif args.country == "JPN": write_bundle(import_japan())
    elif args.country == "COL": write_bundle(import_colombia())
    elif args.country == "GEO": write_bundle(import_georgia())
    elif args.country == "ITA": write_bundle(import_italy())
    elif args.country == "BOL": write_bundle(import_bolivia())
    elif args.country == "SLV": write_bundle(import_el_salvador())
    elif args.country == "MEX": write_bundle(import_mexico())
    elif args.country == "CRI": write_bundle(import_costa_rica())
    elif args.country == "GTM": write_bundle(import_guatemala())
    elif args.country == "PER": write_bundle(import_peru())
    elif args.country == "KOR": write_bundle(import_south_korea())
    elif args.country == "CHL": write_bundle(import_chile())
    else:
        if not 0 <= args.shard_index < args.shard_count:
            parser.error("--shard-index must be between zero and --shard-count minus one")
        filename = f"BRA-shard-{args.shard_index}.json" if args.shard_count > 1 else None
        write_bundle(import_brazil(limit=args.limit, shard_index=args.shard_index, shard_count=args.shard_count), filename)


if __name__ == "__main__":
    main()
