#!/usr/bin/env python3
"""Build the defence deep-dive dataset from public, reusable source files.

The comparison series is downloaded from World Bank WDI (SIPRI source,
CC BY 4.0). Native budget detail reuses the official national files already
tracked by the country-spending pipeline and adds title/account-level adapters
where the downloaded files expose that level cleanly.
"""

from __future__ import annotations

import csv
import json
import urllib.request
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = ROOT.parent / "data" / "sources"
DEFENCE_SOURCE_ROOT = SOURCE_ROOT / "defense"
OUT = ROOT / "data" / "defense-deep-dive.v1.json"
SPENDING = ROOT / "data" / "country-spending-2025-2026.v1.json"
COMPARISON = ROOT / "data" / "country-spending-comparison.v1.json"

COUNTRY_NAMES = {
    "USA": ("Spojené státy", "United States"), "CZE": ("Česko", "Czechia"),
    "DEU": ("Německo", "Germany"), "DNK": ("Dánsko", "Denmark"),
    "FIN": ("Finsko", "Finland"), "FRA": ("Francie", "France"),
    "GBR": ("Spojené království", "United Kingdom"), "POL": ("Polsko", "Poland"),
    "SWE": ("Švédsko", "Sweden"), "CHE": ("Švýcarsko", "Switzerland"),
    "UKR": ("Ukrajina", "Ukraine"), "BRA": ("Brazílie", "Brazil"),
    "ESP": ("Španělsko", "Spain"), "JPN": ("Japonsko", "Japan"),
    "NLD": ("Nizozemsko", "Netherlands"), "NOR": ("Norsko", "Norway"),
    "GRC": ("Řecko", "Greece"),
}

NATO = {"USA", "CZE", "DEU", "DNK", "FIN", "FRA", "GBR", "POL", "SWE", "ESP", "NLD", "NOR", "GRC"}

WDI_URL = (
    "https://api.worldbank.org/v2/country/"
    + ";".join(COUNTRY_NAMES)
    + "/indicator/MS.MIL.XPND.GD.ZS?format=json&per_page=2000"
)


def download_wdi() -> dict:
    DEFENCE_SOURCE_ROOT.mkdir(parents=True, exist_ok=True)
    target = DEFENCE_SOURCE_ROOT / "world-bank-military-expenditure-pct-gdp.json"
    request = urllib.request.Request(WDI_URL, headers={"User-Agent": "publicspendingdata.org data build"})
    with urllib.request.urlopen(request, timeout=60) as response:
        payload = json.load(response)
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return payload


def wdi_series(payload: dict) -> dict[str, list[list[float]]]:
    result: dict[str, list[list[float]]] = defaultdict(list)
    for row in payload[1]:
        code = row.get("countryiso3code")
        value = row.get("value")
        if code in COUNTRY_NAMES and value is not None:
            result[code].append([int(row["date"]), round(float(value), 4)])
    for rows in result.values():
        rows.sort(key=lambda item: item[0])
    return result


def baseline_budgets() -> dict[str, dict]:
    spending = json.loads(SPENDING.read_text(encoding="utf-8"))
    comparison = json.loads(COMPARISON.read_text(encoding="utf-8"))
    spending_by_code = {row["code"]: row for row in spending["countries"]}
    result = {}
    for compact in comparison["countries"]:
        code = compact["code"]
        native = spending_by_code[code]
        defence_group = next(group for group in compact["groups"] if group["category_id"] == "defence")
        defence_codes = {row["code"] for row in defence_group["source_rows"]}
        rows = [row for row in native["rows"] if row["code"] in defence_codes]
        items = [{
            "id": row["code"], "parent": native["scope_en"],
            "label_native": row["label_native"], "label_en": row.get("label_en"),
            "amount": round(row["amounts"]["current"], 6),
        } for row in rows]
        result[code] = {
            "period": native["periods"]["current"]["label"],
            "status_cs": native["periods"]["current"]["status_cs"],
            "status_en": native["periods"]["current"]["status_en"],
            "currency": native["currency"], "unit": native["unit"],
            "dimension": native["dimension"],
            "scope_cs": native["scope_cs"], "scope_en": native["scope_en"],
            "granularity": "national_budget_top_line",
            "coverage_note_cs": "Nejjemnější strojově čitelná obranná položka v aktuálně staženém národním rozpočtovém zdroji.",
            "coverage_note_en": "Finest machine-readable defence line in the currently downloaded national budget source.",
            "items": items, "sources": native["sources"],
        }
    return result


def german_detail() -> dict:
    path = SOURCE_ROOT / "ministries" / "DEU" / "federal-budget-2026.csv"
    items = []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle, delimiter=";"):
            if row["einzelplan"] != "14" or row["einahmen-ausgaben"] != "A":
                continue
            amount = float(row["soll "] or 0) / 1000  # source EUR thousands -> EUR millions
            if amount == 0:
                continue
            items.append({
                "id": f'{row["kapitel"]}-{row["titel"]}',
                "parent": row["kapitel-text"], "label_native": row["titel-text"],
                "label_en": None, "amount": round(amount, 6),
                "classification": row["einahmen-ausgaben-text"],
            })
    return {
        "period": "2026", "status_cs": "schválený spolkový rozpočet", "status_en": "enacted federal budget",
        "currency": "EUR", "unit": "million_local_currency", "dimension": "administrative_title",
        "scope_cs": "Oddíl 14 spolkového rozpočtu — jednotlivé kapitoly a tituly",
        "scope_en": "Federal budget section 14 — individual chapters and titles",
        "granularity": "budget_title", "items": sorted(items, key=lambda row: row["amount"], reverse=True),
        "coverage_note_cs": "Každý nenulový výdajový titul oddílu 14; zvláštní mimorozpočtové fondy nejsou sloučeny do této tabulky.",
        "coverage_note_en": "Every non-zero expenditure title in section 14; separate off-budget special funds are not merged into this table.",
        "sources": [{"title": "Bundeshaushalt 2026 — Open Data", "url": "https://www.bundeshaushalt.de/DE/Bundeshaushalt-digital/bundeshaushalt-digital.html"}],
    }


def french_detail() -> dict:
    path = SOURCE_ROOT / "ministries" / "FRA" / "state-budget-execution-2024.csv"
    items = []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle, delimiter=";"):
            if row["Mission"].strip().casefold() != "défense":
                continue
            amount = float(row["Depenses_constatees"].replace(",", ".") or 0) / 1_000_000
            if amount == 0:
                continue
            items.append({
                "id": row["Programme"].rsplit(" - ", 1)[-1] + "-" + row["Titre"].split(" - ", 1)[0],
                "parent": row["Programme"], "label_native": row["Titre"], "label_en": None,
                "amount": round(amount, 6),
            })
    return {
        "period": "2024", "status_cs": "skutečné výdaje", "status_en": "actual expenditure",
        "currency": "EUR", "unit": "million_local_currency", "dimension": "programme_economic_title",
        "scope_cs": "Mise Défense podle programů a ekonomických titulů",
        "scope_en": "Défense mission by programme and economic title",
        "granularity": "programme_by_economic_title", "items": sorted(items, key=lambda row: row["amount"], reverse=True),
        "coverage_note_cs": "Skutečné platby mise Défense. Samostatná mise pro veterány a některé meziresortní bezpečnostní výdaje jsou mimo záběr.",
        "coverage_note_en": "Actual payments of the Défense mission. The separate veterans mission and some cross-ministry security spending are outside scope.",
        "sources": [{"title": "data.gouv.fr — exécution du budget de l’État 2024", "url": "https://www.data.gouv.fr/fr/datasets/execution-du-budget-de-letat/"}],
    }


def us_detail() -> dict:
    path = SOURCE_ROOT / "ministries" / "USA" / "public-budget-database-outlays-fy2027.xlsx"
    sheet = load_workbook(path, read_only=True, data_only=True).active
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    positions = {name: header.index(name) for name in (
        "Agency Code", "Agency Name", "Bureau Code", "Bureau Name", "Account Code", "Account Name",
        "Subfunction Code", "Subfunction Title", "BEA Category", "2026",
    )}
    items = []
    for row in rows:
        subfunction = str(row[positions["Subfunction Code"]] or "").zfill(3)
        amount_raw = row[positions["2026"]]
        if subfunction not in {"051", "053", "054"} or not isinstance(amount_raw, (int, float)) or amount_raw == 0:
            continue
        agency_code = str(row[positions["Agency Code"]] or "")
        bureau_code = str(row[positions["Bureau Code"]] or "")
        account_code = str(row[positions["Account Code"]] or "")
        items.append({
            "id": f"{agency_code}-{bureau_code}-{account_code}-{subfunction}",
            "parent": f'{row[positions["Agency Name"]]} / {row[positions["Bureau Name"]]}',
            "label_native": str(row[positions["Account Name"]] or row[positions["Subfunction Title"]]),
            "label_en": str(row[positions["Account Name"]] or row[positions["Subfunction Title"]]),
            "amount": round(float(amount_raw) / 1000, 6),  # source USD thousands -> USD millions
            "classification": str(row[positions["BEA Category"]] or ""),
            "subfunction": subfunction,
        })
    return {
        "period": "FY 2026", "status_cs": "odhad výdajů v rozpočtu FY 2027", "status_en": "FY 2027 Budget estimate",
        "currency": "USD", "unit": "million_local_currency", "dimension": "federal_budget_account",
        "scope_cs": "Federální funkce National Defense 050 — účty v podfunkcích 051, 053 a 054",
        "scope_en": "Federal National Defense function 050 — accounts in subfunctions 051, 053 and 054",
        "granularity": "federal_budget_account", "items": sorted(items, key=lambda row: row["amount"], reverse=True),
        "coverage_note_cs": "Jednotlivé federální účty napříč ministerstvy; záporné položky jsou v datech zachovány jako zápočty nebo příjmy.",
        "coverage_note_en": "Individual federal accounts across agencies; negative lines remain visible as offsets or receipts.",
        "sources": [{"title": "US OMB — Public Budget Database, FY 2027 Budget", "url": "https://www.govinfo.gov/app/collection/budget/2027/BUDGET-2027-DB"}],
    }


def main() -> None:
    wdi_payload = download_wdi()
    history = wdi_series(wdi_payload)
    budgets = baseline_budgets()
    budgets["DEU"] = german_detail()
    budgets["FRA"] = french_detail()
    budgets["USA"] = us_detail()

    countries = []
    for code in COUNTRY_NAMES:
        budget = budgets[code]
        budget["item_count"] = len(budget["items"])
        budget["total_amount"] = round(sum(row["amount"] for row in budget["items"]), 6)
        series = history.get(code, [])
        cs, en = COUNTRY_NAMES[code]
        countries.append({
            "code": code, "name_cs": cs, "name_en": en, "nato_member": code in NATO,
            "comparison": {"indicator": "MS.MIL.XPND.GD.ZS", "unit": "pct_gdp", "series": series,
                           "latest": {"year": series[-1][0], "value": series[-1][1]} if series else None},
            "budget": budget,
        })

    output = {
        "schema_version": "1.0.0", "dataset_id": "DEFENSE_DEEP_DIVE_V1",
        "generated_at": date.today().isoformat(), "default_country": "USA",
        "commitments": {
            "nato_core_pct_gdp_2035": 3.5, "nato_broader_security_pct_gdp_2035": 1.5,
            "nato_total_pct_gdp_2035": 5.0, "legacy_nato_floor_pct_gdp": 2.0,
            "source_url": "https://www.nato.int/en/about-us/official-texts-and-resources/official-texts/2025/06/25/the-hague-summit-declaration",
        },
        "comparison_source": {
            "title": "World Bank WDI — Military expenditure (% of GDP)", "indicator": "MS.MIL.XPND.GD.ZS",
            "underlying_source": "SIPRI Military Expenditure Database", "license": "CC BY 4.0",
            "url": "https://data.worldbank.org/indicator/MS.MIL.XPND.GD.ZS",
            "last_updated": wdi_payload[0].get("lastupdated"),
            "method_note_en": "SIPRI-based military expenditure and national/NATO budget definitions can differ. The target marker is context, not a compliance ruling.",
            "method_note_cs": "Výdaje podle SIPRI a národní či NATO rozpočtové definice se mohou lišit. Cílová čára je kontext, nikoli verdikt o plnění.",
        },
        "countries": countries,
    }
    OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT}: {len(countries)} countries, {sum(c['budget']['item_count'] for c in countries)} budget lines")


if __name__ == "__main__":
    main()
