#!/usr/bin/env python3
"""Store official annual age/sex projections and derive comparable age bands."""

from __future__ import annotations

import csv
import gzip
import io
import json
import re
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
COMMON_YEARS = tuple(range(2025, 2046))
DETAIL_COLUMNS = ["year", "age_start", "age_end", "male", "female", "total"]
EUROSTAT = "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/proj_23np"
US_CENSUS = "https://www2.census.gov/programs-surveys/popproj/datasets/2023/2023-popproj/np2023_d1_mid.csv"
ONS_ZIP = "https://www.ons.gov.uk/file?uri=%2Fpeoplepopulationandcommunity%2Fpopulationandmigration%2Fpopulationprojections%2Fdatasets%2Fz1zippedpopulationprojectionsdatafilesuk%2F2024based%2Fuk1.zip"
UN_WPP = "https://population.un.org/wpp/assets/Excel%20Files/1_Indicator%20(Standard)/CSV_FILES/WPP2024_PopulationBySingleAgeSex_Medium_2024-2100.csv.gz"
CZE_FILES = {
    "total": "https://csu.gov.cz/docs/107508/fd8ee566-2c71-69c4-50d4-72d9639079ac/1301392301.xlsx?version=1.0",
    "male": "https://csu.gov.cz/docs/107508/e579b776-3689-aa8e-34a1-19710cf9eb89/1301392302.xlsx?version=1.0",
    "female": "https://csu.gov.cz/docs/107508/ed564e3c-bdfd-2c7d-9b9f-24ceab02da9b/1301392303.xlsx?version=1.0",
}


def download(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "PublicSpendingData/1.0"})
    with urllib.request.urlopen(request, timeout=240) as response:
        return response.read()


def rounded(value) -> int:
    return round(float(value or 0))


def age_group(label):
    text = str(label).strip()
    numbers = [int(value) for value in re.findall(r"\d+", text)]
    if not numbers:
        raise ValueError(f"Unrecognised age label: {label}")
    start = numbers[0]
    if "+" in text or "over" in text.lower() or "GE" in text:
        return start, None
    return start, numbers[-1]


def compact_rows(values: dict) -> list:
    rows = []
    for (year, start, end), sexes in sorted(values.items(), key=lambda item: (item[0][0], item[0][1])):
        male, female = rounded(sexes.get("male")), rounded(sexes.get("female"))
        total = rounded(sexes.get("total", male + female))
        if abs(total - (male + female)) > 2:
            raise ValueError(f"Sex totals do not reconcile for {year}/{start}: {male}+{female}!={total}")
        rows.append([year, start, end, male, female, male + female])
    return rows


def xlsx_age_matrix(content: bytes) -> dict:
    sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
    matrix = list(sheet.iter_rows(values_only=True))
    years = {}
    for index, value in enumerate(matrix[2][1:], start=1):
        match = re.search(r"\d{4}", str(value))
        if match:
            years[index] = int(match.group())
    output = {}
    for values in matrix[4:]:
        if values[0] is None:
            break
        try:
            start, end = age_group(values[0])
        except ValueError:
            continue
        for column, year in years.items():
            output[(year, start, end)] = rounded(values[column])
    return output


def czechia() -> dict:
    matrices = {sex: xlsx_age_matrix(download(url)) for sex, url in CZE_FILES.items()}
    values = {key: {sex: matrix[key] for sex, matrix in matrices.items()} for key in matrices["total"]}
    return {
        "coverage": "national_middle_variant_full_age_sex",
        "projection": "ČSÚ 2023–2100, střední varianta",
        "reference_date": "1 January",
        "rows": compact_rows(values),
        "source": {
            "publisher": "Český statistický úřad",
            "dataset": "Projekce obyvatelstva České republiky 2023–2100",
            "url": "https://csu.gov.cz/produkty/projekce-obyvatelstva-ceske-republiky-2023-2100",
            "download_urls": list(CZE_FILES.values()),
            "location": "Tables 1–3, middle variant, both sexes/males/females; annual columns; ages 0–99 and 100+",
        },
    }


def jsonstat_value(payload: dict, coordinates: dict):
    flat = 0
    for dimension, size in zip(payload["id"], payload["size"]):
        category = payload["dimension"][dimension]["category"]["index"]
        selected = coordinates.get(dimension)
        position = category[selected] if selected is not None else 0
        flat = (flat * size) + position
    return payload.get("value", {}).get(str(flat))


def eurostat_country(geo: str) -> dict:
    params = [("lang", "en"), ("geo", geo), ("projection", "BSL")]
    payload = json.loads(download(EUROSTAT + "?" + urllib.parse.urlencode(params)))
    ages = payload["dimension"]["age"]["category"]["index"]
    times = payload["dimension"]["time"]["category"]["index"]
    single_ages = []
    for code in ages:
        if code == "Y_LT1":
            single_ages.append((code, 0, 0))
        elif re.fullmatch(r"Y\d+", code):
            value = int(code[1:])
            single_ages.append((code, value, value))
        elif code == "Y_GE100":
            single_ages.append((code, 100, None))
    values = {}
    for year_code in times:
        year = int(year_code)
        for age_code, start, end in single_ages:
            values[(year, start, end)] = {
                "male": jsonstat_value(payload, {"sex": "M", "age": age_code, "time": year_code}),
                "female": jsonstat_value(payload, {"sex": "F", "age": age_code, "time": year_code}),
                "total": jsonstat_value(payload, {"sex": "T", "age": age_code, "time": year_code}),
            }
    return {
        "coverage": "eurostat_baseline_full_age_sex",
        "projection": "EUROPOP2023 baseline scenario",
        "reference_date": "1 January",
        "rows": compact_rows(values),
        "source": {
            "publisher": "Eurostat",
            "dataset": "Population projections in the EU (proj_23np)",
            "url": "https://ec.europa.eu/eurostat/databrowser/view/proj_23np/default/table",
            "api_url": EUROSTAT,
            "location": f"API proj_23np; geo={geo}; projection=BSL; sex=M/F/T; annual time dimension; ages Y_LT1, Y1…Y99 and Y_GE100",
        },
    }


def united_kingdom() -> dict:
    archive = zipfile.ZipFile(io.BytesIO(download(ONS_ZIP)))
    workbook = load_workbook(io.BytesIO(archive.read("uk_ppp_machine_readable.xlsx")), read_only=True, data_only=True)
    rows = workbook["Population"].iter_rows(values_only=True)
    years = {index: int(value) for index, value in enumerate(next(rows)) if str(value).isdigit()}
    values = {}
    for record in rows:
        sex = {"Males": "male", "Females": "female"}.get(record[0])
        if not sex:
            continue
        start, end = age_group(record[1])
        for column, year in years.items():
            values.setdefault((year, start, end), {})[sex] = rounded(record[column])
    return {
        "coverage": "national_principal_full_age_sex",
        "projection": "ONS 2024-based principal projection",
        "reference_date": "mid-year",
        "rows": compact_rows(values),
        "source": {
            "publisher": "Office for National Statistics",
            "dataset": "National population projections: 2024-based",
            "url": "https://www.ons.gov.uk/peoplepopulationandcommunity/populationandmigration/populationprojections/datasets/z1zippedpopulationprojectionsdatafilesuk",
            "download_urls": [ONS_ZIP],
            "location": "UK1.zip → uk_ppp_machine_readable.xlsx → Population; males/females; annual columns; ages 0–104, 105–109 and 110+",
        },
    }


def united_states() -> dict:
    records = csv.DictReader(io.StringIO(download(US_CENSUS).decode("utf-8-sig")))
    values = {}
    for record in records:
        if record["ORIGIN"] != "0" or record["RACE"] != "0" or record["SEX"] not in ("0", "1", "2"):
            continue
        year = int(record["YEAR"])
        sex = {"0": "total", "1": "male", "2": "female"}[record["SEX"]]
        for age in range(101):
            values.setdefault((year, age, age if age < 100 else None), {})[sex] = int(record[f"POP_{age}"])
    return {
        "coverage": "national_middle_series_full_age_sex",
        "projection": "2023 National Population Projections, middle series",
        "reference_date": "1 July",
        "rows": compact_rows(values),
        "source": {
            "publisher": "U.S. Census Bureau",
            "dataset": "2023 National Population Projections — Main Series",
            "url": "https://www.census.gov/data/tables/2023/demo/popproj/2023-summary-tables.html",
            "download_urls": [US_CENSUS],
            "location": "np2023_d1_mid.csv; ORIGIN=0; RACE=0; SEX=0/1/2; every annual row; POP_0…POP_99 and POP_100 (100+)",
        },
    }


def ukraine() -> dict:
    compressed = io.BytesIO(download(UN_WPP))
    records = csv.DictReader(io.TextIOWrapper(gzip.GzipFile(fileobj=compressed), encoding="utf-8-sig"))
    values = {}
    for record in records:
        if record["ISO3_code"] != "UKR":
            continue
        year, start = int(record["Time"]), int(record["AgeGrpStart"])
        span = int(record["AgeGrpSpan"])
        end = None if span < 0 or "+" in record["AgeGrp"] else start + span - 1
        values[(year, start, end)] = {
            "male": float(record["PopMale"]) * 1000,
            "female": float(record["PopFemale"]) * 1000,
            "total": float(record["PopTotal"]) * 1000,
        }
    return {
        "coverage": "un_medium_variant_full_age_sex",
        "projection": "UN World Population Prospects 2024, medium variant",
        "reference_date": "1 July",
        "rows": compact_rows(values),
        "source": {
            "publisher": "United Nations, Population Division",
            "dataset": "World Population Prospects 2024",
            "url": "https://population.un.org/wpp/Download/Standard/CSV/",
            "download_urls": [UN_WPP],
            "location": "WPP2024_PopulationBySingleAgeSex_Medium_2024-2100.csv.gz; ISO3_code=UKR; annual Time; AgeGrpStart/AgeGrpSpan; PopMale/PopFemale/PopTotal",
        },
    }


def aggregate(rows: list) -> list:
    bands = ("age_0_19", "age_20_64", "age_65_79", "age_80_plus")
    by_year = {}
    for year, start, _end, male, female, total in rows:
        if year not in COMMON_YEARS:
            continue
        band = "age_0_19" if start < 20 else "age_20_64" if start < 65 else "age_65_79" if start < 80 else "age_80_plus"
        target = by_year.setdefault(year, {
            "year": year, "total": 0, "male": 0, "female": 0,
            **{key: 0 for key in bands},
            "male_by_age": {key: 0 for key in bands},
            "female_by_age": {key: 0 for key in bands},
        })
        target["total"] += total
        target["male"] += male
        target["female"] += female
        target[band] += total
        target["male_by_age"][band] += male
        target["female_by_age"][band] += female
    output = []
    for year in COMMON_YEARS:
        target = by_year[year]
        if target["total"] != target["male"] + target["female"]:
            raise ValueError(f"Annual sex totals do not reconcile for {year}")
        if target["total"] != sum(target[key] for key in bands):
            raise ValueError(f"Annual age bands do not reconcile for {year}")
        target["shares_pct"] = {key: round(target[key] / target["total"] * 100, 4) for key in bands}
        target["old_age_dependency_per_100_working_age"] = round(
            (target["age_65_79"] + target["age_80_plus"]) / target["age_20_64"] * 100, 4
        )
        output.append(target)
    return output


def profile_and_store(code: str, detail: dict, generated_at: str) -> dict:
    years = sorted({row[0] for row in detail["rows"]})
    age_groups = sorted({(row[1], row[2]) for row in detail["rows"]}, key=lambda value: value[0])
    shard = {
        "schema_version": "1.0.0", "contract": "country-demography-detail.v1", "generated_at": generated_at,
        "country_code": code, "projection": detail["projection"], "coverage": detail["coverage"],
        "reference_date": detail["reference_date"], "period": {"from": years[0], "to": years[-1]},
        "columns": DETAIL_COLUMNS, "row_count": len(detail["rows"]), "age_group_count": len(age_groups),
        "age_resolution": "single-year ages with a source-native open/grouped oldest-age tail",
        "source": detail["source"], "rows": detail["rows"],
    }
    directory = ROOT / "data" / "countries" / code.lower()
    directory.mkdir(parents=True, exist_ok=True)
    target = directory / "demography.v1.json"
    target.write_text(json.dumps(shard, ensure_ascii=False, separators=(",", ":")) + "\n")
    return {
        "coverage": detail["coverage"], "projection": detail["projection"], "reference_date": detail["reference_date"],
        "period": {"from": years[0], "to": years[-1]}, "detail": str(target.relative_to(ROOT)),
        "detail_row_count": len(detail["rows"]), "age_group_count": len(age_groups), "years": aggregate(detail["rows"]),
        "source": {**detail["source"], "period": f"{years[0]}–{years[-1]}"},
    }


def main() -> None:
    generated_at = datetime.now(timezone.utc).isoformat()
    details = {"CZE": czechia()}
    for code, geo in {"POL": "PL", "DEU": "DE", "FRA": "FR", "CHE": "CH", "SWE": "SE", "DNK": "DK"}.items():
        details[code] = eurostat_country(geo)
    details["GBR"] = united_kingdom()
    details["USA"] = united_states()
    details["UKR"] = ukraine()
    countries = {code: profile_and_store(code, detail, generated_at) for code, detail in details.items()}
    payload = {
        "schema_version": "2.0.0", "generated_at": generated_at, "contract": "country-demography.v1",
        "common_period": {"from": COMMON_YEARS[0], "to": COMMON_YEARS[-1], "frequency": "annual"},
        "common_age_bands": [
            {"id": "age_0_19", "from": 0, "to": 19}, {"id": "age_20_64", "from": 20, "to": 64},
            {"id": "age_65_79", "from": 65, "to": 79}, {"id": "age_80_plus", "from": 80, "to": None},
        ],
        "methodology": {
            "cs": "Oficiální hlavní/střední varianta je uložena po jednotlivých letech, věku a pohlaví v národních souborech. Společná pásma 0–19, 20–64, 65–79 a 80+ jsou počítána výhradně z těchto uložených řádků. Referenční datum zdroje zůstává uvedeno u země.",
            "en": "Each official principal/middle variant is stored by year, age and sex in a country shard. The common 0–19, 20–64, 65–79 and 80+ bands are calculated exclusively from those stored rows. Each country's source reference date is retained.",
        },
        "countries": countries,
    }
    (ROOT / "data" / "country-demography.v1.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
    print(f"Stored full demographic projection detail and annual aggregates for {len(countries)} countries")


if __name__ == "__main__":
    main()
