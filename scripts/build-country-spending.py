#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the two-period national spending explorer from official source files."""

from __future__ import annotations

import csv
import json
import re
import tempfile
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT.parent / "data" / "sources" / "ministries"
OUT = ROOT / "data" / "country-spending-2025-2026.v1.json"

TRANSLATIONS = {
    "CZE": {
        "Ministerstvo práce a sociálních věcí": "Ministry of Labour and Social Affairs", "Ministerstvo školství, mládeže a tělovýchovy": "Ministry of Education, Youth and Sports", "Všeobecná pokladní správa": "General Treasury Administration", "Ministerstvo obrany": "Ministry of Defence", "Ministerstvo dopravy": "Ministry of Transport", "Ministerstvo vnitra": "Ministry of the Interior", "Státní dluh": "State debt", "Ministerstvo průmyslu a obchodu": "Ministry of Industry and Trade", "Ministerstvo spravedlnosti": "Ministry of Justice", "Ministerstvo zemědělství": "Ministry of Agriculture", "Ministerstvo financí": "Ministry of Finance", "Ministerstvo kultury": "Ministry of Culture", "Ministerstvo zdravotnictví": "Ministry of Health", "Ministerstvo zahraničních věcí": "Ministry of Foreign Affairs", "Národní sportovní agentura": "National Sports Agency", "Ministerstvo pro místní rozvoj": "Ministry for Regional Development", "Akademie věd České republiky": "Czech Academy of Sciences", "Technologická agentura České republiky": "Technology Agency of the Czech Republic", "Správa státních hmotných rezerv": "Administration of State Material Reserves", "Ministerstvo životního prostředí": "Ministry of the Environment", "Grantová agentura České republiky": "Czech Science Foundation", "Český úřad zeměměřický a katastrální": "Czech Office for Surveying, Mapping and Cadastre", "Český telekomunikační úřad": "Czech Telecommunication Office", "Bezpečnostní informační služba": "Security Information Service", "Digitální a informační agentura": "Digital and Information Agency", "Poslanecká sněmovna Parlamentu": "Chamber of Deputies of Parliament", "Úřad vlády České republiky": "Office of the Government of the Czech Republic", "Český statistický úřad": "Czech Statistical Office", "Národní úřad pro kybernetickou a informační bezpečnost": "National Cyber and Information Security Agency", "Senát Parlamentu": "Senate of Parliament", "Nejvyšší kontrolní úřad": "Supreme Audit Office", "Generální inspekce bezpečnostních sborů": "General Inspection of Security Forces", "Státní úřad pro jadernou bezpečnost": "State Office for Nuclear Safety", "Národní bezpečnostní úřad": "National Security Authority", "Kancelář prezidenta republiky": "Office of the President of the Republic", "Energetický regulační úřad": "Energy Regulatory Office", "Úřad pro ochranu hospodářské soutěže": "Office for the Protection of Competition", "Ústavní soud": "Constitutional Court", "Úřad průmyslového vlastnictví": "Industrial Property Office", "Ústav pro studium totalitních režimů": "Institute for the Study of Totalitarian Regimes", "Kancelář veřejného ochránce práv a ochránce práv dětí": "Office of the Public Defender of Rights and Children's Ombudsman", "Český báňský úřad": "Czech Mining Authority", "Úřad pro ochranu osobních údajů": "Office for Personal Data Protection", "Rada pro rozhlasové a televizní vysílání": "Council for Radio and Television Broadcasting", "Úřad pro dohled nad hospodařením politických stran a hnutí": "Office for Oversight of Political Parties and Political Movements", "Úřad Národní rozpočtové rady": "Office of the Czech Fiscal Council", "Operace státních finančních aktiv": "State financial asset operations",
    },
    "DEU": {
        "Bundesministerium für Arbeit und Soziales": "Federal Ministry of Labour and Social Affairs", "Allgemeine Finanzverwaltung": "General financial administration", "Bundesministerium der Verteidigung": "Federal Ministry of Defence", "Bundesschuld": "Federal debt", "Bundesministerium für Verkehr": "Federal Ministry of Transport", "Bundesministerium für Forschung, Technologie und Raumfahrt": "Federal Ministry of Research, Technology and Space", "Bundesministerium für Gesundheit": "Federal Ministry of Health", "Bundesministerium für Bildung, Familie, Senioren, Frauen und Jugend": "Federal Ministry of Education, Family Affairs, Senior Citizens, Women and Youth", "Bundesministerium des Innern": "Federal Ministry of the Interior", "Bundesministerium für Landwirtschaft, Ernährung und Heimat": "Federal Ministry of Agriculture, Food and Regional Identity", "Bundesministerium der Finanzen": "Federal Ministry of Finance", "Bundesministerium für wirtschaftliche Zusammenarbeit und Entwicklung": "Federal Ministry for Economic Cooperation and Development", "Bundesministerium für Wohnen, Stadtentwicklung und Bauwesen": "Federal Ministry for Housing, Urban Development and Building", "Auswärtiges Amt": "Federal Foreign Office", "Bundesministerium für Wirtschaft und Energie": "Federal Ministry for Economic Affairs and Energy", "Bundeskanzler und Bundeskanzleramt": "Federal Chancellor and Federal Chancellery", "Bundesministerium für Umwelt, Klimaschutz, Naturschutz und nukleare Sicherheit": "Federal Ministry for the Environment, Climate Action, Nature Conservation and Nuclear Safety", "Bundesministerium für Digitales und Staatsmodernisierung": "Federal Ministry for Digital Transformation and Government Modernisation", "Deutscher Bundestag": "German Bundestag", "Bundesministerium der Justiz und für Verbraucherschutz": "Federal Ministry of Justice and Consumer Protection", "Bundesrechnungshof": "Federal Court of Audit", "Bundespräsident und Bundespräsidialamt": "Federal President and Office of the Federal President", "Die Bundesbeauftragte für den Datenschutz und die Informationsfreiheit": "Federal Commissioner for Data Protection and Freedom of Information", "Bundesverfassungsgericht": "Federal Constitutional Court", "Bundesrat": "Federal Council", "Unabhängiger Kontrollrat": "Independent Oversight Council",
    },
    "DNK": {
        "Indenrigs- og Sundhedsministeriet": "Ministry of the Interior and Health", "Beskæftigelsesministeriet": "Ministry of Employment", "Forsvarsministeriet": "Ministry of Defence", "Uddannelses- og Forskningsministeriet": "Ministry of Higher Education and Science", "Børne- og Undervisningsministeriet": "Ministry of Children and Education", "Finansministeriet": "Ministry of Finance", "Generelle reserver": "General reserves", "Pensionsvæsenet": "Pensions administration", "Transportministeriet": "Ministry of Transport", "Udenrigsministeriet": "Ministry of Foreign Affairs", "Justitsministeriet": "Ministry of Justice", "Kulturministeriet": "Ministry of Culture", "Social- og Boligministeriet": "Ministry of Social Affairs and Housing", "Skatteministeriet": "Ministry of Taxation", "Ministeriet for Grøn Trepart": "Ministry for the Green Tripartite", "Klima-, Energi- og Forsyningsministeriet": "Ministry of Climate, Energy and Utilities", "Udlændinge- og Integrationsministeriet": "Ministry of Immigration and Integration", "Erhvervsministeriet": "Ministry of Industry, Business and Financial Affairs", "Fødevareministeriet": "Ministry of Food, Agriculture and Fisheries", "Ældreministeriet": "Ministry for Senior Citizens", "Økonomiministeriet": "Ministry of Economic Affairs", "Ministeriet f. Samfundssikkerhed og Beredskab": "Ministry of Societal Security and Emergency Management", "Miljø- og Ligestillingsministeriet": "Ministry of Environment and Gender Equality", "Statsministeriet": "Prime Minister's Office", "Folketinget": "Danish Parliament", "By-, Land- og Kirkeministeriet": "Ministry of Urban, Rural and Church Affairs", "Kongen": "The King", "Digitaliseringsministeriet": "Ministry of Digital Government", "Medlemmer af det kongelige hus m.fl.": "Members of the Royal House and related expenditure",
    },
    "FRA": {
        "Enseignement scolaire": "School education", "Engagements financiers de l’État": "State financial commitments", "Défense": "Defence", "Recherche et enseignement supérieur": "Research and higher education", "Solidarité, insertion et égalité des chances": "Solidarity, inclusion and equal opportunities", "Cohésion des territoires": "Territorial cohesion", "Écologie, développement et mobilité durables": "Ecology, sustainable development and mobility", "Sécurités": "Security", "Travail, emploi et administration des ministères sociaux": "Labour, employment and social-ministry administration", "Justice": "Justice", "Gestion des finances publiques": "Public-finance management", "Régimes sociaux et de retraite": "Social and pension schemes", "Investir pour la France de 2030": "Investing for France 2030", "Administration générale et territoriale de l’État": "General and territorial state administration", "Relations avec les collectivités territoriales": "Relations with local authorities", "Agriculture, alimentation, forêt et affaires rurales": "Agriculture, food, forestry and rural affairs", "Aide publique au développement": "Official development assistance", "Culture": "Culture", "Action extérieure de l’État": "State external action", "Économie": "Economic affairs", "Outre-mer": "Overseas territories", "Immigration, asile et intégration": "Immigration, asylum and integration", "Monde combattant, mémoire et liens avec la nation": "Veterans, remembrance and links with the nation", "Santé": "Health", "Sport, jeunesse et vie associative": "Sport, youth and community life", "Pouvoirs publics": "Constitutional public authorities", "Direction de l’action du Gouvernement": "Direction of government action", "Conseil et contrôle de l’État": "State advisory and oversight bodies", "Médias, livre et industries culturelles": "Media, books and cultural industries", "Crédits non répartis": "Unallocated appropriations", "Transformation et fonction publiques": "Public-sector transformation and civil service",
    },
    "POL": {
        "Obowiązkowe ubezpieczenia społeczne": "Compulsory social insurance", "Różne rozliczenia": "Miscellaneous settlements", "Obrona narodowa": "National defence", "Rodzina": "Family", "Obsługa długu publicznego": "Public-debt service", "Ochrona zdrowia": "Health care", "Szkolnictwo wyższe i nauka": "Higher education and science", "Bezpieczeństwo publiczne i ochrona przeciwpożarowa": "Public safety and fire protection", "Administracja publiczna": "Public administration", "Wymiar sprawiedliwości": "Justice system", "Transport i łączność": "Transport and communications", "Rolnictwo i łowiectwo": "Agriculture and hunting", "Gospodarka mieszkaniowa": "Housing", "Pomoc społeczna": "Social assistance", "Górnictwo i kopalnictwo": "Mining and quarrying", "Kultura i ochrona dziedzictwa narodowego": "Culture and protection of national heritage", "Oświata i wychowanie": "School education", "Urzędy naczelnych organów władzy państwowej, kontroli i ochrony prawa oraz sądownictwa": "Supreme state authorities, audit and legal-protection bodies, and the judiciary", "Gospodarka komunalna i ochrona środowiska": "Municipal services and environmental protection", "Przetwórstwo przemysłowe": "Manufacturing", "Pozostałe zadania w zakresie polityki społecznej": "Other social-policy tasks", "Handel": "Trade", "Działalność usługowa": "Service activities", "Kultura fizyczna": "Sport and physical activity", "Informatyka": "Information technology", "Edukacyjna opieka wychowawcza": "Educational support and care", "Rybołówstwo i rybactwo": "Fisheries", "Ogrody botaniczne i zoologiczne oraz obszary chronione": "Botanical and zoological gardens and protected areas", "Turystyka": "Tourism", "Hotele i restauracje": "Hotels and restaurants", "Leśnictwo": "Forestry",
    },
    "CHE": {
        "Soziale Wohlfahrt": "Social welfare", "Finanzen und Steuern": "Finance and taxes", "Bildung und Forschung": "Education and research", "Verkehr": "Transport", "Übrige Aufgabengebiete": "Other task areas", "Sicherheit": "Security", "Beziehungen zum Ausland – Internationale Zusammenarbeit": "Foreign relations – international cooperation", "Landwirtschaft und Ernährung": "Agriculture and food",
    },
    "UKR": {
        "Оборона": "Defence", "Загальнодержавні функції": "General public services", "Громадський порядок, безпека та судова влада": "Public order, security and the judiciary", "Соціальний захист та соціальне забезпечення": "Social protection and social security", "Охорона здоров'я": "Health care", "Економічна діяльність": "Economic affairs", "Освіта": "Education", "Духовний та фізичний розвиток": "Culture, religion and physical development", "Охорона навколишнього природного середовища": "Environmental protection", "Житлово-комунальне господарство": "Housing and communal services",
    },
}
ENGLISH_SOURCE_COUNTRIES = {"GBR", "SWE", "USA"}

COFOG_LABELS = {
    "GF01": "General public services", "GF02": "Defence", "GF03": "Public order and safety",
    "GF04": "Economic affairs", "GF05": "Environmental protection", "GF06": "Housing and community amenities",
    "GF07": "Health", "GF08": "Recreation, culture and religion", "GF09": "Education", "GF10": "Social protection",
}
for _code in ("ESP", "NLD", "NOR"):
    TRANSLATIONS[_code] = dict(COFOG_LABELS)
TRANSLATIONS["BRA"] = {
    "Serviços públicos gerais": "General public services", "Defesa": "Defence",
    "Ordem pública e segurança": "Public order and safety", "Assuntos econômicos": "Economic affairs",
    "Proteção ambiental": "Environmental protection", "Habitação e serviços comunitários": "Housing and community amenities",
    "Saúde": "Health", "Recreação, cultura e religião": "Recreation, culture and religion",
    "Educação": "Education", "Proteção social": "Social protection",
}
TRANSLATIONS["JPN"] = {
    "皇室費":"Imperial Household expenses", "国会":"National Diet", "裁判所":"Courts", "会計検査院":"Board of Audit",
    "内閣":"Cabinet", "内閣府":"Cabinet Office", "デジタル庁":"Digital Agency", "防災庁":"Disaster Management Agency",
    "総務省":"Ministry of Internal Affairs and Communications", "法務省":"Ministry of Justice", "外務省":"Ministry of Foreign Affairs",
    "財務省":"Ministry of Finance", "文部科学省":"Ministry of Education, Culture, Sports, Science and Technology",
    "厚生労働省":"Ministry of Health, Labour and Welfare", "農林水産省":"Ministry of Agriculture, Forestry and Fisheries",
    "経済産業省":"Ministry of Economy, Trade and Industry", "国土交通省":"Ministry of Land, Infrastructure, Transport and Tourism",
    "環境省":"Ministry of the Environment", "防衛省":"Ministry of Defence",
}


def row(code, label, a, b):
    return {"code": str(code), "label_native": str(label).strip(), "amounts": {"previous": round(float(a), 4), "current": round(float(b), 4)}}


def country(code, currency, dimension, scope_cs, scope_en, previous, current, rows, sources, note_cs="", note_en=""):
    clean = [r for r in rows if r["amounts"]["previous"] != 0 or r["amounts"]["current"] != 0]
    for item in clean:
        item["label_en"] = item["label_native"] if code in ENGLISH_SOURCE_COUNTRIES else TRANSLATIONS.get(code, {}).get(item["label_native"])
        if not item["label_en"]:
            raise ValueError(f"Missing English budget-line translation: {code} / {item['label_native']}")
    return {
        "code": code, "currency": currency, "unit": "billion_local_currency",
        "dimension": dimension, "scope_cs": scope_cs, "scope_en": scope_en,
        "periods": {"previous": previous, "current": current},
        "rows": sorted(clean, key=lambda x: x["amounts"]["current"], reverse=True),
        "totals": {
            "previous": round(sum(x["amounts"]["previous"] for x in clean), 4),
            "current": round(sum(x["amounts"]["current"] for x in clean), 4),
        },
        "note_cs": note_cs, "note_en": note_en, "sources": sources,
    }


def czechia():
    data = json.loads((ROOT / "data" / "cz-spending-2026.v1.json").read_text())
    rows = [row(x["code"], x["label_cs"], x["amount_2025_czk"] / 1e9, x["amount_2026_czk"] / 1e9) for x in data["chapters"]]
    return country("CZE", "CZK", "administrative", "Schválené výdaje státního rozpočtu podle kapitol", "Approved state-budget expenditure by chapter",
        {"label":"2025", "status_cs":"schválený rozpočet", "status_en":"approved budget"},
        {"label":"2026", "status_cs":"schválený rozpočet", "status_en":"approved budget"}, rows, data["sources"],
        "Součet kapitol nezahrnuje samostatně vykázané prostředky EU a finančních mechanismů.",
        "The chapter total excludes separately reported EU and financial-mechanism funds.")


def germany():
    years = {}
    labels = {}
    for year in (2025, 2026):
        vals = defaultdict(float)
        with (RAW / "DEU" / f"federal-budget-{year}.csv").open(encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh, delimiter=";"):
                if r.get("einahmen-ausgaben") != "A": continue
                code = r["einzelplan"].strip(); labels[code] = r["einzelplan-text"].strip()
                raw = r.get("soll ", r.get("soll", "0"))
                vals[code] += float(raw.replace(".", "").replace(",", ".") or 0) / 1e6
        years[year] = vals
    rows = [row(k, labels[k], years[2025].get(k,0), years[2026].get(k,0)) for k in sorted(set(years[2025])|set(years[2026]))]
    return country("DEU","EUR","administrative","Federální rozpočet podle jednotlivých plánů (Einzelpläne)","Federal budget by departmental section (Einzelplan)",
        {"label":"2025","status_cs":"schválený plán","status_en":"approved plan"},{"label":"2026","status_cs":"schválený plán","status_en":"approved plan"},rows,
        [{"title":"Bundeshaushalt — Download-Portal","url":"https://www.bundeshaushalt.de/DE/Download-Portal/download-portal.html"}],
        "Položky příjmů jsou vyloučeny; zobrazeny jsou výdajové záznamy federálního rozpočtu.","Revenue records are excluded; the chart sums federal expenditure records only.")


def denmark():
    t = pd.read_html(RAW / "DNK" / "finance-act-report-2026.html")[-1]
    rows=[]
    for _,r in t.iterrows():
        label=str(r.iloc[0]); m=re.match(r"(\d{2})\s+(.+?)(?:\s+\(Anm\.\))?$",label)
        if not m or int(m.group(1)) >= 37: continue
        parse=lambda v: float(str(v).replace(".","").replace(",",".")) / 1000
        rows.append(row(m.group(1),m.group(2),parse(r.iloc[2]),parse(r.iloc[3])))
    return country("DNK","DKK","administrative","Finanční zákon podle ministerských paragrafů","Finance Act by ministerial section",
        {"label":"2025","status_cs":"rozpočet","status_en":"budget"},{"label":"2026","status_cs":"finanční zákon","status_en":"Finance Act"},rows,
        [{"title":"Finansministeriet — Finanslovsdatabasen","url":"https://fm.dk/arbejdsomraader/finanslov-og-offentlige-finanser/arbejdet-med-finansloven/finanslovsdatabasen/"}],
        "Částky jsou netto výdaje ministerských paragrafů; příjmové a finanční paragrafy 37–42 jsou vyloučeny.","Amounts are net expenditure by ministerial section; revenue and financing sections 37–42 are excluded.")


FRANCE = [
 ("Action extérieure de l’État",3.3,3.3),("Administration générale et territoriale de l’État",4.1,4.2),("Agriculture, alimentation, forêt et affaires rurales",4.0,3.8),("Aide publique au développement",4.4,3.7),("Cohésion des territoires",23.1,22.2),("Conseil et contrôle de l’État",.7,.7),("Crédits non répartis",.2,.5),("Culture",3.7,3.5),("Défense",50.5,57.1),("Direction de l’action du Gouvernement",1,1),("Écologie, développement et mobilité durables",17.8,20.9),("Économie",3.5,3.3),("Engagements financiers de l’État",55.9,60.4),("Enseignement scolaire",64.3,64.5),("Gestion des finances publiques",8.1,8.2),("Immigration, asile et intégration",2.1,2.2),("Investir pour la France de 2030",5.5,5.5),("Justice",10.5,10.6),("Médias, livre et industries culturelles",.7,.7),("Monde combattant, mémoire et liens avec la nation",1.9,1.7),("Outre-mer",2.9,2.8),("Pouvoirs publics",1.1,1.1),("Recherche et enseignement supérieur",30.6,31.3),("Régimes sociaux et de retraite",6,6),("Relations avec les collectivités territoriales",4,3.9),("Santé",1.5,1.7),("Sécurités",17.3,17.7),("Solidarité, insertion et égalité des chances",30.3,29.5),("Sport, jeunesse et vie associative",1.5,1.2),("Transformation et fonction publiques",.7,.5),("Travail, emploi et administration des ministères sociaux",19.9,17.4)]


def france():
    return country("FRA","EUR","functional","Mise rozpočtu státu; bez přímých příspěvků do důchodového účtu CAS","State-budget missions; excluding direct contributions to the CAS pensions account",
        {"label":"2025","status_cs":"LFI 2025","status_en":"2025 initial Finance Act"},{"label":"2026","status_cs":"PLF 2026 ve srovnávací tabulce","status_en":"2026 draft in comparison table"},[row(i+1,*x) for i,x in enumerate(FRANCE)],
        [{"title":"Direction du Budget — chiffres clés PLF 2026","url":"https://www.budget.gouv.fr/documentation/file-download/30562"},{"title":"Direction du Budget — LFI 2026","url":"https://www.budget.gouv.fr/reperes/loi_de_finances/articles/chiffres-cles-budget-etat-2026"}],
        "Dvouletá mise-po-misi tabulka používá společný perimeter publikace PLF 2026; karta zdrojů odkazuje i na pozdější schválený LFI 2026.","The two-period mission table uses the common perimeter published with the 2026 draft; sources also link the later enacted 2026 budget.")


def britain():
    ws=load_workbook(RAW/"GBR"/"pesa-2026-chapter-1.xlsx",read_only=True,data_only=True)["Table_1_9"]
    rows=[]
    for r in ws.iter_rows(min_row=7,max_row=30,values_only=True):
        if r[0] and isinstance(r[5],(int,float)) and isinstance(r[6],(int,float)): rows.append(row(len(rows)+1,r[0],r[5]/1000,r[6]/1000))
    return country("GBR","GBP","administrative","Celkové rezortní výdajové limity (resource + capital DEL)","Total departmental expenditure limits (resource + capital DEL)",
        {"label":"2025–26","status_cs":"skutečnost","status_en":"outturn"},{"label":"2026–27","status_cs":"plán","status_en":"plan"},rows,
        [{"title":"HM Treasury — Public Expenditure Statistical Analyses 2026","url":"https://www.gov.uk/government/statistics/public-expenditure-statistical-analyses-2026"}],
        "Jde o kontrolované rezortní limity DEL, nikoli o celé britské veřejné výdaje; mandatorní AME je mimo graf.","This is controlled departmental DEL, not total UK public spending; demand-led AME is outside the chart.")


POL26 = [("010","Rolnictwo i łowiectwo",9575450),("020","Leśnictwo",7658),("050","Rybołówstwo i rybactwo",243330),("100","Górnictwo i kopalnictwo",5634233),("150","Przetwórstwo przemysłowe",2224221),("500","Handel",1324009),("550","Hotele i restauracje",70718),("600","Transport i łączność",21820688),("630","Turystyka",131585),("700","Gospodarka mieszkaniowa",7033790),("710","Działalność usługowa",1029995),("720","Informatyka",298326),("730","Szkolnictwo wyższe i nauka",41241437),("750","Administracja publiczna",30207860),("751","Urzędy naczelnych organów władzy państwowej, kontroli i ochrony prawa oraz sądownictwa",4690337),("752","Obrona narodowa",109231480),("753","Obowiązkowe ubezpieczenia społeczne",195970695),("754","Bezpieczeństwo publiczne i ochrona przeciwpożarowa",31139276),("755","Wymiar sprawiedliwości",28682245),("757","Obsługa długu publicznego",90388313),("758","Różne rozliczenia",162527526),("801","Oświata i wychowanie",4700145),("851","Ochrona zdrowia",54242749),("852","Pomoc społeczna",6911630),("853","Pozostałe zadania w zakresie polityki społecznej",2213500),("854","Edukacyjna opieka wychowawcza",290956),("855","Rodzina",97921791),("900","Gospodarka komunalna i ochrona środowiska",2692013),("921","Kultura i ochrona dziedzictwa narodowego",5488396),("925","Ogrody botaniczne i zoologiczne oraz obszary chronione",232531),("926","Kultura fizyczna",773117)]


def poland():
    ws=load_workbook(RAW/"POL"/"state-budget-execution-2025.xlsx",read_only=True,data_only=True)["TABLICA 7"]
    old={}; labels={}
    for r in ws.iter_rows(min_row=12,values_only=True):
        if r[0] and r[3]=="a" and isinstance(r[4],(int,float)): old[str(r[0]).zfill(3)]=r[4]/1e9; labels[str(r[0]).zfill(3)]=str(r[2]).strip()
    new={c:v/1e6 for c,_,v in POL26}; labels.update({c:l for c,l,_ in POL26})
    rows=[row(c,labels[c],old.get(c,0),new.get(c,0)) for c in sorted(set(old)|set(new))]
    return country("POL","PLN","functional","Státní rozpočet podle funkčních oddílů (działy)","State budget by functional division (działy)",
        {"label":"2025","status_cs":"schválený zákon","status_en":"approved act"},{"label":"2026","status_cs":"schválený zákon","status_en":"approved act"},rows,
        [{"title":"Ministerstwo Finansów — wykonanie budżetu 2025","url":"https://www.gov.pl/web/finanse/sprawozdania-miesieczne-2025"},{"title":"Ministerstwo Finansów — ustawa budżetowa 2026","url":"https://www.gov.pl/web/finanse/ustawa-2026"}])


SWEDEN = [("Governance",20,22),("Economy and financial administration",22,22),("Taxes, customs and enforcement",14,16),("Justice",86,94),("International cooperation",3,3),("Defence and contingency measures",164,221),("International development cooperation",45,46),("Migration",10,13),("Health care, medical care and social services",116,127),("Financial security for those with illnesses and disabilities",120,122),("Financial security for the elderly",62,60),("Financial security for families and children",101,102),("Integration and gender equality",6,6),("Labour market and working life",91,90),("Financial support for students",31,31),("Education and academic research",102,105),("Culture, media, religious communities and leisure",17,17),("Planning, housing, construction and consumer policy",3,2),("Regional development",4,5),("Climate, environment and nature",13,18),("Energy",5,10),("Transport and communications",89,104),("Land- and water-based industries, rural areas and food",21,22),("Industry and trade",8,9),("General grants to local government",173,181),("Interest on central government debt",30,26),("Contribution to the European Union",46,56)]


def sweden():
    return country("SWE","SEK","functional","27 výdajových oblastí centrálního státního rozpočtu","27 central-government expenditure areas",
        {"label":"2025","status_cs":"skutečnost","status_en":"outcome"},{"label":"2026","status_cs":"prognóza","status_en":"forecast"},[row(i+1,*x) for i,x in enumerate(SWEDEN)],
        [{"title":"Government of Sweden — Central government budget in figures","url":"https://www.government.se/government-of-sweden/ministry-of-finance/central-government-budget/central-government-budget-in-figures/"}],
        "Částky jsou zaokrouhlené na miliardy SEK; součet se může lišit vlivem zaokrouhlení.","Amounts are rounded to SEK billions; components may not sum exactly.")


SWISS = [("Soziale Wohlfahrt",30106,31612),("Finanzen und Steuern",13903,15038),("Verkehr",8566,8699),("Bildung und Forschung",8039,8704),("Sicherheit",5121,5372),("Landwirtschaft und Ernährung",3588,3597),("Beziehungen zum Ausland – Internationale Zusammenarbeit",3614,3693),("Übrige Aufgabengebiete",7158,7577)]


def switzerland():
    return country("CHE","CHF","functional","Běžné výdaje Konfederace podle oblastí úkolů","Confederation current expenditure by task area",
        {"label":"2025","status_cs":"rozpočet","status_en":"budget"},{"label":"2026","status_cs":"rozpočet","status_en":"budget"},[row(i+1,n,a/1000,b/1000) for i,(n,a,b) in enumerate(SWISS)],
        [{"title":"Eidgenössische Finanzverwaltung — Voranschlag 2026, Band 1","url":"https://www.efv.admin.ch/dam/en/sd-web/npeosCYca2do/Voranschlag%202026%20mit%20IAFP%202027%E2%80%932029%20%E2%80%93%20Band%201%20DE.pdf"}],
        "Graf pokrývá běžné výdaje; investiční účet je v oficiálním rozpočtu vykázán odděleně.","The chart covers current expenditure; the official budget reports the investment account separately.")


UA_LABELS={"01":"Загальнодержавні функції","02":"Оборона","03":"Громадський порядок, безпека та судова влада","04":"Економічна діяльність","05":"Охорона навколишнього природного середовища","06":"Житлово-комунальне господарство","07":"Охорона здоров'я","08":"Духовний та фізичний розвиток","09":"Освіта","10":"Соціальний захист та соціальне забезпечення"}


def ukraine():
    years={}
    for year in (2025,2026):
        with tempfile.TemporaryDirectory() as d:
            with ZipFile(RAW/"UKR"/f"state-budget-law-{year}.zip") as z:z.extractall(d)
            x=next(Path(d).glob("*.xlsx")); ws=load_workbook(x,read_only=True,data_only=True)[load_workbook(x,read_only=True).sheetnames[2]]
            vals=defaultdict(float)
            for r in ws.iter_rows(min_row=8,values_only=True):
                p=str(r[1] or "").strip(); f=str(r[2] or "").strip(); amount=r[14]
                if len(p)==7 and not p.endswith("000") and len(f)==4 and f.isdigit() and isinstance(amount,(int,float)): vals[f[:2]] += amount/1e6
            years[year]=vals
    rows=[row(c,UA_LABELS[c],years[2025].get(c,0),years[2026].get(c,0)) for c in UA_LABELS]
    return country("UKR","UAH","functional","Výdaje státního rozpočtu podle funkční klasifikace","State-budget expenditure by functional classification",
        {"label":"2025","status_cs":"zákon po změnách","status_en":"amended budget act"},{"label":"2026","status_cs":"zákon po změnách","status_en":"amended budget act"},rows,
        [{"title":"Verkhovna Rada — State Budget Act 2025","url":"https://zakon.rada.gov.ua/laws/show/4059-20"},{"title":"Verkhovna Rada — State Budget Act 2026","url":"https://zakon.rada.gov.ua/laws/show/4695-20"},{"title":"Open Budget — functional expenditure","url":"https://openbudget.gov.ua/en/national-budget/expenses?class=functional&view=table"}],
        "Částky agregujeme z programových řádků přílohy 3 podle první úrovně funkční klasifikace.","Amounts are aggregated from Appendix 3 programme rows to the first functional-classification level.")


def usa():
    ws=load_workbook(RAW/"USA"/"public-budget-database-outlays-fy2027.xlsx",read_only=True,data_only=True).active
    hdr=[x.value for x in next(ws.iter_rows(min_row=1,max_row=1))]; i25=hdr.index("2025"); i26=hdr.index("2026")
    vals=defaultdict(lambda:[0.0,0.0])
    for r in ws.iter_rows(min_row=2,values_only=True):
        name=r[1]
        if not name: continue
        vals[str(name)][0]+=float(r[i25] or 0)/1000; vals[str(name)][1]+=float(r[i26] or 0)/1000
    rows=[row(i+1,n,*v) for i,(n,v) in enumerate(vals.items())]
    return country("USA","USD","administrative","Federální výdaje (outlays) podle agentury, včetně netto kompenzačních příjmů","Federal outlays by agency, including net offsetting receipts",
        {"label":"FY 2025","status_cs":"skutečnost","status_en":"actual"},{"label":"FY 2026","status_cs":"odhad","status_en":"estimate"},rows,
        [{"title":"OMB — Public Budget Database, Outlays","url":"https://www.whitehouse.gov/omb/information-resources/budget/supplemental-materials/"}],
        "Záporné částky jsou zachovány: jde o netto kompenzační příjmy a úvěrové operace v oficiální databázi OMB.","Negative values are retained: they are net offsetting receipts and credit operations in the official OMB database.")


EUROSTAT_NATIVE = {
    "ESP": ["Servicios públicos generales", "Defensa", "Orden público y seguridad", "Asuntos económicos", "Protección del medio ambiente", "Vivienda y servicios comunitarios", "Salud", "Actividades recreativas, cultura y religión", "Educación", "Protección social"],
    "NLD": ["Algemeen overheidsbestuur", "Defensie", "Openbare orde en veiligheid", "Economische aangelegenheden", "Milieubescherming", "Huisvesting en gemeenschapsvoorzieningen", "Volksgezondheid", "Recreatie, cultuur en religie", "Onderwijs", "Sociale bescherming"],
    "NOR": ["Alminnelig offentlig tjenesteyting", "Forsvar", "Offentlig orden og trygghet", "Næringsøkonomiske formål", "Miljøvern", "Boliger og nærmiljø", "Helse", "Fritid, kultur og religion", "Utdanning", "Sosial beskyttelse"],
}
for _country_code, _labels in EUROSTAT_NATIVE.items():
    TRANSLATIONS[_country_code] = {_labels[i]: COFOG_LABELS[f"GF{i + 1:02d}"] for i in range(10)}


def _jsonstat_value(payload, wanted):
    """Read one value from Eurostat's compact JSON-stat response."""
    flat = 0
    stride = 1
    for dimension_id, size in reversed(list(zip(payload["id"], payload["size"]))):
        index = payload["dimension"][dimension_id]["category"]["index"]
        position = index[wanted[dimension_id]] if isinstance(index, dict) else index.index(wanted[dimension_id])
        flat += position * stride
        stride *= size
    values = payload["value"]
    value = values.get(str(flat)) if isinstance(values, dict) else values[flat]
    return float(value)


def eurostat_cofog(code, geo):
    params = [("freq", "A"), ("unit", "MIO_NAC"), ("sector", "S13"), ("na_item", "TE"), ("geo", geo), ("time", "2023"), ("time", "2024")]
    url = "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/gov_10a_exp?" + urllib.parse.urlencode(params)
    with urllib.request.urlopen(url, timeout=90) as response:
        payload = json.load(response)
    rows = []
    for i, cofog in enumerate(COFOG_LABELS, 1):
        values = []
        for year in ("2023", "2024"):
            wanted = {dimension_id: next(iter(payload["dimension"][dimension_id]["category"]["index"])) for dimension_id in payload["id"]}
            wanted.update({"freq":"A", "unit":"MIO_NAC", "sector":"S13", "na_item":"TE", "geo":geo, "cofog":cofog, "time":year})
            values.append(_jsonstat_value(payload, wanted) / 1000)
        rows.append(row(cofog, EUROSTAT_NATIVE[code][i - 1], *values))
    currency = {"ESP":"EUR", "NLD":"EUR", "NOR":"NOK"}[code]
    return country(code, currency, "functional", "Výdaje vládních institucí podle COFOG", "General-government expenditure by COFOG",
        {"label":"2023", "status_cs":"skutečnost", "status_en":"actual"}, {"label":"2024", "status_cs":"skutečnost", "status_en":"actual"}, rows,
        [{"title":"Eurostat — General government expenditure by function (gov_10a_exp)", "url":"https://ec.europa.eu/eurostat/databrowser/view/gov_10a_exp/default/table"}],
        "Částky pokrývají celý sektor vládních institucí S.13 a jsou v milionech národní měny převedených na miliardy.",
        "Amounts cover the full S.13 general-government sector and are converted from millions to billions of national currency.")


def brazil():
    labels = ["Serviços públicos gerais", "Defesa", "Ordem pública e segurança", "Assuntos econômicos", "Proteção ambiental", "Habitação e serviços comunitários", "Saúde", "Recreação, cultura e religião", "Educação", "Proteção social"]
    values = [(1242018,1301580),(55791,59397),(311371,346997),(241967,265604),(54293,65313),(126088,159650),(516411,590982),(37092,47214),(539973,599178),(1833347,1941604)]
    rows = [row(f"70{i}", labels[i - 1], a / 1000, b / 1000) for i, (a, b) in enumerate(values, 1)]
    return country("BRA", "BRL", "functional", "Výdaje vládních institucí podle COFOG", "General-government expenditure by COFOG",
        {"label":"2023", "status_cs":"skutečnost", "status_en":"actual"}, {"label":"2024", "status_cs":"skutečnost", "status_en":"actual"}, rows,
        [{"title":"Tesouro Nacional / SOF / IBGE / Banco Central — COFOG 2024 bulletin", "url":"https://www.gov.br/planejamento/pt-br/assuntos/orcamento/publicaoes-sobre-orcamento/classificacao-das-funcoes-de-governo-cofog/arquivos/boletim_cofog_2024.pdf/@@display-file/file"}],
        "Oficiální konsolidované výdaje všech vládních institucí; původní tabulka je v milionech BRL.", "Official consolidated expenditure of general government; the source table is in BRL millions.")


def japan():
    labels = ["皇室費","国会","裁判所","会計検査院","内閣","内閣府","デジタル庁","防災庁","総務省","法務省","外務省","財務省","文部科学省","厚生労働省","農林水産省","経済産業省","国土交通省","環境省","防衛省"]
    values = [(13430532,13648753),(136723343,145773232),(361086867,387221498),(17350449,18106659),(165299563,175400615),(5573103413,5534898968),(233910824,294543189),(0,2414573),(20266185387,22703714567),(885963789,944653365),(886514990,889281550),(29934682990,33313616029),(6848151456,7720525067),(35054552982,36528554845),(3274647710,3372806072),(1655514359,2005220872),(9926649047,9952532626),(533758685,565810507),(9673584875,9919530135)]
    rows = [row(f"{i:02d}", label, a / 1e6, b / 1e6) for i, (label, (a, b)) in enumerate(zip(labels, values), 1)]
    return country("JPN", "JPY", "administrative", "Výdajový rozpočet podle ministerstva a ústavní instituce", "Expenditure budget by ministry and constitutional institution",
        {"label":"FY 2025", "status_cs":"běžný rozpočet", "status_en":"current budget"}, {"label":"FY 2026", "status_cs":"běžný rozpočet", "status_en":"current budget"}, rows,
        [{"title":"Ministry of Finance Japan — FY2025 revenue and expenditure budget", "url":"https://www.mof.go.jp/policy/budget/report/revenue_and_expenditure/fy2025/0705b.html"},{"title":"Ministry of Finance Japan — FY2026 revenue and expenditure budget", "url":"https://www.mof.go.jp/policy/budget/report/revenue_and_expenditure/fy2026/0805b.html"}],
        "Oficiální částky jsou v tisících JPY; zde jsou převedeny na miliardy. Jde o rozpočtová oprávnění, nikoli průběžné čerpání.", "Official amounts are in JPY thousands and are converted to billions. These are budget authorities, not year-to-date spending.")


def main():
    countries=[czechia(),germany(),denmark(),france(),britain(),poland(),sweden(),switzerland(),ukraine(),usa(),brazil(),eurostat_cofog("ESP","ES"),japan(),eurostat_cofog("NLD","NL"),eurostat_cofog("NOR","NO")]
    payload={
      "schema_version":"1.0.0","generated_at":"2026-08-26",
      "methodology_cs":"Každý profil srovnává dvě období ve stejném národním členění. Rozsahy mezi státy nejsou účetně totožné.",
      "methodology_en":"Each profile compares two periods under one national classification. Accounting perimeters are not identical across countries.",
      "fx":{"reference_date":"2026-08-18","local_per_eur":{"EUR":1,"CZK":24.179,"DKK":7.4759,"GBP":0.85585,"PLN":4.319,"SEK":11.045,"CHF":0.9406,"USD":1.1576,"UAH":51.8082,"BRL":6.0281,"JPY":184.87,"NOK":10.9025},"sources":[{"title":"ECB euro reference rates","url":"https://www.ecb.europa.eu/stats/policy_and_exchange_rates/euro_reference_exchange_rates/html/index.en.html"},{"title":"National Bank of Ukraine EUR/UAH official rate","url":"https://bank.gov.ua/NBUStatService/v1/statdirectory/exchange?valcode=EUR&date=20260818&json"}]},
      "countries":countries,
    }
    OUT.write_text(json.dumps(payload,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(f"wrote {OUT} ({len(countries)} countries, {sum(len(c['rows']) for c in countries)} rows)")


if __name__ == "__main__": main()
