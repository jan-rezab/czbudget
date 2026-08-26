#!/usr/bin/env python3
"""Download and normalize the portal's long-run econometric panel.

Raw, dated source snapshots live outside the public web root in
../data/sources/economy. Compact report-ready files are written to data/economy.
The normalized fact grain is one source observation for one country and period;
no interpolation, forecast, seasonal conversion, or unit conversion is applied.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import io
import json
import shutil
import sys
import urllib.request
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = ROOT.parent
CONTRACT_PATH = ROOT / "data/contracts/economic-series.v1.json"
OUTPUT_DIR = ROOT / "data/economy"
RAW_ROOT = WORKSPACE / "data/sources/economy"
FACT_FIELDS = [
    "country_code", "indicator_code", "source_series", "source_key", "topic",
    "frequency", "period", "observation_date", "value", "unit",
    "seasonal_adjustment", "transformation", "observation_status", "source_id",
    "source_url", "source_vintage", "retrieved_at", "quality_flags",
]

ISO3_TO_ISO2 = {
    "BRA": "BR", "CHE": "CH", "CZE": "CZ", "DEU": "DE", "DNK": "DK",
    "ESP": "ES", "FIN": "FI", "FRA": "FR", "GBR": "GB", "GRC": "GR",
    "JPN": "JP", "NLD": "NL", "NOR": "NO", "POL": "PL", "SWE": "SE",
    "UKR": "UA", "USA": "US",
}
ISO2_TO_ISO3 = {value: key for key, value in ISO3_TO_ISO2.items()}

COUNTRY_NAMES = {
    "BRA": ("Brazílie", "Brazil"), "CHE": ("Švýcarsko", "Switzerland"),
    "CZE": ("Česko", "Czechia"), "DEU": ("Německo", "Germany"),
    "DNK": ("Dánsko", "Denmark"), "ESP": ("Španělsko", "Spain"),
    "FIN": ("Finsko", "Finland"), "FRA": ("Francie", "France"),
    "GBR": ("Spojené království", "United Kingdom"), "GRC": ("Řecko", "Greece"),
    "JPN": ("Japonsko", "Japan"), "NLD": ("Nizozemsko", "Netherlands"),
    "NOR": ("Norsko", "Norway"), "POL": ("Polsko", "Poland"),
    "SWE": ("Švédsko", "Sweden"), "UKR": ("Ukrajina", "Ukraine"),
    "USA": ("Spojené státy", "United States"),
}

IMF_LABELS = {
    "balance_pct_gdp": ("Saldo vládních institucí", "General-government balance", "fiscal"),
    "expenditure_pct_gdp": ("Výdaje vládních institucí", "General-government expenditure", "fiscal"),
    "gross_debt_pct_gdp": ("Hrubý vládní dluh", "Gross government debt", "fiscal"),
    "primary_balance_pct_gdp": ("Primární saldo", "Primary balance", "fiscal"),
    "revenue_pct_gdp": ("Příjmy vládních institucí", "General-government revenue", "fiscal"),
    "structural_balance_pct_potential_gdp": ("Strukturální saldo", "Structural balance", "fiscal"),
    "real_gdp_growth_pct": ("Růst reálného HDP", "Real GDP growth", "output"),
    "inflation_pct": ("Inflace", "Inflation", "prices_finance"),
    "unemployment_pct": ("Míra nezaměstnanosti", "Unemployment rate", "labour"),
}

IMF_METRICS = {
    "GGXCNL_NGDP": ("balance_pct_gdp", "pct_gdp"),
    "GGX_NGDP": ("expenditure_pct_gdp", "pct_gdp"),
    "NGDPPC": ("gdp_per_capita_local", "local_currency_per_capita"),
    "PPPPC": ("gdp_per_capita_ppp", "international_dollar_per_capita"),
    "NGDPDPC": ("gdp_per_capita_usd", "usd_per_capita"),
    "GGXWDG_NGDP": ("gross_debt_pct_gdp", "pct_gdp"),
    "PCPIPCH": ("inflation_pct", "percent"),
    "NGDP": ("nominal_gdp_local_bn", "local_currency_bn"),
    "NGDPD": ("nominal_gdp_usd_bn", "usd_bn"),
    "PPPEX": ("ppp_conversion_factor", "local_currency_per_international_dollar"),
    "GGXONLB_NGDP": ("primary_balance_pct_gdp", "pct_gdp"),
    "NGDP_RPCH": ("real_gdp_growth_pct", "percent"),
    "GGR_NGDP": ("revenue_pct_gdp", "pct_gdp"),
    "GGSB_NPGDP": ("structural_balance_pct_potential_gdp", "pct_potential_gdp"),
    "LUR": ("unemployment_pct", "percent"),
}


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def download(url: str, target: Path, accept: str | None = None, refresh: bool = False) -> None:
    if target.exists() and not refresh:
        return
    target.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(url, headers={"User-Agent": "PublicSpendingData/1.0"})
    if accept:
        request.add_header("Accept", accept)
    temporary = target.with_suffix(target.suffix + ".part")
    with urllib.request.urlopen(request, timeout=180) as response, temporary.open("wb") as output:
        shutil.copyfileobj(response, output)
    temporary.replace(target)


def fetch_country_dimension(raw_dir: Path, refresh: bool) -> list[dict]:
    url = "https://api.worldbank.org/v2/country?format=json&per_page=400"
    target = raw_dir / "world_bank" / "countries.json"
    download(url, target, refresh=refresh)
    payload = json.loads(target.read_text(encoding="utf-8"))
    countries = []
    for record in payload[1]:
        iso3 = record.get("id", "")
        iso2 = record.get("iso2Code", "")
        if record.get("region", {}).get("id") == "NA" or len(iso3) != 3 or len(iso2) != 2:
            continue
        countries.append({
            "code": iso3, "iso2": iso2, "name_en": record.get("name", iso3),
            "name_cs": COUNTRY_NAMES.get(iso3, (record.get("name", iso3), ""))[0],
            "region": record.get("region", {}).get("value", ""),
            "income_level": record.get("incomeLevel", {}).get("value", ""),
        })
        ISO3_TO_ISO2[iso3] = iso2
        ISO2_TO_ISO3[iso2] = iso3
    return sorted(countries, key=lambda item: item["code"])


def numeric(value) -> float | None:
    if value in (None, "", "NA", "NaN"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def period_date(period: str, frequency: str) -> str:
    if frequency == "A" and len(period) >= 4:
        return f"{period[:4]}-12-31"
    if frequency == "Q" and "Q" in period:
        year, quarter = period.replace("-", "").split("Q", 1)
        month_day = {"1": "03-31", "2": "06-30", "3": "09-30", "4": "12-31"}.get(quarter[-1], "12-31")
        return f"{year[:4]}-{month_day}"
    if frequency == "M" and len(period) >= 7:
        return f"{period[:7]}-01"
    return ""


def source_code(value: str) -> str:
    return (value or "").split(":", 1)[0].strip()


def make_row(**values) -> dict[str, str]:
    row = {field: "" for field in FACT_FIELDS}
    row.update({key: str(value) for key, value in values.items() if value is not None})
    return row


def fetch_world_bank(contract: dict, raw_dir: Path, retrieved_at: str, refresh: bool) -> tuple[list[dict], list[dict]]:
    countries = "all" if contract["scope"].get("download_all_countries") else ";".join(contract["scope"]["countries"])
    allowed = set(contract["scope"]["countries"])
    start = contract["scope"]["annual_start"]
    end = datetime.now().year - 1
    observations: list[dict] = []
    raw_bundle = {"retrieved_at": retrieved_at, "series": {}}
    definitions = []
    for definition in contract["world_bank_indicators"]:
        code = definition["source_code"]
        url = f"https://api.worldbank.org/v2/country/{countries}/indicator/{code}?date={start}:{end}&format=json&per_page=20000"
        target = raw_dir / "world_bank" / f"{code}.json"
        download(url, target, refresh=refresh)
        payload = json.loads(target.read_text(encoding="utf-8"))
        metadata = payload[0] if isinstance(payload, list) and payload else {}
        records = payload[1] if isinstance(payload, list) and len(payload) > 1 and payload[1] else []
        raw_bundle["series"][code] = {"url": url, "metadata": metadata, "file": str(target.relative_to(raw_dir))}
        definitions.append({**definition, "source_id": "world_bank", "frequencies": ["A"]})
        for record in records:
            value = numeric(record.get("value"))
            country = record.get("countryiso3code")
            if value is None or country not in allowed:
                continue
            period = str(record["date"])
            observations.append(make_row(
                country_code=country, indicator_code=definition["indicator_code"], source_series=code,
                source_key=code, topic=definition["topic"], frequency="A", period=period,
                observation_date=period_date(period, "A"), value=format(value, ".15g"), unit=definition["unit"],
                observation_status=record.get("obs_status", ""), source_id="world_bank", source_url=url,
                source_vintage=metadata.get("lastupdated", ""), retrieved_at=retrieved_at,
            ))
    summary = raw_dir / "world_bank-manifest.json"
    summary.write_text(json.dumps(raw_bundle, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return observations, definitions


def fetch_oecd(contract: dict, raw_dir: Path, retrieved_at: str, refresh: bool) -> tuple[list[dict], list[dict]]:
    countries = "." if contract["scope"].get("download_all_countries") else "+".join(contract["scope"]["countries"])
    allowed = set(contract["scope"]["countries"])
    start = contract["scope"]["high_frequency_start"]
    end = datetime.now().year
    key = countries if countries == "." else f"{countries}......"
    url = f"{contract['sources']['oecd_kei']['api']}/{key}?startPeriod={start}-01&endPeriod={end}-12"
    raw_csv = raw_dir / "oecd-kei.csv"
    gzip_target = raw_dir / "oecd-kei.csv.gz"
    if not refresh and not raw_csv.exists() and gzip_target.exists():
        with gzip.open(gzip_target, "rb") as source, raw_csv.open("wb") as target:
            shutil.copyfileobj(source, target)
    else:
        download(url, raw_csv, accept="text/csv", refresh=refresh)
    observations: list[dict] = []
    definitions = []
    measures = contract["oecd_measure_map"]
    for code, details in measures.items():
        definitions.append({
            "indicator_code": f"oecd_kei_{code.lower()}", "source_code": code,
            "label_en": details["label_en"], "label_cs": details["label_cs"],
            "topic": details["topic"], "unit": "source_defined", "source_id": "oecd_kei",
            "frequencies": ["A", "Q", "M"],
        })
    with raw_csv.open(newline="", encoding="utf-8-sig") as handle:
        for record in csv.DictReader(handle):
            value = numeric(record.get("OBS_VALUE"))
            measure = record.get("MEASURE", "")
            country = record.get("REF_AREA", "")
            if value is None or measure not in measures or country not in allowed:
                continue
            frequency = record.get("FREQ", "")
            period = record.get("TIME_PERIOD", "")
            dimensions = [measure, record.get("UNIT_MEASURE", ""), record.get("ACTIVITY", ""), record.get("ADJUSTMENT", ""), record.get("TRANSFORMATION", "")]
            observations.append(make_row(
                country_code=country, indicator_code=f"oecd_kei_{measure.lower()}",
                source_series=".".join(dimensions), source_key="|".join([country, frequency, *dimensions]),
                topic=measures[measure]["topic"], frequency=frequency, period=period,
                observation_date=period_date(period, frequency), value=format(value, ".15g"),
                unit=record.get("UNIT_MEASURE", ""), seasonal_adjustment=record.get("ADJUSTMENT", ""),
                transformation=record.get("TRANSFORMATION", ""), observation_status=record.get("OBS_STATUS", ""),
                source_id="oecd_kei", source_url=url, source_vintage="4.0", retrieved_at=retrieved_at,
                quality_flags="" if country != "UKR" else "limited_country_coverage",
            ))
    if refresh or not gzip_target.exists():
        with raw_csv.open("rb") as source, gzip.open(gzip_target, "wb", compresslevel=9) as target:
            shutil.copyfileobj(source, target)
    raw_csv.unlink()
    return observations, definitions


def fetch_bis(contract: dict, raw_dir: Path, retrieved_at: str, refresh: bool) -> tuple[list[dict], list[dict]]:
    observations: list[dict] = []
    definitions = []
    allowed = set(contract["scope"]["countries"])
    for dataset in contract["bis_datasets"]:
        code = dataset["dataset"]
        target = raw_dir / "bis" / f"{code}_csv_flat.zip"
        download(dataset["url"], target, refresh=refresh)
        definitions.append({
            "indicator_code": dataset["indicator_code"], "source_code": code,
            "label_en": dataset["label_en"], "label_cs": dataset["label_cs"],
            "topic": dataset["topic"], "unit": "source_defined", "source_id": "bis",
            "frequencies": ["A", "Q", "M"],
        })
        with zipfile.ZipFile(target) as archive:
            csv_name = next(name for name in archive.namelist() if name.endswith(".csv"))
            with archive.open(csv_name) as binary:
                reader = csv.DictReader(io.TextIOWrapper(binary, encoding="utf-8-sig", newline=""))
                for record in reader:
                    country_field = "BORROWERS_CTY:Borrowers' country" if code in {"WS_TC", "WS_DSR"} else "REF_AREA:Reference area"
                    country = ISO2_TO_ISO3.get(source_code(record.get(country_field, "")))
                    value = numeric(record.get("OBS_VALUE:Observation Value"))
                    period = record.get("TIME_PERIOD:Time period or range", "")
                    frequency = source_code(record.get("FREQ:Frequency", ""))
                    if country not in allowed or value is None or not period or frequency not in {"A", "Q", "M"}:
                        continue
                    # BIS places true series dimensions before TIME_PERIOD and
                    # observation attributes after OBS_VALUE. Status, decimals
                    # and confidentiality must not fragment one time series.
                    field_order = list(record)
                    time_index = field_order.index("TIME_PERIOD:Time period or range")
                    dimension_fields = [
                        key for key in field_order[3:time_index]
                        if key != country_field and record.get(key, "")
                    ]
                    series_parts = [code] + [f"{key.split(':', 1)[0]}={source_code(record[key])}" for key in dimension_fields]
                    unit = record.get("UNIT_MEASURE:Unit of measure", "")
                    observations.append(make_row(
                        country_code=country, indicator_code=dataset["indicator_code"],
                        source_series="|".join(series_parts), source_key="|".join([country, *series_parts]),
                        topic=dataset["topic"], frequency=frequency, period=period,
                        observation_date=period_date(period, frequency), value=format(value, ".15g"), unit=unit,
                        observation_status=source_code(record.get("OBS_STATUS:Observation Status", "")),
                        source_id="bis", source_url=dataset["url"], source_vintage="bulk_download",
                        retrieved_at=retrieved_at,
                    ))
    return observations, definitions


def load_imf(contract: dict, retrieved_at: str) -> tuple[list[dict], list[dict]]:
    from openpyxl import load_workbook

    source_path = WORKSPACE / "data/sources/international_fiscal/WEOApr2026all.xlsx"
    url = contract["sources"]["imf_weo"]["homepage"]
    observations: list[dict] = []
    definitions_by_code = {}
    allowed = set(contract["scope"]["countries"])
    sheet = load_workbook(source_path, read_only=True, data_only=True)["Countries"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    years = [(index, int(value)) for index, value in enumerate(headers) if isinstance(value, int)]
    for values in rows:
        record = dict(zip(headers[:29], values[:29]))
        source_indicator = record.get("INDICATOR.ID")
        country = record.get("COUNTRY.ID")
        if source_indicator not in IMF_METRICS or country not in allowed:
            continue
        code, unit = IMF_METRICS[source_indicator]
        label_cs, label_en, topic = IMF_LABELS.get(code, (code, code, "macro"))
        definitions_by_code[code] = {
            "indicator_code": code, "source_code": source_indicator,
            "label_en": label_en, "label_cs": label_cs, "topic": topic,
            "unit": unit, "source_id": "imf_weo", "frequencies": ["A"],
        }
        latest_actual = numeric(record.get("LATEST_ACTUAL_ANNUAL_DATA"))
        if latest_actual is None:
            continue
        for index, year in years:
            if year > int(latest_actual):
                continue
            value = numeric(values[index])
            if value is None:
                continue
            period = str(year)
            observations.append(make_row(
                country_code=country, indicator_code=code,
                source_series=source_indicator, source_key=source_indicator,
                topic=topic, frequency="A", period=period, observation_date=period_date(period, "A"),
                value=format(value, ".15g"), unit=unit, observation_status="actual",
                source_id="imf_weo", source_url=url, source_vintage="April 2026",
                retrieved_at=retrieved_at,
            ))
    return observations, list(definitions_by_code.values())


def build_coverage(observations: list[dict], definitions: list[dict], countries: list[str]) -> list[dict]:
    grouped = defaultdict(list)
    for row in observations:
        grouped[(row["country_code"], row["indicator_code"], row["frequency"], row["source_id"])].append(row)
    coverage = []
    for (country, indicator, frequency, source_id), rows in sorted(grouped.items()):
        periods = sorted(row["period"] for row in rows)
        coverage.append({
            "country_code": country, "indicator_code": indicator, "frequency": frequency,
            "source_id": source_id, "first_period": periods[0], "last_period": periods[-1],
            "observation_count": len(rows), "series_count": len(set(row["source_key"] for row in rows)),
        })
    return coverage


def select_report_series(observations: list[dict], benchmark_countries: set[str]) -> list[dict]:
    """Small, unambiguous series for browser charts; the fact file retains everything."""
    selected = []
    for row in observations:
        if row["country_code"] not in benchmark_countries:
            continue
        if row["source_id"] == "imf_weo" and row["indicator_code"] in {
            "real_gdp_growth_pct", "inflation_pct", "unemployment_pct", "balance_pct_gdp", "gross_debt_pct_gdp",
        }:
            selected.append(row)
        elif row["source_id"] == "world_bank" and row["indicator_code"] in {
            "labour_force_participation", "employment_to_population", "gross_fixed_capital_formation",
            "current_account_balance", "private_credit", "gini_index",
        }:
            selected.append(row)
        elif row["source_id"] == "oecd_kei":
            measure = row["indicator_code"].replace("oecd_kei_", "").upper()
            wanted = (
                (measure == "B1GQ_Q" and row["frequency"] == "Q" and row["unit"] == "GR" and row["transformation"] == "G1") or
                (measure == "UNEMP" and row["frequency"] == "M" and row["unit"] == "PT_LF") or
                (measure == "CP" and row["frequency"] == "M" and row["unit"] == "GR" and row["transformation"] == "GY") or
                (measure == "TOVM" and row["frequency"] == "M" and row["unit"] == "IX") or
                (measure == "IRSTCI" and row["frequency"] == "M" and row["unit"] == "PA") or
                (measure == "IRLT" and row["frequency"] == "M" and row["unit"] == "PA") or
                (measure == "BCICP" and row["frequency"] == "M") or
                (measure == "CCICP" and row["frequency"] == "M")
            )
            if wanted and row["seasonal_adjustment"] in {"Y", "_Z"}:
                selected.append(row)
        elif row["source_id"] == "bis" and row["indicator_code"] in {
            "central_bank_policy_rate", "residential_property_prices", "debt_service_ratio",
        }:
            selected.append(row)
    return selected


def write_outputs(contract: dict, observations: list[dict], definitions: list[dict], countries: list[dict], raw_dir: Path, retrieved_at: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    observations.sort(key=lambda row: (row["country_code"], row["indicator_code"], row["source_key"], row["period"]))
    fact_path = OUTPUT_DIR / "economic-observations.v1.csv.gz"
    with gzip.open(fact_path, "wt", newline="", encoding="utf-8", compresslevel=9) as handle:
        writer = csv.DictWriter(handle, fieldnames=FACT_FIELDS)
        writer.writeheader()
        writer.writerows(observations)
    unique_definitions = {}
    for definition in definitions:
        unique_definitions[(definition["source_id"], definition["indicator_code"])] = definition
    definitions = sorted(unique_definitions.values(), key=lambda item: (item["topic"], item["indicator_code"], item["source_id"]))
    coverage = build_coverage(observations, definitions, contract["scope"]["countries"])
    (OUTPUT_DIR / "economic-indicators.v1.json").write_text(json.dumps({
        "schema_version": "1.0.0", "generated_at": retrieved_at,
        "countries": countries,
        "indicators": definitions, "sources": contract["sources"],
    }, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (OUTPUT_DIR / "economic-coverage.v1.json").write_text(json.dumps({
        "schema_version": "1.0.0", "generated_at": retrieved_at, "coverage": coverage,
    }, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    benchmark = set(contract["scope"]["benchmark_countries"])
    report_rows = select_report_series(observations, benchmark)
    compact_series = []
    grouped_series = defaultdict(list)
    for row in report_rows:
        grouped_series[(
            row["country_code"], row["indicator_code"], row["source_key"], row["source_series"],
            row["topic"], row["frequency"], row["unit"], row["seasonal_adjustment"],
            row["transformation"], row["source_id"], row["source_url"], row["source_vintage"],
        )].append((row["period"], float(row["value"]), row["observation_status"]))
    for key, values in sorted(grouped_series.items()):
        (country, indicator, source_key, source_series, topic, frequency, unit,
         adjustment, transformation, source_id, source_url, vintage) = key
        compact_series.append({
            "country_code": country, "indicator_code": indicator, "source_key": source_key,
            "source_series": source_series, "topic": topic, "frequency": frequency, "unit": unit,
            "seasonal_adjustment": adjustment, "transformation": transformation,
            "source_id": source_id, "source_url": source_url, "source_vintage": vintage,
            "values": [[period, value, status] for period, value, status in sorted(values)],
        })
    report_payload = {
        "schema_version": "1.0.0", "generated_at": retrieved_at,
        "methodology": {
            "annual_recommendation": "Use 30 years where available; 40-60 years for structural relationships and regime checks.",
            "high_frequency_recommendation": "Use at least 80 quarters or 120 monthly observations; 2000 onward is the practical comparison window.",
            "warning": "Observational macroeconomic series support description and association, not causal claims without an identification design."
        },
        "countries": [country for country in countries if country["code"] in benchmark],
        "global_summary": {
            "country_count": len(countries), "observation_count": len(observations),
            "indicator_definition_count": len(definitions),
            "source_counts": dict(sorted(Counter(row["source_id"] for row in observations).items())),
            "frequency_counts": dict(sorted(Counter(row["frequency"] for row in observations).items())),
        },
        "definitions": definitions,
        "coverage": coverage,
        "series": compact_series,
    }
    (OUTPUT_DIR / "economy-deep-dive.v1.json").write_text(json.dumps(report_payload, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    manifest = {
        "schema_version": "1.0.0", "generated_at": retrieved_at,
        "raw_snapshot": str(raw_dir.relative_to(WORKSPACE)),
        "raw_assets": [{"path": str(path.relative_to(WORKSPACE)), "bytes": path.stat().st_size, "sha256": sha256(path)} for path in sorted(raw_dir.rglob("*")) if path.is_file()],
        "outputs": [
            {"path": str(path.relative_to(ROOT)), "bytes": path.stat().st_size, "sha256": sha256(path)}
            for path in sorted(OUTPUT_DIR.glob("*")) if path.is_file() and path.name != "manifest.v1.json"
        ],
        "row_count": len(observations), "report_row_count": len(report_rows),
        "source_counts": dict(sorted(Counter(row["source_id"] for row in observations).items())),
        "country_counts": dict(sorted(Counter(row["country_code"] for row in observations).items())),
    }
    (OUTPUT_DIR / "manifest.v1.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--refresh", action="store_true", help="redownload an existing snapshot")
    parser.add_argument("--snapshot-date", help="YYYY-MM-DD; defaults to current UTC date")
    args = parser.parse_args()
    retrieved_at = now_iso()
    snapshot_date = args.snapshot_date or retrieved_at[:10]
    contract = json.loads(CONTRACT_PATH.read_text(encoding="utf-8"))
    raw_dir = RAW_ROOT / snapshot_date
    raw_dir.mkdir(parents=True, exist_ok=True)
    countries = fetch_country_dimension(raw_dir, args.refresh)
    contract["scope"]["countries"] = [country["code"] for country in countries]
    observations: list[dict] = []
    definitions: list[dict] = []
    for loader in (fetch_world_bank, fetch_oecd, fetch_bis):
        rows, items = loader(contract, raw_dir, retrieved_at, args.refresh)
        observations.extend(rows)
        definitions.extend(items)
        print(f"{loader.__name__}: {len(rows):,} observations", file=sys.stderr)
    rows, items = load_imf(contract, retrieved_at)
    observations.extend(rows)
    definitions.extend(items)
    print(f"load_imf: {len(rows):,} observations", file=sys.stderr)
    write_outputs(contract, observations, definitions, countries, raw_dir, retrieved_at)
    print(f"total: {len(observations):,} observations -> {OUTPUT_DIR}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
